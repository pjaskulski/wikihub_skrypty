""" wybór z wyników """

from pathlib import Path
import openpyxl


input_path = Path('..') / 'data' / 'dopasowanie_prng_v5.xlsx'
output_path = Path('..') / 'data' / 'dopasowanie_prng_v5x.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['dopasowanie_prng_v5']

col_names = {}
nr_col = 0
for column in ws.iter_cols(1, ws.max_column):
    col_names[column[1].value] = nr_col
    nr_col += 1

rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row))

for row in reversed(rows):
    nazwa_wspolczesna = row[col_names['nazwa_wspolczesna']].value
    nazwa16w = row[col_names['nazwa16w']].value
    nazwa_slow = row[col_names['nazwa_slow']].value
    nowa_nazwa_prng = row[col_names['Nowa_nazwa_PRNG']].value
    if nowa_nazwa_prng not in [nazwa_wspolczesna, nazwa16w, nazwa_slow]:
        ws.delete_rows(row[0].row, 1)

wb.save(output_path)

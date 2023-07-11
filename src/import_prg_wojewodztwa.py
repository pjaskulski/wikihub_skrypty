""" import województw z pliku wojewodztwa.xlsx z danymi z PRG"""
import os
import sys
import time
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import element_search_adv
from wikidariahtools import get_properties, get_elements
from property_import import create_statement_data


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

# czy zapis w wikibase czy tylko test
WIKIBASE_WRITE = False

# standardowe właściwości
properties = get_properties(['instance of', 'stated as', 'reference URL', 'retrieved',
                            'id SDI', 'part of', 'has part or parts', 'TERYT',
                            'stated in', 'point in time'])

# elementy definicyjne (purl to voivodship (The Republic of Poland (1999-2016))
elements = get_elements(['administrative unit', 'onto.kul.pl/ontohgis/administrative_type_47'])

# wspólna referencja dla wszystkich deklaracji z PRG
references = {}
references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGIK/PRG/WFS/AdministrativeBoundaries'
references[properties['retrieved']] = '2022-09-05'

# wspólna referencja do onto (poprawić Q w docelowej!)
onto_references = {}
onto_references[properties['stated in']] = 'Q233549'

# kwalifikator z punktem czasowym
qualifiers = {}
qualifiers[properties['point in time']] = '+2022-00-00T00:00:00Z/9' # rok 2022


def get_label_en(qid: str) -> str:
    """ zwraca angielską etykietę dla podanego QID """

    wb_item_test = wbi_core.ItemEngine(item_id=qid)
    result = wb_item_test.get_label('en')

    return result


# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':

    # logowanie do instancji wikibase
    login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                     consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                     access_token=WIKIDARIAH_ACCESS_TOKEN,
                                     access_secret=WIKIDARIAH_ACCESS_SECRET,
                                     token_renew_period=14400)

    xlsx_input = '../data_prng/wojewodztwa.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["wojewodztwa"]

    # nazwy kolumn z xlsx
    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1
        # wczytanie danych z xlsx
        nazwa = row[col_names['JPT_NAZWA_']].value
        if not nazwa:
            continue

        label_pl = 'województwo' + ' ' + nazwa
        label_en = 'voivodship' + ' ' + nazwa

        teryt = row[col_names['JPT_KOD_JE']].value
        idiip = row[col_names['IIP_IDENTY']].value

        description_pl = 'województwo - współczesna jednostka administracyjna według Państwowego Rejestru Granic (PRG)'
        description_en = 'voivodship - a modern administrative unit according to the National Register of Boundaries (PRG)'

        # przygotowanie struktur wikibase
        data = []
        aliasy = []

        # instance of
        statement = create_statement_data(properties['instance of'],
                                          elements['onto.kul.pl/ontohgis/administrative_type_47'],
                                          None, None, add_ref_dict=onto_references)
        if statement:
            data.append(statement)

        # id SDI
        if idiip:
            statement = create_statement_data(prop=properties['id SDI'],
                                              value=idiip,
                                              reference_dict=None,
                                              qualifier_dict=qualifiers,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)

        # TERYT
        if teryt:
            statement = create_statement_data(properties['TERYT'],
                                              teryt,
                                              None,
                                              None,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)

        # JPT_NAZWA_
        aliasy.append(nazwa)
        statement = create_statement_data(properties['stated as'], f'pl:"{nazwa}"', None, None, add_ref_dict=references)
        if statement:
            data.append(statement)

        # etykiety, description, aliasy
        wb_item = wbi_core.ItemEngine(new_item=True, data=data)
        wb_item.set_label(label_en, lang='en')
        wb_item.set_label(label_pl,lang='pl')

        # description
        wb_item.set_description(description_en, 'en')
        wb_item.set_description(description_pl, 'pl')

        if aliasy:
            for value_alias in aliasy:
                wb_item.set_aliases(value_alias, 'pl')

        # wyszukiwanie po etykiecie, właściwości instance of oraz po opisie
        parameters = [(properties['instance of'], elements['onto.kul.pl/ontohgis/administrative_type_47'])]
        ok, item_id = element_search_adv(label_en, 'en', parameters, description_en)

        if not ok:
            if WIKIBASE_WRITE:
                test = 1
                while True:
                    try:
                        new_id = wb_item.write(login_instance, bot_account=True, entity_type='item')
                        print(f'{index}/{max_row - 1} Dodano nowy element: {label_en} / {label_pl} = {new_id}')
                        break
                    except MWApiError as wb_error:
                        err_code = wb_error.error_msg['error']['code']
                        message = wb_error.error_msg['error']['info']
                        print(f'ERROR: {err_code}, {message}')
                        # jeżeli jest to problem z tokenem to próba odświeżenia tokena i powtórzenie
                        # zapisu, ale tylko raz, w razie powtórnego błędu bad token, skrypt kończy pracę
                        if err_code in ['assertuserfailed', 'badtoken']:
                            if test == 1:
                                print('Generate edit credentials...')
                                login_instance.generate_edit_credentials()
                                test += 1
                                continue
                        sys.exit(1)
            else:
                new_id = 'TEST'
                print(f"{index}/{max_row - 1} Przygotowano dodanie elementu - {label_en} / {label_pl}  = {new_id}")
        else:
            print(f'{index}/{max_row - 1} Element: {label_en} / {label_pl} już istnieje: {item_id}')

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

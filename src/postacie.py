""" skrypt generujący quickstatements dla postaci z PSB na podstawie
    indeksu Bożeny Bigaj
"""

import sys
import re
import pickle
import os
from pathlib import Path
from time import sleep
from urllib.parse import quote
import requests
import roman as romenum
from openpyxl import load_workbook
from wikibaseintegrator.wbi_config import config as wbi_config
from wikidariahtools import format_date, text_clear, \
                            get_last_nawias, short_names_in_autor
from postacietools import get_name
from wikidariahtools import element_search, gender_detector
from wyjatki_postacie import ETYKIETY_WYJATKI


# adresy
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# stałe
P_INSTANCE_OF = 'P47'
Q_HUMAN = 'Q32'
P_IMIE = 'P184'
P_NAZWISKO = 'P183'
P_VIAF = 'P79'
P_REFERENCE_URL = 'S182'
P_DATE_OF_BIRTH = 'P7'
P_DATE_OF_DEATH = 'P8'
P_DESCRIBED_BY_SOURCE = 'P17'
P_BIRTH_NAME = 'P63'
P_EARLIEST_DATE = 'P38'
P_LATEST_DATE = 'P39'
P_FLUORIT = 'P54'
Q_CIRCA = 'Q37979'
P_SOURCING_CIRCUMSTANCES = 'P189'

BIOGRAMY = {}
VIAF_ID = {}
VIAF_BIRTH = {}
VIAF_DEATH = {}
WERYFIKACJA_VIAF = {}
WYJATKI = {}

MALE_FEMALE_NAME = ['Maria', 'Anna']

# czy wczytywanie i zapisywanie słowników z/do pickle
LOAD_DICT = True
SAVE_DICT = False


def double_space(value:str) -> str:
    """ usuwa podwójne spacje """
    return ' '.join(value.strip().split())


def postac_etykieta(imie_1:str, imie_2:str, imie_3:str, imie_4:str, 
                    nazwisko_1:str, nazwisko_2:str):
    """ ustala etykietę dla postaci (imiona nazwiska) """
    result = f'{imie_1} {imie_2} {imie_3} {imie_4} {nazwisko_2} {nazwisko_1}'
    result = result.strip()
    return double_space(result)


def load_wyjatki(path: str) -> dict:
    """ load wyjatki"""
    result = {}

    try:
        work_book = load_workbook(path)
    except IOError:
        print(f"ERROR. Can't open and process file: {path}")
        sys.exit(1)

    sheet = work_book['Arkusz1']
    columns = {'POSTAĆ':0, 'VIAF':1}
    for current_row in sheet.iter_rows(2, sheet.max_row):
        u_osoba = current_row[columns['POSTAĆ']].value
        u_viaf = current_row[columns['VIAF']].value
        if u_osoba:
            u_osoba = u_osoba.strip()
        if u_viaf:
            u_viaf = u_viaf.strip()
        if u_osoba and u_viaf:
            result[u_osoba.strip()] = u_viaf.strip()
        else:
            break

    return result


def get_viaf_data(v_url: str) -> tuple:
    """ get_viaf_data """
    v_id = v_birth = v_death = ''
    response = requests.get(v_url + 'viaf.json')
    result = response.json()
    if 'viafID' in result:
        v_id = result['viafID']
    if 'birthDate' in result:
        v_birth = result['birthDate']
    if 'deathDate' in result:
        v_death = result['deathDate']

    return v_id, v_birth, v_death


def viaf_search(person_name: str, s_birth: str = '', s_death: str = '',
                offline: bool = False) -> tuple:
    """ szukanie identyfikatora VIAF """

    info = id_url = birthDate = deathDate = ''
    result = False

    # jeżeli osoba jest w wyjątkach to pobieramy dane ze znanego adresu
    # lub od razu NOT FOUND
    if name in WYJATKI:
        if WYJATKI[name].strip() == 'BRAK':
            return False, "NOT_FOUND", '', '', ''

        info, birthDate, deathDate = get_viaf_data(WYJATKI[name])
        id_url = WYJATKI[name]
        VIAF_ID[name] = info
        if birthDate:
            VIAF_BIRTH[name] = birthDate
        if deathDate:
            VIAF_DEATH[name] = deathDate

        return True, info, id_url, birthDate, deathDate

    # jeżeli identyfikator jest już znany to nie ma potrzeby szukania
    # przez api
    if person_name in VIAF_ID:
        info = VIAF_ID[person_name]
        if 'http' in info:
            match = re.search(r'\d{3,25}', info)
            if match:
                info = match.group()
        id_url = f"http://viaf.org/viaf/{info}/"

        if person_name in VIAF_BIRTH:
            birthDate = VIAF_BIRTH[person_name]

        if person_name in VIAF_DEATH:
            deathDate = VIAF_DEATH[person_name]

        return True, info, id_url, birthDate, deathDate

    # jeżeli nie chcemy wyszukiwać online w viaf.org
    if offline:
        return False, "NOT FOUND", '', '', ''

    identyfikatory = []
    urls = {}
    base = 'https://viaf.org/viaf/search'
    format_type = 'application/json'
    search_person = quote(f'"{person_name}"')
    adres = f'{base}?query=local.personalNames+=+{search_person}'
    adres+= f'&local.sources+=+"plwabn"&sortKeys=holdingscount&httpAccept={format_type}'

    # mały odstęp między poszukiwaniami
    sleep(0.03)

    try:
        response = requests.get(adres)
        result = response.json()
        if 'records' in result['searchRetrieveResponse']:
            rekordy = result['searchRetrieveResponse']['records']

            for rekord in rekordy:
                v_id = rekord['record']['recordData']['viafID']
                if v_id:
                    url = rekord['record']['recordData']['Document']['@about']
                    if isinstance(rekord['record']['recordData']['mainHeadings']['data'], list):
                        label = rekord['record']['recordData']['mainHeadings']['data'][0]['text']
                    elif isinstance(rekord['record']['recordData']['mainHeadings']['data'], dict):
                        label = rekord['record']['recordData']['mainHeadings']['data']['text']

                    if label:
                        label = label.replace(",", "")
                        l_name = person_name.split(" ")
                        find_items = True

                        for item_name in l_name:
                            if len(item_name) > 2 and not item_name in label:
                                find_items = False
                                break

                        if find_items:
                            if 'birthDate' in rekord['record']['recordData']:
                                birthDate = rekord['record']['recordData']['birthDate']

                            if 'deathDate' in rekord['record']['recordData']:
                                deathDate = rekord['record']['recordData']['deathDate']

                            # jeżeli mamy podane daty w viaf i w indeksie to mogą się różnić
                            # o maksymalnie 3 lata
                            if len(birthDate) == 10:
                                int_birth_date = int(birthDate[:4])
                            elif len(birthDate) == 9:
                                int_birth_date = int(birthDate[:3])
                            elif len(birthDate) == 4 and birthDate.isnumeric():
                                int_birth_date = int(birthDate)
                            elif len(birthDate) == 3 and birthDate.isnumeric():
                                int_birth_date = int(birthDate)
                            else:
                                int_birth_date = -1

                            if len(deathDate) == 10:
                                int_death_date = int(deathDate[:4])
                            elif len(deathDate) == 9:
                                int_death_date = int(deathDate[:3])
                            elif len(deathDate) == 4 and deathDate.isnumeric():
                                int_death_date = int(deathDate)
                            elif len(deathDate) == 3 and deathDate.isnumeric():
                                int_death_date = int(deathDate)
                            else:
                                int_death_date = -1

                            if s_birth and int_birth_date > 0:
                                y_diff = abs(int(s_birth) - int_birth_date)
                                if not birthDate.startswith(s_birth) and y_diff > 3:
                                    continue

                            if s_death and int_death_date > 0:
                                y_diff = abs(int(s_death) - int_death_date)
                                if not deathDate.startswith(s_death) and y_diff > 3:
                                    continue

                            identyfikatory.append(v_id)
                            urls[v_id] = url

                            VIAF_ID[person_name] = v_id  # zapis identyfikatora w słowniku

                            if birthDate:
                                VIAF_BIRTH[person_name] = birthDate
                            if deathDate:
                                VIAF_DEATH[person_name] = deathDate

                            break

    except requests.exceptions.RequestException as e_info:
        print(f'Name: {name} ERROR {e_info}')

    if len(identyfikatory) == 1:
        return True, identyfikatory[0], urls[identyfikatory[0]], birthDate, deathDate

    return False, "NOT FOUND", '', '', ''


def ustal_etykiete(value: str, title_value: str) -> str:
    """ ustala etykiete biogramu """
    l_nawias = value.split(",")
    if len(l_nawias) != 4:
        print(f'ERROR: {l_nawias}')
        sys.exit(1)
    autor = text_clear(l_nawias[0])
    autor_in_title = short_names_in_autor(autor)
    tom = text_clear(l_nawias[1])
    tom = tom.replace("t.","").strip()
    strony = text_clear(l_nawias[3])

    if ";" in autor_in_title:
        autor_in_title = autor_in_title.replace(';', ',')

    return f"{autor_in_title}, {title_value}, w: PSB {tom}, {strony}"


def roman_numeric(value:str) -> bool:
    """ czy liczba rzymska?"""
    pattern = r'[IVX]{1,5}\s+w\.{0,1}'
    match = re.search(pattern, value)
    if not match:
        pattern = r'[IVX]{1,5}'
        match = re.search(pattern, value)

    return bool(match)


def date_kwal(value: str, typ:str = '') -> tuple:
    """ data kwalifikator? dla dat niepwenych opisanych tekstem oprócz
        podania samej daty np. 'zm. ok. 1459'
        zwraca typ daty: 'B' data urodzenia 'D' data śmierci
        i rodzaj niepewności: 'before', 'after' , 'about', 'or',
        'between', 'roman', 'turn'
    """
    if value.strip().isnumeric():
        return '', 'certain'

    result = result_type = ''
    if 'zm. przed' in value:    # jeszcze nie obsługiwane
        result = 'before'
        result_type = 'B'
    elif 'zm. po' in value:     # jeszcze nie obsługiwane
        result = 'after'
        result_type = 'D'
    elif 'zm. prawdopodobnie' in value: # jeszcze nie obsługiwane
        result = 'about'
        result_type = 'D'
    elif 'zm. ok.' in value:
        result = 'about'
        result_type = 'D'
    elif 'zm. między' in value:
        result = 'between'
        result_type = 'D'
    elif 'zm.' in value and 'lub nieco później' in value:
        result = 'about'
        result_type = 'D'
    elif 'zm. w lub po' in value:
        result = 'after'
        result_type = 'D'
    elif 'zm.' in value and 'lub' in value:
        result = 'or'
        result_type = 'D'
    elif 'zm. w/przed' in value:
        result = 'before'
        result_type = 'D'
    elif 'ur. ok.' in value:
        result = 'about'
        result_type = 'B'
    elif 'ur. między' in value:
        result = 'between'
        result_type = 'B'

    elif 'przed lub w' in value: # wpis z rokiem i kwalifikatorem latest date jednocześnie?
        result = 'before'
    elif 'w lub po' in value: # wpis z rokiem i kwalifikatorem earliest date jednocześnie?
        result = 'after'
    elif 'po lub w' in value: # wpis z rokiem i kwalifikatorem earliest date jednocześnie?
        result = 'after'
    elif 'między' in value and '?' in value:
        result = 'between?'
    elif 'między' in value:
        result = 'between'
    elif 'miedzy' in value:
        result = 'between'
    elif 'ok.' in value:
        result = 'about'
    elif 'około' in value:
        result = 'about'
    elif 'prawdopodobnie' in value:
        result = 'about'
    elif 'zapewne' in value:
        result = 'about'
    elif 'lub' in value:
        result = 'or'
    elif 'przed' in value:
        result = 'before'
    elif 'po ' in value:
        result = 'after'
    elif '?' in value:
        result = 'about'
    elif 'nie później niż' in value:
        result = 'before'
    elif 'najpóźniej' in value:
        result = 'before'

    elif 'zm.' in value: # data pewna tylko z określeniem tekstowym zm.
        result_type = 'D'
        result = 'certain'
    elif 'um.' in value: # data pewna tylko z określeniem tesktowym um.
        result_type = 'D'
        result = 'certain'
    elif 'ur.' in value: # data pewna tylko z określeniem tesktowym ur.
        result_type = 'B'
        result = 'certain'
    elif roman_numeric(value):
        result = 'roman'

    match = re.search(r'\d{3,4}/\d', value)
    if match:
        result = 'turn' # przełom lat

    if result_type == '':
        result_type = typ

    # test czy obsłużono wszystkie przypadki
    #if result == '':
    #    print(value)

    return result_type, result


def get_date(value: str, typ = '') -> tuple:
    """ get date """
    date_of = dod_info = ''
    pattern = r'\d{3,4}'
    matches = [x.group() for x in re.finditer(pattern, value)]
    if len(matches) == 1:
        date_of = matches[0]
    elif len(matches) > 1:
        date_of = '|'.join(matches)
    typ, dod_info = date_kwal(value, typ)
    if dod_info == 'roman':
        matches = [x.group() for x in re.finditer(r'[IVX]{1,5}', value)]
        if len(matches) == 1:
            date_of = str(romenum.fromRoman(matches[0]))
        else:
            matches = [str(romenum.fromRoman(x)) for x in matches]
            date_of = '|'.join(matches)
    elif dod_info == 'turn':
        match = re.search(r'\d{3,4}/\d', value)
        if match:
            v_list = match.group().split('/')
            v_list1 = v_list[0].strip()
            v_list2 = v_list1[:len(v_list1)-1] + v_list[1].strip()
            date_of = f'{v_list1}|{v_list2}'

    return date_of, typ, dod_info


def date_birth_death(value: str) -> tuple:
    """ dateBirthDeath """
    date_of_birth = date_of_death = dod_b = dod_d = ""
    # if '1. poł. XIII w.' in value:
    #     print()

    if value == "":
        return date_of_birth, date_of_death, dod_b, dod_d

    # jeżeli dwie daty (urodzin i śmierci)
    if ',' in value:
        separator = ','
    else:
        separator = '-'

    # jeżeli zakres dat
    if separator in value and not roman_numeric(value):
        tmp = value.split(separator)
        date_of_birth, typ, dod_b = get_date(tmp[0].strip(), 'B')
        date_of_death, typ, dod_d = get_date(tmp[1].strip(), 'D')
    # jeżeli tylko jedna z dat lub ogólny opis np. XVII wiek
    else:
        date_of_one, typ, dod_inf = get_date(value, '')
        if typ == 'B':
            date_of_birth = date_of_one
            dod_b = dod_inf
        elif typ == 'D':
            date_of_death = date_of_one
            dod_d = dod_inf
        elif typ == '' and dod_inf == 'roman':
            date_of_birth = date_of_one # dla określeń typu 'XV w.' lub 'XV/XVI w.'
            dod_b = dod_inf

    return date_of_birth, date_of_death, dod_b, dod_d


if __name__ == "__main__":
    file_path = Path('.').parent / 'data/lista_hasel_PSB_2020.txt'
    uzup_path = Path('.').parent / 'data/postacie_viaf_uzup.xlsx'
    output = Path('.').parent / 'out/postacie.qs'
    output_daty = Path('.').parent / 'out/postacie_daty.qs'
    postacie_pickle = Path('.').parent / 'out/postacie.pickle'
    postacie_birth_pickle = Path('.').parent / 'out/postacie_birth.pickle'
    postacie_death_pickle = Path('.').parent / 'out/postacie_death.pickle'
    biogramy_pickle = Path('.').parent / 'out/biogramy.pickle'
    postacie_viaf_html = Path('.').parent / 'out/postacie_viaf.html'

    # odmrażanie słowników
    if LOAD_DICT:
        if os.path.isfile(postacie_pickle):
            with open(postacie_pickle, 'rb') as handle:
                VIAF_ID = pickle.load(handle)

        if os.path.isfile(biogramy_pickle):
            with open(biogramy_pickle, 'rb') as handle:
                BIOGRAMY = pickle.load(handle)

        if os.path.isfile(postacie_birth_pickle):
            with open(postacie_birth_pickle, 'rb') as handle:
                VIAF_BIRTH = pickle.load(handle)

        if os.path.isfile(postacie_death_pickle):
            with open(postacie_death_pickle, 'rb') as handle:
                VIAF_DEATH = pickle.load(handle)

    with open(file_path, "r", encoding='utf-8') as f:
        indeks = f.readlines()

    if not indeks:
        print('ERROR: empty index')
        sys.exit(1)

    load_wyjatki(uzup_path)

    with open(output, "w", encoding='utf-8') as f, open(output_daty, "w", encoding='utf-8') as fd:
        licznik = 0
        for line in indeks:
            licznik += 1
            nawias, title_stop = get_last_nawias(line)
            title = line[:title_stop].strip()
            # etykieta biogramu do wyszukania w wikibase
            etykieta = ustal_etykiete(nawias, title)

            isYears = title.count('(') == 1 and title.count(')') == 1
            isAlias = title.count('(') == 2 and title.count(')') == 2

            name = years = author = psb = dateOfBirth = dateOfDeath = ''
            imie = imie2 = imie3 = imie4 = nazwisko = nazwisko2 = ''
            stop = 0

            # lata życia
            # jeżeli we wpisie jest alias, przydomek itp. to daty są w drugim nawiasie
            start = title.find('(')
            name = title[:start].strip()
            nazwisko, imie, imie2, nazwisko2, imie3, imie4 = get_name(name)

            if isAlias:
                #print("ALIAS: ", title)
                start = title.find('(', start + 1)

            if isYears:
                stop = title.find(')', start)
                years = title[start + 1: stop].strip()
                years = years.replace('–', '-')

            #ok, q_biogram = element_search(etykieta, 'item', 'pl')
            ok = False
            if not ok:
                q_biogram = '{Q:biogram}'
            else:
                BIOGRAMY[name] = q_biogram

            # daty urodzenia i śmierci
            dateB, dateD, dateB_dod, dateD_dod = date_birth_death(years)

            #print(dateB, dateD, dateB_dod, dateD_dod, name)

            # if "Stefan I" in name:
            #     print()

            # jeżeli znamy tylko imię postaci odpytywanie VIAF nie ma sensu (?)
            viaf_ok = False
            viaf_id = viaf_url = viaf_date_b = viaf_date_d = ''
            if ' ' in name:
                viaf_ok, viaf_id, viaf_url, viaf_date_b, viaf_date_d = viaf_search(name,
                                                                                   s_birth=dateB,
                                                                                   s_death=dateD,
                                                                                   offline=True)
            # jeżeli VIAF ma daty to zakładam że są lepsze niż w PSB
            # tylko jeżeli dat nie ma w VIAF lub są identyczne jak w PSB
            # to obsługa dat z PSB, z niepewnościami włącznie typu:
            # 'before', 'after' , 'about', 'or', 'between', 'roman', 'turn'
            dateB_1 = dateB_2 = dateD_1 = dateD_2 = ''
            if viaf_ok and viaf_date_b and viaf_date_b != dateB:
                dateB = viaf_date_b
            else:
                if dateB_dod == 'between':
                    t_date_b = dateB.split('|')
                    dateB_1 = format_date(t_date_b[0])
                    dateB_2 = format_date(t_date_b[1])
                    dateB = 'somevalue'
                elif dateB_dod == 'or':
                    t_date_b = dateB.split('|')
                    dateB_1 = format_date(t_date_b[0])
                    dateB_2 = format_date(t_date_b[1])
                elif dateB_dod in ('before', 'after'):
                    dateB_1 = dateB
                    dateB = 'somevalue'
                elif dateB_dod == 'turn':
                    t_date_b = dateB.split('|')
                    dateB_1 = format_date(t_date_b[0])
                    dateB_2 = format_date(t_date_b[1])
                elif dateB_dod == 'roman': # fluorit!
                    t_date_b = dateB.split('|')
                    dateB_1 = format_date(t_date_b[0])
                    if len(t_date_b) == 2:
                        dateB_2 = format_date(t_date_b[1])

            if viaf_ok and viaf_date_d and viaf_date_d != dateD:
                dateD = viaf_date_d
            else:
                if dateD_dod == 'between':
                    t_date_d = dateD.split('|')
                    dateD_1 = format_date(t_date_d[0])
                    dateD_2 = format_date(t_date_d[1])
                    dateD = 'somevalue'
                elif dateD_dod == 'or':
                    t_date_d = dateD.split('|')
                    dateD_1 = format_date(t_date_d[0])
                    dateD_2 = format_date(t_date_d[1])
                elif dateD_dod in ('before', 'after'):
                    dateD_1 = dateD
                    dateD = 'somevalue'
                elif dateD_dod == 'turn':
                    t_date_d = dateD.split('|')
                    dateD_1 = format_date(t_date_d[0])
                    dateD_2 = format_date(t_date_d[1])

            if dateB and dateB != 'somevalue':
                dateB = format_date(dateB)

            if dateD and dateD != 'somevalue':
                dateD = format_date(dateD)

            # konstrukcja etykiety z uwzględnieniem przestawienia kolejności
            # imienia i nazwiska, imion zakonnych itp.
            name_etykieta = postac_etykieta(imie, imie2, imie3, imie4, nazwisko, nazwisko2)
            if ' zwany ' in name:
                t_mark = ' zwany '
            elif ' zw. ' in name:
                t_mark = ' zw. '
            elif ' z ' in name:
                t_mark = ' z '
            elif ' ze ' in name:
                t_mark = ' ze '
            elif ' h.' in name:
                t_mark = ' h.'
            else:
                t_mark = ''

            if t_mark:
                pos = name.find(t_mark)
                toponimik = name[pos:]
                name_etykieta = name_etykieta + toponimik

            if 'starszy' in name and 'starszy' not in name_etykieta:
                name_etykieta += ' ' + 'starszy'
            if 'Starszy' in name and 'Starszy' not in name_etykieta:
                name_etykieta += ' ' + 'Starszy'
            if 'młodszy' in name and 'młodszy' not in name_etykieta:
                name_etykieta += ' ' + 'młodszy'
            if 'Młodszy' in name and 'Młodszy' not in name_etykieta:
                name_etykieta += ' ' + 'Młodszy'
            if 'junior' in name and 'junior' not in name_etykieta:
                name_etykieta += ' ' + 'junior'
            if 'senior' in name and 'senior' not in name_etykieta:
                name_etykieta += ' ' + 'senior'

            # imona zakonne
            if ' w zak. ' in name:
                z_mark = ' w zak.'
            elif ' w zakonie ' in name:
                z_mark = ' w zakonie '
            elif ' w zak ' in name:
                z_mark = ' w zak '
            elif ' zak. ' in name:
                z_mark = ' zak. '
            else:
                z_mark = ''

            birth_name = ''
            if z_mark:
                pos = name.find(z_mark)
                pos2 = pos + len(z_mark)
                birth_name = name_etykieta
                zakonimik = name[pos:]
                zakon_names = name[pos2:].strip()
                name_etykieta = f'{zakon_names} {nazwisko2} {nazwisko}'.strip()

            name_etykieta = double_space(name_etykieta)

            # dla postaci typu 'Szneur Zalman ben Baruch' na razie tak jak w oryginale
            if ' ben ' in name:
                name_etykieta = name
            if ' Ben ' in name:
                name_etykieta = name

            # dla postaci typu 'Salomon syn Joela' na razie tak jak w oryginale
            if ' syn ' in name:
                name_etykieta = name

            # dla władców tak jak w oryginale
            is_king = False
            roman = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 
             'XI', 'XII', 'XIII', 'XIV', 'XV', 'XVI', 'XVII', 'XVIII', 'XIX', 'XX']
            for item in roman:
                if f' {item} ' in name or name.endswith(f' {item}'):
                    is_king = True
                    break
            if is_king:
                name_etykieta = name

            # konstrukcja typu Mikołaj z Jaroszowa zw. Kornicz lub Siestrzeniec
            # zostaje bez zmian w etykiecie
            if ' z ' in name and ' zw. ' in name:
                name_etykieta = name
            if ' z ' in name and ' zwany ' in name:
                name_etykieta = name

            # jeżeli nie było rozpoznanego imienia tylko przydomek/przezwisko
            if name_etykieta.startswith('z '):
                name_etykieta = name

            # super wyjątki nie podpadające gdzie indziej
            if name in ETYKIETY_WYJATKI:
                name_etykieta = ETYKIETY_WYJATKI[name]

            #if len(name) != len(name_etykieta):
            # if z_mark:
            #     print(name_etykieta, '=', birth_name)

            years = f'({years})'

            #else:
                #print(name, '=', name_etykieta)
            #continue # tymczasowo

            # zapis quickstatements
            f.write('CREATE\n')
            f.write(f'LAST\tLpl\t"{name_etykieta}"\n')
            f.write(f'LAST\tLen\t"{name_etykieta}"\n')
            if years:
                f.write(f'LAST\tDpl\t"{years}"\n')
                f.write(f'LAST\tDen\t"{years}"\n')
            if imie:
                gender1 = gender_detector(imie)
                #ok, q_imie = element_search(imie, 'item', 'pl', description=gender1)
                ok = False
                if not ok:
                    q_imie = '{Q:' + f'{imie}' + '}'
                f.write(f'LAST\t{P_IMIE}\t{q_imie}\n')
            if imie2:
                gender = gender_detector(imie2)
                # czy to przypadek 'Maria', 'Anna'?
                if imie2 in MALE_FEMALE_NAME and gender1 == 'imię męskie' and gender != gender1:
                    gender = gender1
                #ok, q_imie = element_search(imie2, 'item', 'pl', description=gender)
                ok = False
                if not ok:
                    q_imie = '{Q:' + f'{imie2}' + '}'
                f.write(f'LAST\t{P_IMIE}\t{q_imie}\n')
            if imie3:
                gender = gender_detector(imie3)
                # czy to przypadek 'Maria', 'Anna'?
                if imie3 in MALE_FEMALE_NAME and gender1 == 'imię męskie' and gender != gender1:
                    gender = gender1
                #ok, q_imie = element_search(imie3, 'item', 'pl', description=gender)
                ok = False # na razie nie szukamy
                if not ok:
                    q_imie = '{Q:' + f'{imie3}' + '}'
                f.write(f'LAST\t{P_IMIE}\t{q_imie}\n')
            if imie4:
                gender = gender_detector(imie4)
                # czy to przypadek 'Maria', 'Anna'?
                if imie4 in MALE_FEMALE_NAME and gender1 == 'imię męskie' and gender != gender1:
                    gender = gender1
                #ok, q_imie = element_search(imie4, 'item', 'pl', description=gender)
                ok = False # na razie nie szukamy
                if not ok:
                    q_imie = '{Q:' + f'{imie4}' + '}'
                f.write(f'LAST\t{P_IMIE}\t{q_imie}\n')
            if nazwisko:
                #ok, q_nazwisko = element_search(nazwisko, 'item', 'en', description='family name')
                ok = False
                if not ok:
                    q_nazwisko = '{Q:' + f'{nazwisko}' + '}'
                f.write(f'LAST\t{P_NAZWISKO}\t{q_nazwisko}\n')
            if nazwisko2:
                #ok, q_nazwisko = element_search(nazwisko2, 'item', 'en', description='family name')
                ok = False
                if not ok:
                    q_nazwisko = '{Q:' + f'{nazwisko2}' + '}'
                f.write(f'LAST\t{P_NAZWISKO}\t{q_nazwisko}\n')

            # imię i nazwisko przy urodzeniu (głównie dla zakoników/zakonnic)
            if z_mark:
                f.write(f'LAST\t{P_BIRTH_NAME}\tpl:"{birth_name}"')

            f.write(f'LAST\t{P_INSTANCE_OF}\t{Q_HUMAN}\n')

            # daty życia - z obsługą kwalifikatorów

            # data urodzin
            if dateB and dateB != 'somevalue' and dateB_dod == 'certain':
                f.write(f'LAST\t{P_DATE_OF_BIRTH}\t{dateB}\n')
            elif dateB == 'somevalue' and dateB_dod == 'between':
                if dateB_1 and dateB_2:
                    fd.write(f'Q:{name_etykieta}|{years}\t{P_DATE_OF_BIRTH}\t{dateB}\t{P_EARLIEST_DATE}\t{dateB_1}\t{P_LATEST_DATE}\t{dateB_2}\n')
            elif dateB == 'somevalue' and dateB_dod == 'before':
                fd.write(f'Q:{name_etykieta}|{years}\t{P_DATE_OF_BIRTH}\t{dateB}\t{P_LATEST_DATE}\t{dateB_1}\n')
            elif dateB == 'somevalue' and dateB_dod == 'after':
                fd.write(f'Q:{name_etykieta}|{years}\t{P_DATE_OF_BIRTH}\t{dateB}\t{P_EARLIEST_DATE}\t{dateB_1}\n')
            elif dateB and dateB_dod == 'about':
                f.write(f'LAST\t{P_DATE_OF_BIRTH}\t{dateB}\t{P_SOURCING_CIRCUMSTANCES}\t{Q_CIRCA}\n')
            elif dateB_1 and dateB_2 and dateB_dod == 'or':
                f.write(f'LAST\t{P_DATE_OF_BIRTH}\t{dateB_1}\n')
                f.write(f'LAST\t{P_DATE_OF_BIRTH}\t{dateB_2}\n')
            elif dateB_1 and dateB_2 and dateB_dod == 'turn':
                f.write(f'LAST\t{P_DATE_OF_BIRTH}\t{dateB_1}\n')
                f.write(f'LAST\t{P_DATE_OF_BIRTH}\t{dateB_2}\n')
            elif dateB_1  and dateB_dod == 'roman':
                f.write(f'LAST\t{P_FLUORIT}\t{dateB_1}\n')
                if dateB_2:
                    f.write(f'LAST\t{P_FLUORIT}\t{dateB_2}\n')

            # data śmierci
            if dateD and dateD != 'somevalue' and dateD_dod == 'certain':
                f.write(f'LAST\t{P_DATE_OF_DEATH}\t{dateD}\n')
            elif dateD == 'somevalue' and dateD_dod == 'between':
                if dateD_1 and dateD_2:
                    fd.write(f'Q:{name_etykieta}|{years}\t{P_DATE_OF_DEATH}\t{dateD}\t{P_EARLIEST_DATE}\t{dateD_1}\t{P_LATEST_DATE}\t{dateD_2}\n')
            elif dateD == 'somevalue' and dateD_dod == 'before':
                fd.write(f'Q:{name_etykieta}|{years}\t{P_DATE_OF_DEATH}\t{dateD}\t{P_LATEST_DATE}\t{dateD_1}\n')
            elif dateD == 'somevalue' and dateD_dod == 'after':
                fd.write(f'Q:{name_etykieta}|{years}\t{P_DATE_OF_DEATH}\t{dateD}\t{P_EARLIEST_DATE}\t{dateD_1}\n')
            elif dateD and dateD_dod == 'about':
                f.write(f'LAST\t{P_DATE_OF_DEATH}\t{dateD}\t{P_SOURCING_CIRCUMSTANCES}\t{Q_CIRCA}\n')
            elif dateD_1 and dateD_2 and dateD_dod == 'or':
                f.write(f'LAST\t{P_DATE_OF_DEATH}\t{dateD_1}\n')
                f.write(f'LAST\t{P_DATE_OF_DEATH}\t{dateD_2}\n')
            elif dateD_1 and dateD_2 and dateD_dod == 'turn':
                f.write(f'LAST\t{P_DATE_OF_DEATH}\t{dateD_1}\n')
                f.write(f'LAST\t{P_DATE_OF_DEATH}\t{dateD_2}\n')
            elif dateD and dateD_dod == 'roman':
                f.write(f'LAST\t{P_FLUORIT}\t{dateD}\n')

            # opisany w źródle
            f.write(f'LAST\t{P_DESCRIBED_BY_SOURCE}\t{q_biogram}\n')
            if viaf_ok and viaf_id and viaf_url:
                # wystarczy samo id, reference url jest już zbędny
                f.write(f'LAST\t{P_VIAF}\t"{viaf_id}"\n')
                WERYFIKACJA_VIAF[name] = viaf_url

    # zamrażanie słownika identyfikatów VIAF_ID
    if SAVE_DICT:
        with open(biogramy_pickle, 'wb') as handle:
            pickle.dump(BIOGRAMY, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(postacie_pickle, 'wb') as handle:
            pickle.dump(VIAF_ID, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(postacie_birth_pickle, 'wb') as handle:
            pickle.dump(VIAF_BIRTH, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(postacie_death_pickle, 'wb') as handle:
            pickle.dump(VIAF_DEATH, handle, protocol=pickle.HIGHEST_PROTOCOL)

    # zapis wyszukiwań VIAF w HTML dla łatwiejszej weryfikacji
    with open(postacie_viaf_html, "w", encoding='utf-8') as h:
        h.write('<html>\n')
        h.write('<head>\n')
        h.write('<meta charset="utf-8">\n')
        h.write('<title>Weryfikacja VIAF dla postaci PSB</title>\n')
        h.write('</head>\n')
        h.write('<body>\n')
        h.write('<h2>Weryfikacja VIAF dla postaci PSB</h2>\n')
        h.write('<table>\n')
        for key, val in WERYFIKACJA_VIAF.items():
            h.write(f'<tr><td>{key}</td><td><a href="{val}">{val}</td></tr>\n')
        h.write('</table>\n')
        h.write('</body>\n')
        h.write('</html>\n')

""" import gmin z pliku gminy.xlsx z danymi z PRG"""
import os
import time
import pickle
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikidariahtools import element_search_adv, search_by_purl, get_properties, get_elements
from property_import import create_statement_data, has_statement


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

BOT_LOGIN = os.environ.get('WIKIDARIAH_USER')
BOT_PASSWORD = os.environ.get('WIKIDARIAH_PWD')

# pomiar czasu wykonania
start_time = time.time()

WIKIBASE_WRITE = True

# standardowe właściwości
properties = get_properties(['instance of', 'stated as', 'reference URL', 'retrieved',
                            'has part or parts', 'TERYT'])

# elementy definicyjne
elements = get_elements(['http://purl.org/ontohgis#administrative_type_45'])

# typy gmin
typ_gminy_text = {}
typ_gminy_text['1'] = 'gmina miejska'
typ_gminy_text['2'] = 'gmina wiejska'
typ_gminy_text['3'] ='gmina miejsko-wiejska'
typ_gminy_text['4'] = 'miasto w gminie miejsko-wiejskiej'
typ_gminy_text['5'] = 'obszar wiejski w gminie miejsko-wiejskiej'
typ_gminy_text['8'] = 'dzielnice m. st. Warszawy'
typ_gminy_text['9'] = 'delegatury w miastach: Kraków, Łódź, Poznań i Wrocław'

# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':

    # wspólna referencja dla wszystkich deklaracji z PRG
    references = {}
    references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGIK/PRG/WFS/AdministrativeBoundaries'
    references[properties['retrieved']] = '2022-09-05'


    xlsx_input = '../data_prng/gminy.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["gminy"]

    q_gmina = elements['http://purl.org/ontohgis#administrative_type_45']

    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    parts = {}

    parts_pickle = '../data_prng/gminy_parts.pickle'
    if os.path.isfile(parts_pickle):
        with open(parts_pickle, 'rb') as handle:
            parts = pickle.load(handle)

    # tylko jak nie udało się wczytać z dysku
    if not parts:
        for index, row in enumerate(ws.iter_rows(2, ws.max_row), start=1):
            # wczytanie danych z xlsx
            nazwa = row[col_names['JPT_NAZWA_']].value
            if not nazwa:
                continue

            label_pl = label_en = nazwa

            print(f'{index}/{ws.max_row - 1} Przetwarzanie gminy: {label_pl}')

            teryt = row[col_names['JPT_KOD_JE']].value
            idiip = row[col_names['IIP_IDENTY']].value
            part_of = teryt[:4]
            typ_gm = teryt[-1]

            description_pl = 'gmina - współczesna jednostka administracyjna'
            description_en = 'gmina - współczesna jednostka administracyjna'

            # przygotowanie struktur wikibase
            data = []

            # part of i has part or parts
            ok, pow_qid = search_by_purl(properties['TERYT'], part_of)
            if not ok:
                pow_qid = ''

            # description
            if pow_qid:
                wb_pow = wbi_core.ItemEngine(item_id=pow_qid)
                pow_label_pl = wb_pow.get_label('pl')
            else:
                pow_label_pl = ''

            if pow_qid and pow_label_pl:
                parameters = [(properties['instance of'], q_gmina)]
                ok, item_id = element_search_adv(label_en, 'en', parameters, f"{description_en} ({pow_label_pl}, {typ_gminy_text[typ_gm]})")
                if not ok:
                    ok, item_id = element_search_adv(label_en, 'en', parameters, f"{description_en} ({pow_label_pl})")
                if ok and pow_qid:
                    if pow_qid in parts:
                        parts[pow_qid].append(item_id)
                    else:
                        parts[pow_qid] = [item_id]

        # save pickle
        with open(parts_pickle, 'wb') as handle:
            pickle.dump(parts, handle, protocol=pickle.HIGHEST_PROTOCOL)


    # uzupełnienie właściwości 'has part or parts' dla powiatów
    if parts:
        print("\nUzupełnienie 'has part or parts' dla powiatów.\n")

        # logowanie do instancji wikibase
        login_instance = wbi_login.Login(user=BOT_LOGIN, pwd=BOT_PASSWORD, token_renew_period=28800)

        for item_qid, part_qids in parts.items():
            tmp_data = []
            data = []
            for part_qid in part_qids:
                if not has_statement(item_qid, properties['has part or parts'], part_qid):
                    if WIKIBASE_WRITE:
                        statement = create_statement_data(properties['has part or parts'], part_qid, references, None, if_exists='APPEND')
                        if statement:
                            data.append(statement)
                    else:
                        tmp_data.append(part_qid)

            if WIKIBASE_WRITE:
                if data:
                    wd_item = wbi_core.ItemEngine(item_id=item_qid, data=data, debug=False)
                    wd_item.write(login_instance, entity_type='item')
                    print(f"Do elementu {item_qid} dodano właściwości {properties['has part or parts']}.")
                else:
                    print(f"ERROR: {item_qid}, wystąpił problem z przygotowaniem danych dla właściwości 'has part or parts'")
            else:
                if tmp_data:
                    print(f"Przygotowano dodanie właściwości {properties['has part or parts']}: {','.join(tmp_data)}")
                else:
                    print(f"ERROR: {item_qid}, wystąpił problem z przygotowaniem danych dla właściwości 'has part or parts'")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

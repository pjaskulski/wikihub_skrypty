""" Uzupełnianie danych miejscowosci z pliku miejscowosciU.xlsx (dane z PRG) """
import os
import sys
import time
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import element_search_adv, get_properties, get_elements
from property_import import create_statement_data


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

#BOT_LOGIN = os.environ.get('WIKIDARIAH_USER')
#BOT_PASSWORD = os.environ.get('WIKIDARIAH_PWD')

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = relogin_time = time.time()

WIKIBASE_WRITE = True

# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':
    # standardowe właściwości
    print('Przygotowanie właściwości...')
    properties = get_properties(['instance of', 'stated as', 'reference URL', 'retrieved',
                                'id SDI', 'part of', 'has part or parts', 'TERYT', 'settlement type',
                                'coordinate location', 'located in the administrative territorial entity',
                                'name status', 'inflectional ending', 'adjective form',
                                'located in the administrative territorial entity'
                                ])

    # elementy definicyjne
    print('Przygotowanie elementów definicyjnych...')
    elements = get_elements([ 'official name', 'human settlement',
        'part of a colony', 'part of a city', 'part of a settlement', 'part of a village',
        'colony', 'colony of a colony', 'colony of a settlement', 'colony of a village',
        'city/town', 'settlement', 'settlement of a colony', 'forest settlement',
        'forest settlement of a village', 'settlement of a settlement', 'settlement of a village',
        'housing developments', 'housing estate of a village', 'hamlet', 'hamlet of a colony',
        'hamlet of a settlement', 'hamlet of a village', 'tourist shelter', 'village'
                            ])

    settlement_type_map = {}
    settlement_type_map['część kolonii'] = 'part of a colony'
    settlement_type_map['część miasta'] = 'part of a city'
    settlement_type_map['część osady'] = 'part of a settlement'
    settlement_type_map['część wsi'] = 'part of a village'
    settlement_type_map['kolonia'] = 'colony'
    settlement_type_map['kolonia kolonii'] = 'colony of a colony'
    settlement_type_map['kolonia osady'] = 'colony of a settlement'
    settlement_type_map['kolonia wsi'] = 'colony of a village'
    settlement_type_map['miasto'] = 'city/town'
    settlement_type_map['osada'] = 'human settlement'
    settlement_type_map['osada kolonii'] = 'settlement of a colony'
    settlement_type_map['osada leśna'] = 'forest settlement'
    settlement_type_map['osada leśna wsi'] = 'forest settlement of a village'
    settlement_type_map['osada osady'] = 'settlement of a settlement'
    settlement_type_map['osada wsi'] = 'settlement of a village'
    settlement_type_map['osiedle'] = 'housing developments'
    settlement_type_map['osiedle wsi'] = 'housing estate of a village'
    settlement_type_map['przysiółek'] = 'hamlet'
    settlement_type_map['przysiółek kolonii'] = 'hamlet of a colony'
    settlement_type_map['przysiółek osady'] = 'hamlet of a settlement'
    settlement_type_map['przysiółek wsi'] = 'hamlet of a village'
    settlement_type_map['schronisko turystyczne'] = 'tourist shelter'
    settlement_type_map['wieś'] = 'village'


    # wspólna referencja dla wszystkich deklaracji z PRG
    references = {}
    references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'
    references[properties['retrieved']] = '2022-09-23'

    # logowanie do instancji wikibase
    #login_instance = wbi_login.Login(user=BOT_LOGIN, pwd=BOT_PASSWORD)

    login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                     consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                     access_token=WIKIDARIAH_ACCESS_TOKEN,
                                     access_secret=WIKIDARIAH_ACCESS_SECRET,
                                     token_renew_period=14400)

    xlsx_input = '../data_prng/miejscowosciU.xlsx'
    xlsx_input = '/home/piotr/ihpan/wikihub_skrypty/data_prng/miejscowosciU.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["miejscowosciU"]

    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    for index, row in enumerate(ws.iter_rows(2, ws.max_row), start=1):
        if index < 99638:
            continue
        # wczytanie danych z xlsx
        nazwa = row[col_names['NAZWAGLOWN']].value
        if not nazwa:
            continue

        label_pl = label_en = nazwa

        nazwa_miejsc = row[col_names['NAZWAMIEJS']].value
        rodzajobie = row[col_names['RODZAJOBIE']].value
        wgs84 = row[col_names['WGS84']].value
        identyfi_2 = row[col_names['IDENTYFI_2']].value
        gmina = row[col_names['GMINA']].value
        idiip = row[col_names['IDIIP']].value
        if gmina:
            gmina = gmina.split('-gmina')[0]
        powiat = row[col_names['POWIAT']].value
        wojewodztw = row[col_names['WOJEWODZTW']].value

        rodzaje_czesci_miejscowosci = ['część wsi', 'przysiółek osady', 'kolonia wsi',
                                       'część miasta', 'część kolonii', 'przysiółek wsi']

        # opis elementu
        description_pl = f'{rodzajobie} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'
        description_en = f'{rodzajobie} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'

        # przygotowanie struktur wikibase
        data = []

        # WGS84 - Point (23.29833332 52.68194448)
        if wgs84:
            wgs84 = wgs84.replace('Point', '').replace('(', '').replace(')','').strip()
            tmp = wgs84.split(' ')
            longitude = tmp[0]
            latitude = tmp[1]
            coordinate = f'{latitude},{longitude}'

       # id SDI
        if idiip:
            statement = create_statement_data(properties['id SDI'], idiip, None, None, add_ref_dict=references, if_exists='APPEND')
            if statement:
                data.append(statement)

        # szukanie elementu po etykiecie i opisie
        ok, item_id = element_search_adv(label_en, 'en', None, description_en)

        if not ok:
            # są miejscowości występujące w dziesiątkach jak 'Borki'
            ok, item_id = element_search_adv(label_en, 'en', None, description_en, max_results_to_verify=500)

        if not ok:
            # wariant dla części miejscowości z nazwą miejscowości nadrzędnej
            if rodzajobie in rodzaje_czesci_miejscowosci:
                description_en = f'{rodzajobie}: {nazwa_miejsc} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'
                ok, item_id = element_search_adv(label_en, 'en', None, description_en)

                if not ok:
                    # są miejscowości występujące w dziesiątkach jak 'Borki'
                    ok, item_id = element_search_adv(label_en, 'en', None, description_en, max_results_to_verify=500)

        if not ok:
            # szukanie ze współrzędnymi
            description_en = f'{description_en} [{coordinate}]'
            ok, item_id = element_search_adv(label_en, 'en', None, description_en, max_results_to_verify=20)
            if not ok:
                print(f"ERROR: nie znaleziono: {label_en}, {description_en}")
                with open('/home/piotr/ihpan/wikihub_skrypty/data_prng/miejscowosci_u.log', 'a', encoding='utf-8') as f:
                    f.write(f"ERROR: nie znaleziono: {label_en}, {description_en}\n")

        # jeżeli udało się odnaleźć pasujący element do uzupełniania
        if ok:
            if WIKIBASE_WRITE:
                test = 1
                while True:
                    try:
                        wb_item = wbi_core.ItemEngine(item_id=item_id, data=data)
                        wb_item.write(login_instance, bot_account=True, entity_type='item')
                        print(f"{index}/{ws.max_row - 1} {label_en} - {item_id} - Uzupełniono statement {properties['id SDI']}")
                        break
                    except MWApiError as wbdelreference_error:
                        err_code = wbdelreference_error.error_msg['error']['code']
                        message = wbdelreference_error.error_msg['error']['info']
                        print(f'ERROR: {err_code}, {message}')
                        if err_code in ['assertuserfailed', 'badtoken']:
                            if test == 1:
                                login_instance.generate_edit_credentials()
                                test += 1
                                continue

                        print(wbdelreference_error.error_msg)
                        end_time = time.time()
                        elapsed_time = end_time - start_time
                        print(f'Czas wykonania programu do wystąpienia błędu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')
                        sys.exit(1)

            else:
                print(f"{index}/{ws.max_row - 1} {label_en} - {item_id} - Przygotowano uzupełnienie statement {properties['id SDI']}")

        # czy to pomoże na zrywanie połączenia?
        if index % 50 == 0:
            time.sleep(2)


    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

""" import miejscowosci z pliku miejscowosciU.xlsx z danymi z PRNG"""
import os
import time
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import element_search_adv, get_properties, get_elements
from wikidariahtools import read_qid_from_text
from property_import import create_statement_data


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

WIKIBASE_WRITE = False

# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':
    # standardowe właściwości
    print('Przygotowanie właściwości...')
    properties = get_properties(['instance of', 'stated as', 'reference URL', 'retrieved',
                                'id SDI', 'part of', 'has part or parts', 'TERYT', 'settlement type',
                                'coordinate location', 'located in the administrative territorial entity',
                                'name status', 'inflectional ending', 'adjective form',
                                'located in the administrative territorial entity'
                                ])

    # elementy definicyjne
    print('Przygotowanie elementów definicyjnych...')
    elements = get_elements([ 'official name', 'human settlement',
        'part of a colony', 'part of a city', 'part of a settlement', 'part of a village',
        'colony', 'colony of a colony', 'colony of a settlement', 'colony of a village',
        'city/town', 'settlement', 'settlement of a colony', 'forest settlement',
        'forest settlement of a village', 'settlement of a settlement', 'settlement of a village',
        'housing developments', 'housing estate of a village', 'hamlet', 'hamlet of a colony',
        'hamlet of a settlement', 'hamlet of a village', 'tourist shelter', 'village'
                            ])

    settlement_type_map = {}
    settlement_type_map['część kolonii'] = 'part of a colony'
    settlement_type_map['część miasta'] = 'part of a city'
    settlement_type_map['część osady'] = 'part of a settlement'
    settlement_type_map['część wsi'] = 'part of a village'
    settlement_type_map['kolonia'] = 'colony'
    settlement_type_map['kolonia kolonii'] = 'colony of a colony'
    settlement_type_map['kolonia osady'] = 'colony of a settlement'
    settlement_type_map['kolonia wsi'] = 'colony of a village'
    settlement_type_map['miasto'] = 'city/town'
    settlement_type_map['osada'] = 'settlement'
    settlement_type_map['osada kolonii'] = 'settlement of a colony'
    settlement_type_map['osada leśna'] = 'forest settlement'
    settlement_type_map['osada leśna wsi'] = 'forest settlement of a village'
    settlement_type_map['osada osady'] = 'settlement of a settlement'
    settlement_type_map['osada wsi'] = 'settlement of a village'
    settlement_type_map['osiedle'] = 'housing developments'
    settlement_type_map['osiedle wsi'] = 'housing estate of a village'
    settlement_type_map['przysiółek'] = 'hamlet'
    settlement_type_map['przysiółek kolonii'] = 'hamlet of a colony'
    settlement_type_map['przysiółek osady'] = 'hamlet of a settlement'
    settlement_type_map['przysiółek wsi'] = 'hamlet of a village'
    settlement_type_map['schronisko turystyczne'] = 'tourist shelter'
    settlement_type_map['wieś'] = 'village'


    # wspólna referencja dla wszystkich deklaracji z PRG
    references = {}
    references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'
    references[properties['retrieved']] = '2022-09-23'

    # logowanie do instancji wikibase
    if WIKIBASE_WRITE:
        login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                         consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                         access_token=WIKIDARIAH_ACCESS_TOKEN,
                                         access_secret=WIKIDARIAH_ACCESS_SECRET,
                                         token_renew_period=14400)

    xlsx_input = '../data_prng/miejscowosciU.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["miejscowosciU"]

    # nazy kolumn w xlsx
    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    unique_item = {}
    parts = {}

    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1
        # wczytanie danych z xlsx
        nazwa = row[col_names['NAZWAGLOWN']].value
        if not nazwa:
            continue

        label_pl = label_en = nazwa

        dopelniacz = row[col_names['DOPELNIACZ']].value
        przymiotni = row[col_names['PRZYMIOTNI']].value
        nazwy_obocz = row[col_names['NAZWYOBOCZ']].value
        nazwy_dodat = row[col_names['NAZWYDODAT']].value
        nd_jezyk = row[col_names['ND_jezyk']].value
        nazwy_histo = row[col_names['NAZWYHISTO']].value
        nazwa_miejsc = row[col_names['NAZWAMIEJS']].value
        rodzajobie = row[col_names['RODZAJOBIE']].value
        wgs84 = row[col_names['WGS84']].value
        identyfi_2 = row[col_names['IDENTYFI_2']].value
        gmina = row[col_names['GMINA']].value
        if gmina:
            gmina = gmina.split('-gmina')[0]
        powiat = row[col_names['POWIAT']].value
        wojewodztw = row[col_names['WOJEWODZTW']].value
        row_prng = row[col_names['IDENTYFIKA']].value
        row_simc = row[col_names['IDENTYFI_1']].value

        rodzaje_czesci_miejscowosci = ['część wsi', 'przysiółek osady', 'kolonia wsi',
                                       'część miasta', 'część kolonii', 'przysiółek wsi']
        if rodzajobie in rodzaje_czesci_miejscowosci:
            description_pl = f'{rodzajobie}: {nazwa_miejsc} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'
            description_en = f'{rodzajobie}: {nazwa_miejsc} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'
        else:
            description_pl = f'{rodzajobie} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'
            description_en = f'{rodzajobie} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'

        # przygotowanie struktur wikibase
        data = []
        aliasy = {}

        # instance of
        statement = create_statement_data(properties['instance of'], elements['human settlement'],
                                          None, None, add_ref_dict=None)
        if statement:
            data.append(statement)

        # stated as NAZWA_GLOWN
        qualifiers = {}
        if dopelniacz:
            qualifiers[properties['inflectional ending']] = dopelniacz
        if przymiotni:
            qualifiers[properties['adjective form']] = przymiotni
        qualifiers[properties['name status']] = elements['official name']

        statement = create_statement_data(properties['stated as'], f'pl:"{nazwa}"', None,
                                          qualifiers, add_ref_dict=references)
        if statement:
            data.append(statement)

        # NAZWY_OBOCZ
        if nazwy_obocz:
            tmp = nazwy_obocz.split(',')
            for tmp_item in tmp:
                tmp_item = tmp_item.strip()
                if 'pl' in aliasy:
                    aliasy['pl'].append(tmp_item)
                else:
                    aliasy['pl'] = [tmp_item]
                statement = create_statement_data(properties['stated as'], f'pl:"{tmp_item}"',
                    None, None, add_ref_dict=references)
                if statement:
                    data.append(statement)

        # NAZWYDODAT
        if nazwy_dodat:
            statement = create_statement_data(properties['stated as'], f'{nd_jezyk}:"{nazwy_dodat}"',
                None, None, add_ref_dict=references)
            if statement:
                data.append(statement)
            if nd_jezyk in aliasy:
                aliasy[nd_jezyk].append(nazwy_dodat)
            else:
                aliasy[nd_jezyk] = [nazwy_dodat]

        # NAZWYHISTO
        if nazwy_histo:
            tmp = nazwy_histo.split(',')
            for tmp_item in tmp:
                tmp_item = tmp_item.strip()
                if 'pl' in aliasy:
                    aliasy['pl'].append(tmp_item)
                else:
                    aliasy['pl'] = [tmp_item]
                statement = create_statement_data(properties['stated as'], f'pl:"{tmp_item}"',
                    None, None, add_ref_dict=references)
                if statement:
                    data.append(statement)

        # RODZAJOBIE
        statement = create_statement_data(properties['settlement type'], elements[settlement_type_map[rodzajobie]],
            None, None, add_ref_dict=references)
        if statement:
            data.append(statement)

        # WGS84 - Point (23.29833332 52.68194448)
        if wgs84:
            wgs84 = wgs84.replace('Point', '').replace('(', '').replace(')','').strip()
            tmp = wgs84.split(' ')
            longitude = tmp[0]
            latitude = tmp[1]
            coordinate = f'{latitude},{longitude}'
            statement = create_statement_data(properties['coordinate location'], coordinate,
                None, None, add_ref_dict=references)
            if statement:
                data.append(statement)

        # IDENTYFI_2
        if identyfi_2:
            parameters = [(properties['TERYT'], identyfi_2)]
            ok, gmina_qid = element_search_adv(gmina, 'en', parameters)
            if ok:
                statement = create_statement_data(properties['located in the administrative territorial entity'],
                    gmina_qid, None, None, add_ref_dict=references)
                if statement:
                    data.append(statement)

        # identyfikator PRNG
        if row_prng:
            statement = create_statement_data(properties['prng id'], row_prng,
                None, None, add_ref_dict=references)
            if statement:
                data.append(statement)

        # identyfikator SIMC
        if row_simc:
            statement = create_statement_data(properties['SIMC place ID'], row_simc,
                None, None, add_ref_dict=references)
            if statement:
                data.append(statement)

        label_desc = f"{label_en}|{description_en}"
        if label_desc not in unique_item:
            unique_item[label_desc] = index
        else:
            description_en = f'{description_en} [{coordinate}]'
            description_pl = f'{description_pl} [{coordinate}]'
            label_desc = f"{label_en}|{description_en}"
            unique_item[label_desc] = index

        # etykiety, description, aliasy
        wb_item = wbi_core.ItemEngine(new_item=True, data=data)
        wb_item.set_label(label_en, lang='en')
        wb_item.set_label(label_pl, lang='pl')

        wb_item.set_description(description_en, 'en')
        wb_item.set_description(description_pl, 'pl')

        if aliasy:
            for alias_lang, alias_value in aliasy.items():
                for alias_item in alias_value:
                    wb_item.set_aliases(alias_item, alias_lang)

        if WIKIBASE_WRITE:
            try:
                new_id = wb_item.write(login_instance, bot_account=True, entity_type='item')
                if new_id:
                    print(f'{index}/{max_row - 1} Dodano nowy element: {label_en} / {label_pl} = {new_id}')
            except MWApiError as wbdelreference_error:
                err_code = wbdelreference_error.error_msg['error']['code']
                message = wbdelreference_error.error_msg['error']['info']
                if 'already has label' in message and err_code == 'modification-failed':
                    match_qid = read_qid_from_text(message)
                    print(f'{index}/{max_row - 1} Element: {label_en} / {label_pl} już istnieje {match_qid}.')
                elif err_code == 'assertuserfailed':
                    now = datetime.now()
                    date_time = now.strftime("%m/%d/%Y, %H:%M:%S")
                    print(f'{date_time} ERROR: {wbdelreference_error.error_msg}')
                    sys.exit(1)
                else:
                    print(f'ERROR: {wbdelreference_error.error_msg}')
        else:
            # wyszukiwanie po etykiecie, właściwości instance of oraz po opisie
            parameters = [(properties['instance of'], elements['human settlement'])]
            ok, item_id = element_search_adv(label_en, 'en', parameters, description_en)
            if not ok:
                new_id = 'TEST'
                print(f"{index}/{max_row - 1} Przygotowano dodanie elementu - {label_en} / {label_pl}  = {new_id}")
            else:
                print(f'{index}/{max_row - 1} Element: {label_en} / {label_pl} już istnieje: {item_id}')

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

""" Tworzenie xlsx do importu na podstawie xlsx z danymi z PRNG (Kraina, region)  """
import os
import sys
import time
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import element_search_adv, get_coord, get_properties, get_elements
from property_import import create_statement_data


# adresy wikibase
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

WIKIBASE_WRITE = False

# standardowe właściwości i elementy
properties = get_properties(['instance of', 'stated as', 'reference URL', 'retrieved',
                                'coordinate location', 'inflectional form', 'locative form',
                                'adjective form', 'located in (string)', 'located in country',
                                'id SDI'
                                ])

elements = get_elements(['country'])

# wspólna referencja dla wszystkich deklaracji z PRNG
references = {}
references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'

# kwalifikator z punktem czasowym
qualifiers_time = {}
qualifiers_time[properties['point in time']] = '+2022-00-00T00:00:00Z/9' # rok 2022


# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':

    # logowanie do instancji wikibase
    login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                        consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                        access_token=WIKIDARIAH_ACCESS_TOKEN,
                                        access_secret=WIKIDARIAH_ACCESS_SECRET,
                                        token_renew_period=14400)

    xlsx_input = '../data_prng/PRNG_egzonimy_panstwo.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["PRNG_egzonimy_panstwo"]

    # kolumny: nazwaGlown, odmianaNGD, odmianaNGM, odmianaNGP, nazwaDluga, odmianaNDD, odmianaNDM
    # odmianaNDP, nazwaObocz, odmianaNOD, odmianaNOM, odmianaNOP, nazwaHist, polozenieT, idiip,
    # wspGeograf, + instance of: “country”

    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1
        # wczytanie danych z xlsx
        nazwa = row[col_names['nazwaGlown']].value
        if not nazwa:
            continue

        label_pl = label_en = nazwa

        nazwa_hist = row[col_names['nazwaHist']].value
        nazwa_obocz = row[col_names['nazwaObocz']].value
        nazwa_dlug = row[col_names['nazwaDluga']].value

        odmiana_ngd = row[col_names['odmianaNGD']].value
        odmiana_ngm = row[col_names['odmianaNGM']].value
        odmiana_ngp = row[col_names['odmianaNGP']].value

        odmiana_ndd = row[col_names['odmianaNDD']].value
        odmiana_ndm = row[col_names['odmianaNDM']].value
        odmiana_ndp = row[col_names['odmianaNDP']].value

        odmiana_nod = row[col_names['odmianaNOD']].value
        odmiana_nom = row[col_names['odmianaNOM']].value
        odmiana_nop = row[col_names['odmianaNOP']].value

        idiip = row[col_names['idiip']].value
        polozenie_t = row[col_names['polozenieT']].value
        wsp_geo = row[col_names['wspGeograf']].value

        description_pl = 'kraj' # zamiast 'państwo'
        description_en = 'country'

        # przygotowanie struktur wikibase
        data = []
        aliasy = []

        # instance of
        statement = create_statement_data(properties['instance of'], elements['country'], None, qualifier_dict=qualifiers_time, add_ref_dict=None)
        if statement:
            data.append(statement)

        # współrzędne geograficzne
        if wsp_geo:
            coordinate = get_coord(wsp_geo)
            statement = create_statement_data(properties['coordinate location'], coordinate, None, qualifier_dict=qualifiers_time, add_ref_dict=references)
            if statement:
                data.append(statement)

        # id SDI
        if idiip:
            statement = create_statement_data(properties['id SDI'], idiip, None, qualifier_dict=qualifiers_time, add_ref_dict=references)
            if statement:
                data.append(statement)

        # nazwaGlow
        qualifiers = {}
        if odmiana_ngd:
            qualifiers[properties['inflectional form']] = odmiana_ngd
        if odmiana_ngm:
            qualifiers[properties['locative form']] = odmiana_ngm
        if odmiana_ngp:
            qualifiers[properties['adjective form']] = odmiana_ngp
        qualifiers.update(qualifiers_time)
        statement = create_statement_data(properties['stated as'], f'pl:"{nazwa}"', None, qualifiers, add_ref_dict=references)
        if statement:
            data.append(statement)

        # nazwaDlug
        if nazwa_dlug:
            aliasy.append(nazwa_dlug)
            qualifiers = {}
            if odmiana_ndd:
                qualifiers[properties['inflectional form']] = odmiana_ndd
            if odmiana_ndm:
                qualifiers[properties['locative form']] = odmiana_ndm
            if odmiana_ndp:
                qualifiers[properties['adjective form']] = odmiana_ndp
            qualifiers.update(qualifiers_time)
            statement = create_statement_data(properties['stated as'], f'pl:"{nazwa_dlug}"', None, qualifiers, add_ref_dict=references)
            if statement:
                data.append(statement)

        # nazwaObocz
        if nazwa_obocz:
            aliasy.append(nazwa_obocz)
            qualifiers = {}
            if odmiana_nod:
                qualifiers[properties['inflectional form']] = odmiana_nod
            if odmiana_nom:
                qualifiers[properties['locative form']] = odmiana_nom
            if odmiana_nop:
                qualifiers[properties['adjective form']] = odmiana_nop
            qualifiers.update(qualifiers_time)
            statement = create_statement_data(properties['stated as'], f'pl:"{nazwa_obocz}"', None, qualifiers, add_ref_dict=references)
            if statement:
                data.append(statement)

        # nazwaHist
        if nazwa_hist:
            aliasy.append(nazwa_hist)
            statement = create_statement_data(properties['stated as'], f'pl:"{nazwa_hist}"', None, None, add_ref_dict=references)
            if statement:
                data.append(statement)

        # polozenieT
        if polozenie_t:
            statement = create_statement_data(properties['located in (string)'], polozenie_t, None, qualifier_dict=qualifiers_time, add_ref_dict=references)
            if statement:
                data.append(statement)

        # etykiety, description, aliasy
        wb_item = wbi_core.ItemEngine(new_item=True, data=data)
        wb_item.set_label(label_en, lang='en')
        wb_item.set_label(label_pl,lang='pl')

        wb_item.set_description(description_en, 'en')
        wb_item.set_description(description_pl, 'pl')

        if aliasy:
            for value_alias in aliasy:
                wb_item.set_aliases(value_alias, 'pl')

        # wyszukiwanie po etykiecie
        parameters = [(properties['instance of'], elements['country'])]
        ok, item_id = element_search_adv(label_en, 'en', parameters)
        if not ok:
            if WIKIBASE_WRITE:
                test = 1
                while True:
                    try:
                        new_id = wb_item.write(login_instance, bot_account=True, entity_type='item')
                        print(f'Dodano nowy element: {label_en} / {label_pl} = {new_id}')
                        break
                    except MWApiError as wb_error:
                        err_code = wb_error.error_msg['error']['code']
                        message = wb_error.error_msg['error']['info']
                        print(f'ERROR: {err_code}, {message}')
                        # jeżeli jest to problem z tokenem to próba odświeżenia tokena i powtórzenie
                        # zapisu, ale tylko raz, w razie powtórnego błędu bad token, skrypt kończy pracę
                        if err_code in ['assertuserfailed', 'badtoken']:
                            if test == 1:
                                print('Generate edit credentials...')
                                login_instance.generate_edit_credentials()
                                test += 1
                                continue
                        sys.exit(1)
            else:
                new_id = 'TEST'
                print(f"Przygotowano dodanie elementu - {label_en} / {label_pl}  = {new_id}")
        else:
            #print(f'Element: {label_en} / {label_pl} już istnieje: {item_id}')
            print(f'# [https://prunus-208.man.poznan.pl/wiki/Item:{item_id} {label_en} / {label_pl}]')

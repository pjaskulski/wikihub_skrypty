""" skrypt dodaje brakujące deklaracje located in country do regionów """

import os
import time
import sys
from pathlib import Path
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_exceptions import MWApiError
from dotenv import load_dotenv
import openpyxl
from wikidariahtools import find_name_qid, element_search_adv, element_exists
from property_import import create_statement_data


# adresy
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# brak ustawienia tych wartości w wikibase powoduje ostrzeżenia, ale skrypt działa
#wbi_config['PROPERTY_CONSTRAINT_PID'] = 'Pxxx'
#wbi_config['DISTINCT_VALUES_CONSTRAINT_QID'] = 'Qxxx'

ok, p_instance_of = find_name_qid('instance of', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'instance of' w instancji Wikibase")
    sys.exit(1)
ok, p_located_in_country = find_name_qid('located in country', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'located in country' w instancji Wikibase")
    sys.exit(1)
ok, p_reference_url = find_name_qid('reference URL', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'reference URL' w instancji Wikibase")
    sys.exit(1)

# wspólna referencja dla wszystkich deklaracji z PRNG
references = {}
references[p_reference_url] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'

q_country = 'Q86557'

WIKIBASE_WRITE = True


# --------------------------------- MAIN ---------------------------------------

if __name__ == "__main__":
    # pomiar czasu wykonania
    start_time = time.time()

    # login i hasło ze zmiennych środowiskowych (plik .env w folderze ze źródłami)
    env_path = Path('.') / '.env'

    load_dotenv(dotenv_path=env_path)
    BOT_LOGIN = os.environ.get('WIKIDARIAH_USER')
    BOT_PASSWORD = os.environ.get('WIKIDARIAH_PWD')

    login_instance = wbi_login.Login(user=BOT_LOGIN, pwd=BOT_PASSWORD)

    # wczytanie danych z XLSX
    xlsx_input = '../data_prng/PRNG_egzonimy_region_source.xlsx'

    wb = openpyxl.load_workbook(xlsx_input)

    ws = wb["PRNG_egzonimy_region"]

    # kolumny: idiip, nazwaGlown, informDod, nazwaObocz, nazwaHist, Polozeniet, odmianaNGD,
    # odmianaNGM, odmianaNGP, odmianaNOD, odmianaNOM, odmianaNOP, WGS84

    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    inf = {}
    pol_t = {}
    for index, row in enumerate(ws.iter_rows(2, ws.max_row), start=1):
        nazwa = row[col_names['nazwaGlown']].value
        if not nazwa:
            continue

        inform_dod = row[col_names['informDod']].value
        polozenie_t = row[col_names['Polozeniet']].value
        if inform_dod:
            inf[nazwa] = inform_dod
        if polozenie_t:
            pol_t[nazwa] = polozenie_t

    # uzpełniane danych dla regionów - kraje z pola informDod
    start = 86628
    stop = 87091

    for i in range(start, stop + 1):
        item = f'Q{i}'
        if not element_exists(item):
            continue
        wb_item = wbi_core.ItemEngine(item_id=item)
        label_pl = wb_item.get_label('pl')
        if '(region)' in label_pl:
            label_pl = label_pl.replace('(region)','').strip()

        if label_pl in inf and 'obiekt transgraniczny:' in inf[label_pl]:
            data = []
            informacja = inf[label_pl].replace('obiekt transgraniczny:','').strip()

            tmp_tab = informacja.split(',')
            for tmp_tab_item in tmp_tab:
                parameters = [(p_instance_of ,q_country)]
                tmp_tab_item = tmp_tab_item.strip()
                ok, item_id = element_search_adv(tmp_tab_item, 'pl', parameters)
                if not ok:
                    continue

                statement = create_statement_data(p_located_in_country, item_id, None, None, add_ref_dict=references, if_exists='APPEND')
                if statement:
                    data.append(statement)

            if data:
                if WIKIBASE_WRITE:
                    try:
                        wb_update = wbi_core.ItemEngine(item_id=item, data=data, debug=False)
                        wb_update.write(login_instance, entity_type='item')
                        print(f'{item} ({label_pl}) - uzupełniono deklarację {p_located_in_country}')
                    except (MWApiError, KeyError) as err_update:
                        print(f"ERROR: {item} ({err_update.error_msg})")
                else:
                    print(f'{item} ({label_pl}) - dane przygotowane do zapisu.')

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

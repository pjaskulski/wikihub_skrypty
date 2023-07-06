""" import miejscowosci z pliku miejscowosciU.xlsx z danymi z PRNG"""
import os
import time
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import element_search_adv, get_properties, get_elements
from wikidariahtools import read_qid_from_text
from property_import import create_statement_data


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

# czy zapis do wikibase czy tylko test
WIKIBASE_WRITE = True

# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':
    # standardowe właściwości
    print('Przygotowanie właściwości... ', end='', flush=True)
    properties = get_properties(['instance of', 'stated as', 'reference URL', 'retrieved',
                                'id SDI', 'part of', 'has part or parts', 'TERYT', 'settlement type',
                                'coordinate location', 'located in the administrative territorial entity',
                                'name status', 'inflectional ending', 'adjective form',
                                'located in the administrative territorial entity', 'prng id',
                                'SIMC place ID'
                                ])
    print('OK')

    # elementy definicyjne
    print('Przygotowanie elementów definicyjnych... ', end='', flush=True)
    elements = get_elements([ 'official name', 'human settlement',
        'part of a colony', 'part of a city', 'part of a settlement', 'part of a village',
        'colony', 'colony of a colony', 'colony of a settlement', 'colony of a village',
        'city/town', 'settlement of a colony', 'forest settlement',
        'forest settlement of a village', 'settlement of a settlement', 'settlement of a village',
        'housing developments', 'housing estate of a village', 'hamlet', 'hamlet of a colony',
        'hamlet of a settlement', 'hamlet of a village', 'tourist shelter', 'village'
                            ])
    print('OK')

    settlement_type_map = {}
    settlement_type_map['część kolonii'] = 'part of a colony'
    settlement_type_map['część miasta'] = 'part of a city'
    settlement_type_map['część osady'] = 'part of a settlement'
    settlement_type_map['część wsi'] = 'part of a village'
    settlement_type_map['kolonia'] = 'colony'
    settlement_type_map['kolonia kolonii'] = 'colony of a colony'
    settlement_type_map['kolonia osady'] = 'colony of a settlement'
    settlement_type_map['kolonia wsi'] = 'colony of a village'
    settlement_type_map['miasto'] = 'city/town'
    settlement_type_map['osada'] = 'human settlement'
    settlement_type_map['osada kolonii'] = 'settlement of a colony'
    settlement_type_map['osada leśna'] = 'forest settlement'
    settlement_type_map['osada leśna wsi'] = 'forest settlement of a village'
    settlement_type_map['osada osady'] = 'settlement of a settlement'
    settlement_type_map['osada wsi'] = 'settlement of a village'
    settlement_type_map['osiedle'] = 'housing developments'
    settlement_type_map['osiedle wsi'] = 'housing estate of a village'
    settlement_type_map['przysiółek'] = 'hamlet'
    settlement_type_map['przysiółek kolonii'] = 'hamlet of a colony'
    settlement_type_map['przysiółek osady'] = 'hamlet of a settlement'
    settlement_type_map['przysiółek wsi'] = 'hamlet of a village'
    settlement_type_map['schronisko turystyczne'] = 'tourist shelter'
    settlement_type_map['wieś'] = 'village'


    # wspólna referencja dla wszystkich deklaracji z PRG
    references = {}
    references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'
    references[properties['retrieved']] = '2022-09-23'

    # logowanie do instancji wikibase
    print('Logowanie do wikibase... ', end='', flush=True)
    login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                     consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                     access_token=WIKIDARIAH_ACCESS_TOKEN,
                                     access_secret=WIKIDARIAH_ACCESS_SECRET,
                                     token_renew_period=14400)
    print('OK')

    print('Wczytanie danych z pliku XLSX... ', end='', flush=True)
    xlsx_input = '../data_prng/miejscowosciU.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["miejscowosciU"]

    # nazy kolumn w xlsx
    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1
    print('OK')

    parts = {}
    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1

        if index <= 101133:
            continue

        # wczytanie danych z xlsx
        nazwa = row[col_names['NAZWAGLOWN']].value
        if not nazwa:
            continue

        label_pl = label_en = nazwa
        identyfi_2 = row[col_names['IDENTYFI_2']].value
        gmina = row[col_names['GMINA']].value
        if gmina:
            gmina = gmina.split('-gmina')[0]
        update_qid = row[col_names['QID']].value
        # print(nazwa, update_qid)

        # przygotowanie struktur wikibase
        data = []

        # IDENTYFI_2
        if identyfi_2:
            parameters = [(properties['TERYT'], identyfi_2)]
            ok, gmina_qid = element_search_adv('commune' + ' ' + gmina, 'en', parameters)
            if ok:
                statement = create_statement_data(properties['located in the administrative territorial entity'],
                    gmina_qid, None, None, add_ref_dict=references)
                if statement:
                    data.append(statement)
            else:
                print(f'ERROR: nie znaleziono gminy ({gmina})')

        wb_item = wbi_core.ItemEngine(item_id=update_qid, data=data)

        # korekta opisu
        description_en = wb_item.get_description(lang='en')
        if 'część kolonii' in description_en:
            description_en = description_en.replace('część kolonii', 'part of a colony')
        elif 'część miasta' in description_en:
            description_en = description_en.replace('część miasta','part of a city')
        elif 'część osady' in description_en:
            description_en = description_en.replace('część osady','part of a settlement')
        elif 'część wsi' in description_en:
            description_en = description_en.replace('część wsi','part of a village')
        elif 'kolonia kolonii' in description_en:
            description_en = description_en.replace('kolonia kolonii','colony of a colony')
        elif 'kolonia osady' in description_en:
            description_en = description_en.replace('kolonia osady','colony of a settlement')
        elif 'kolonia wsi' in description_en:
            description_en = description_en.replace('kolonia wsi','colony of a village')
        elif 'kolonia' in description_en:
            description_en = description_en.replace('kolonia','colony')
        elif 'miasto' in description_en:
            description_en = description_en.replace('miasto','city/town')
        elif 'osada kolonii' in description_en:
            description_en = description_en.replace('osada kolonii','settlement of a colony')
        elif 'osada leśna wsi' in description_en:
            description_en = description_en.replace('osada leśna wsi','forest settlement of a village')
        elif 'osada leśna' in description_en:
            description_en = description_en.replace('osada leśna','forest settlement')
        elif 'osada osady' in description_en:
            description_en = description_en.replace('osada osady','settlement of a settlement')
        elif 'osada wsi' in description_en:
            description_en = description_en.replace('osada wsi','settlement of a village')
        elif 'osada' in description_en:
            description_en = description_en.replace('osada','human settlement')
        elif 'osiedle wsi' in description_en:
            description_en = description_en.replace('osiedle wsi','housing estate of a village')
        elif 'osiedle' in description_en:
            description_en = description_en.replace('osiedle','housing developments')
        elif 'przysiółek kolonii' in description_en:
            description_en = description_en.replace('przysiółek kolonii','hamlet of a colony')
        elif 'przysiółek osady' in description_en:
            description_en = description_en.replace('przysiółek osady','hamlet of a settlement')
        elif 'przysiółek wsi' in description_en:
            description_en = description_en.replace('przysiółek wsi','hamlet of a village')
        elif 'przysiółek' in description_en:
            description_en = description_en.replace('przysiółek','hamlet')
        elif 'schronisko turystyczne' in description_en:
            description_en = description_en.replace('schronisko turystyczne','tourist shelter')
        elif 'wieś' in description_en:
            description_en = description_en.replace('wieś','village')

        wb_item.set_description(description=description_en, lang='en')

        if WIKIBASE_WRITE:
            test = 1
            while True:
                try:
                    wb_item.write(login_instance, bot_account=True, entity_type='item')
                    print(f'{index}/{max_row - 1} Uzupełniono element: {label_en} / {label_pl} = {update_qid}')
                    break
                except MWApiError as wbdelreference_error:
                    err_code = wbdelreference_error.error_msg['error']['code']
                    message = wbdelreference_error.error_msg['error']['info']
                    print(f'ERROR: {err_code}, {message}')
                    if 'already has label' in message and err_code == 'modification-failed':
                        match_qid = read_qid_from_text(message)
                        print(f'{index}/{max_row - 1} Element: {label_en} / {label_pl} już istnieje {match_qid}.')
                        break
                    else:
                        if err_code in ['assertuserfailed', 'badtoken']:
                            if test == 1:
                                print('Generate edit credentials...')
                                login_instance.generate_edit_credentials()
                                test += 1
                                continue
                        sys.exit(1)
        else:
            print(f"{index}/{max_row - 1} Przygotowano uzupełnienie elementu - {label_en} / {label_pl}  = {update_qid}")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

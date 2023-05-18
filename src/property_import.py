""" skrypt importujący właściwości z xlsx do wikibase """

import os
import sys
import re
import json
import time
import warnings
from pathlib import Path
from typing import Union
from openpyxl import load_workbook
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login, wbi_datatype
from wikibaseintegrator.wbi_functions import mediawiki_api_call_helper
from wikibaseintegrator.wbi_exceptions import MWApiError
from dotenv import load_dotenv
from wikidariahtools import element_search, search_by_purl, get_property_type, statement_value_fix

warnings.filterwarnings("ignore")

# adresy dla API Wikibase
wbi_config["MEDIAWIKI_API_URL"] = "https://prunus-208.man.poznan.pl/api.php"
wbi_config["SPARQL_ENDPOINT_URL"] = "https://prunus-208.man.poznan.pl/bigdata/sparql"
wbi_config["WIKIBASE_URL"] = "https://prunus-208.man.poznan.pl"

# słownik globalnych referencji dla arkuszy (z deklaracjami)
GLOBAL_REFERENCE = {}
# słowniki dodawanych/modyfikowanych właściwości i elementów
GLOBAL_PROPERTY = {}
GLOBAL_ITEM = {}
# do trybu testowego
#QID_LICZNIK = 0
#TEST_QID = {}

# parametr globalny czy zapisywać dane do wikibase, jeżeli = False dla nowych
# właściwości i elementów zwraca QID = TEST
WIKIBASE_WRITE = True

# --- klasy ---
class BasicProp:
    """Identyfikatory podstawowych właściwości"""

    def __init__(self):
        self.wiki_id = ""
        self.wiki_url = ""
        self.inverse = ""

    def get_wiki_properties(self):
        """funkcja ustala nr podstawowych property związanych z wikidata.org"""
        if self.wiki_id == "":
            search_result, pid = element_search("Wikidata ID", "property", "en")
            if search_result:
                self.wiki_id = pid
        if self.wiki_url == "":
            search_result, pid = element_search("reference URL", "property", "en")
            if search_result:
                self.wiki_url = pid
        if self.inverse == "":
            search_result, pid = element_search("inverse property", "property", "en")
            if search_result:
                self.inverse = pid


class WDHSpreadsheet:
    """Plik arkusza kalkulacyjnego z modelem danych dla Wikibase"""

    def __init__(self, path: str):
        self.path = path
        #if self.path.endswith('.yaml'):
        #    self.read_from_yaml()
        self.sheets = ["P_list", "P_statements", "Q_list", "Q_statements", "Globals"]
        self.p_list = None  # arkusz z listą właściwości
        self.p_statements = None  # arkusz z listą deklaracji dla właściwości
        self.workbook = None
        self.property_columns = []
        self.statement_columns = []
        self.i_list = None  # arkusz z listą elementów (item)
        self.i_statements = None  # arkusz z listą deklaracji dla elementów
        self.item_columns = []
        self.item_statement_columns = []
        self.globals = None  # arkusz z globalnymi referencjami
        self.globals_columns = []

    @property
    def path(self) -> str:
        """get path"""
        return self._path

    @path.setter
    def path(self, value: str):
        """set path"""
        self._path = value

    def open(self):
        """odczyt pliku i weryfikacja poprawności"""
        try:
            self.workbook = load_workbook(
                self.path, data_only=True
            )  # czytanie wartości a nie formuł (?)
        except IOError:
            print(f"ERROR. Can't open and process file: {self.path}")
            sys.exit(1)

        # czy to jest właściwy plik? cz. 1
        for sheet in self.sheets:
            if not sheet in self.workbook.sheetnames:
                print(f"ERROR. Expected worksheet '{sheet}' is missing in the file.")
                sys.exit(1)

        # arkusz właściwości
        self.p_list = self.workbook[self.sheets[0]]
        self.property_columns = self.get_col_names(self.p_list)
        p_list_expected = ["Label_en", "Description_en", "Datatype", "Label_pl"]
        res, inf = self.test_columns(self.property_columns, p_list_expected)
        if not res:
            print(
                f"ERROR. Worksheet {self.sheets[0]}. The expected columns ({inf}) are missing."
            )
            sys.exit(1)

        # arkusz deklaracji dla właściwości
        self.p_statements = self.workbook[self.sheets[1]]
        self.statement_columns = self.get_col_names(self.p_statements)
        p_statements_expected = ["Label_en", "P", "Value"]
        res, inf = self.test_columns(self.statement_columns, p_statements_expected)
        if not res:
            print(
                f"ERROR. Worksheet {self.sheets[1]}. The expected columns ({inf}) are missing."
            )
            sys.exit(1)

        # arkusz elementów
        self.i_list = self.workbook[self.sheets[2]]
        self.item_columns = self.get_col_names(self.i_list)
        # możliwe są także opcjonalne kolumny StartsAt, EndsAt, Instance of, Purl_identifier
        i_list_expected = [
            "Label_en",
            "Label_pl",
            "Description_en",
            "Description_pl",
            "Wiki_id",
        ]
        res, inf = self.test_columns(self.item_columns, i_list_expected)
        if not res:
            print(
                f"ERROR. Worksheet {self.sheets[2]}. The expected columns ({inf}) are missing."
            )
            sys.exit(1)

        # arkusz deklaracji dla elementów
        self.i_statements = self.workbook[self.sheets[3]]
        self.item_statement_columns = self.get_col_names(self.i_statements)
        i_statements_expected = [
            "Label_en",
            "P",
            "Value",
            "Qualifier",
            "Qualifier_value",
            "Reference_property",
            "Reference_value"
        ]
        res, inf = self.test_columns(self.item_statement_columns, i_statements_expected)
        if not res:
            print(
                f"ERROR. Worksheet {self.sheets[3]}. The expected columns ({inf}) are missing."
            )
            sys.exit(1)

        # arkusz globalnych referencji dla poszczególnych arkuszy
        self.globals = self.workbook[self.sheets[4]]
        self.globals_columns = self.get_col_names(self.globals)
        globals_expected = ["Sheet", "Reference_property", "Reference_value"]
        res, inf = self.test_columns(self.globals_columns, globals_expected)
        if not res:
            print(
                f"ERROR. Worksheet {self.sheets[4]}. The expected columns ({inf}) are missing."
            )
            sys.exit(1)

    def test_columns(self, t_col_names: dict, expected: list) -> tuple:
        """weryfikuje czy arkusz zawiera oczekiwane kolumny"""
        missing_cols = []
        res = True
        for col in expected:
            if not col in t_col_names:
                missing_cols.append(col)
                res = False

        return res, ",".join(missing_cols)

    def get_col_names(self, sheet) -> dict:
        """funkcja zwraca słownik nazw kolumn"""
        names = {}
        nr_col = 0
        for column in sheet.iter_cols(1, sheet.max_column):
            names[column[0].value] = nr_col
            nr_col += 1

        return names

    def correct_type(self, t_datatype: str) -> str:
        """Funkcja ewentualnie koryguje typ właściwości na właściwy, zgodny z oczekiwanym
        przez Wikibase
        """
        if t_datatype is not None:
            if t_datatype == "item":
                t_datatype = "wikibase-item"
            elif t_datatype == "property":
                t_datatype = "wikibase-property"
            elif t_datatype == "external identifier":
                t_datatype = "external-id"
            elif value == "URL":
                value = "url"
            elif value == "monolingual text":
                value = "monolingualtext"
            elif value == "geographic coordinates":
                value = "globe-coordinate"
            elif value == "point in time":
                value = "time"

        return t_datatype

    def get_property_list(self) -> list:
        """zwraca listę właściwości (w formie obiektów WDHProperty) do dodania"""
        p_list = []
        for row in self.p_list.iter_rows(2, self.p_list.max_row):
            basic_cols = ["Label_en", "Description_en", "Datatype", "Label_pl"]
            p_item = {}
            p_item = WDHProperty()
            for col in basic_cols:
                key = col.lower()
                col_value = row[self.property_columns[col]].value
                if key == "label_en":
                    p_item.label_en = col_value
                elif key == "description_en":
                    p_item.description_en = col_value
                elif key == "datatype":
                    p_item.datatype = col_value
                elif key == "label_pl":
                    p_item.label_pl = col_value

            # tylko jeżeli: etykieta i opis w języku angielskim lub etykiety pl i ang - oraz
            # typ danych są wypełnione dane właściwości są dodawane do listy
            if (
                (p_item.label_en and p_item.description_en)
                or (p_item.label_en and p_item.label_pl)
            ) and p_item.datatype:
                extend_cols = ["Description_pl", "Wiki_id", "Inverse_property"]
                for col in extend_cols:
                    key = col.lower()
                    col_value = row[self.property_columns[col]].value
                    if key == "description_pl":
                        p_item.description_pl = col_value
                    elif key == "wiki_id":
                        p_item.wiki_id = col_value
                    elif key == "inverse_property":
                        p_item.inverse_property = col_value

                p_list.append(p_item)

        return p_list

    def get_statement_list(self) -> list:
        """zwraca listę obiektów deklaracji dla właściwości do dodania"""
        s_list = []
        for row in self.p_statements.iter_rows(2, self.p_statements.max_row):
            basic_cols = [
                "Label_en",
                "P",
                "Value",
                "Reference_property",
                "Reference_value",
            ]
            s_item = WDHStatementProperty()
            reference_property = reference_value = ""
            for col in basic_cols:
                # tylko jeżeli kolumna z listy jest w pliku
                if col in self.statement_columns:
                    key = col.lower()
                    col_value = row[self.statement_columns[col]].value

                    if key == "label_en":
                        s_item.label_en = col_value
                    elif key == "p":
                        s_item.statement_property = col_value
                    elif key == "value":
                        if not isinstance(col_value, str):
                            col_value = str(col_value)
                        s_item.statement_value = col_value
                    elif key == "reference_property":
                        reference_property = col_value
                    elif key == "reference_value":
                        if not isinstance(col_value, str):
                            col_value = str(col_value)
                        reference_value = col_value

                    if reference_property and reference_value:
                        s_item.references[reference_property] = reference_value

            # nazwa arkusza z deklaracjami dla właściwości
            s_item.sheet_name = self.sheets[1]

            # jeżeli są globalne referencje
            if s_item.sheet_name in GLOBAL_REFERENCE:
                g_ref_property, g_ref_value = GLOBAL_REFERENCE[s_item.sheet_name]
                s_item.additional_references[g_ref_property] = g_ref_value

            # tylko jeżeli etykieta w języku angielskim, właściwość i wartość są wypełnione
            # dane deklaracji są dodawane do listy
            if s_item.label_en and s_item.statement_property and s_item.statement_value:
                s_list.append(s_item)
            # jeżeli nie ma wartości etykiety, właściwości i wartości deklaracji
            # a są dane referencji to  dodaje referencje do ostatnio dodanej
            # pozycji z listy
            else:
                if reference_property and reference_value:
                    s_list[-1].references[reference_property] = reference_value

        return s_list

    def get_item_list(self) -> list:
        """zwraca listę elementów (w formie obiektów WDHItem) do dodania"""
        i_list = []
        for row in self.i_list.iter_rows(2, self.i_list.max_row):
            basic_cols = ["Label_en", "Label_pl", "Description_en", "Description_pl"]
            i_item = WDHItem()
            for col in basic_cols:
                key = col.lower()
                col_value = row[self.item_columns[col]].value
                if key == "label_en":
                    i_item.label_en = col_value
                elif key == "description_en":
                    i_item.description_en = col_value
                elif key == "description_pl":
                    i_item.description_pl = col_value
                elif key == "label_pl":
                    i_item.label_pl = col_value

            # tylko jeżeli etykieta i opis w języku angielskim lub etykieta polska
            # są wypełnione dane elementu są dodawane do listy
            if (i_item.label_en and i_item.description_en) or (i_item.label_pl):
                extend_cols = [
                    "Wiki_id",
                    "StartsAt",
                    "EndsAt",
                    "Instance of",
                    "Purl identifier",
                    "Reference_property",
                    "Reference_value"
                ]
                for col in extend_cols:
                    key = col.lower()
                    if col in self.item_columns:
                        col_value = row[self.item_columns[col]].value
                        if key == "wiki_id":
                            i_item.wiki_id = col_value
                        elif key == "startsat":
                            if col_value is None:
                                col_value = ""
                            if not isinstance(col_value, str):
                                col_value = str(col_value)
                            i_item.starts_at = col_value
                        elif key == "endsat":
                            if col_value is None:
                                col_value = ""
                            if not isinstance(col_value, str):
                                col_value = str(col_value)
                            i_item.ends_at = col_value
                        elif key == "instance of":
                            i_item.instance_of = col_value
                        elif key == "purl identifier":
                            if col_value is None:
                                col_value = ""
                            if not isinstance(col_value, str):
                                col_value = str(col_value)
                            i_item.purl_identifier = col_value

                i_list.append(i_item)

        return i_list


    def get_item_statement_list(self) -> list:
        """zwraca listę obiektów deklaracji do dodania do elementów"""

        s_list = []
        for row in self.i_statements.iter_rows(2, self.i_statements.max_row):
            basic_cols = ["Label_en", "P", "Value", "Qualifier", "Qualifier_value",
                          "Reference_property", "Reference_value"]

            label_en = statement_property = statement_value = qualifier = qualifier_value = ""
            reference_property = reference_value = ""
            for col in basic_cols:
                key = col.lower()
                col_value = row[self.item_statement_columns[col]].value

                if key == "label_en":
                    label_en = col_value
                elif key == "p":
                    statement_property = col_value
                elif key == "value":
                    statement_value = col_value
                    if not isinstance(statement_value, str):
                        statement_value = str(statement_value)
                    # jeżeli to multilingual text to weryfikacja cudzysłowów
                    statement_value = monolingual_text_fix(statement_value)
                elif key == "qualifier":
                    qualifier = col_value
                elif key == "qualifier_value":
                    qualifier_value = col_value
                    if not isinstance(qualifier_value, str):
                        qualifier_value = str(qualifier_value)
                    # jeżeli to multilingual text to weryfikacja cudzysłowów
                    qualifier_value = monolingual_text_fix(qualifier_value)
                elif key == "reference_property":
                    reference_property = col_value
                elif key == "reference_value":
                    reference_value = col_value
                    if not isinstance(reference_value, str):
                        reference_value = str(reference_value)

            # tylko jeżeli etykieta w języku angielskim, właściwość i wartość są wypełnione
            # dane deklaracji są dodawane do listy
            if label_en and statement_property and statement_value:
                s_item = WDHStatementItem()
                s_item.label_en = label_en
                s_item.statement_property = statement_property
                s_item.statement_value = statement_value
                if qualifier and qualifier_value:
                    #print('Kwalifikator:', qualifier, qualifier_value)
                    s_item.qualifiers[qualifier] = qualifier_value
                if reference_property and reference_property != 'NONE' and reference_value:
                    s_item.references[reference_property] = reference_value
                s_item.sheet_name = self.sheets[3]

                # jeżeli są globalne referencje
                if s_item.sheet_name in GLOBAL_REFERENCE:
                    if reference_property == "":
                        g_ref_property, g_ref_value = GLOBAL_REFERENCE[s_item.sheet_name]
                        s_item.additional_references[g_ref_property] = g_ref_value

                s_list.append(s_item)
            # jeżeli nie ma wartości etykiety, właściwości i wartości deklaracji
            # a są dane kwalifikatora to  dodaje kwalifikator do ostatnio dodanej
            # pozycji z listy, podobnie dla referencji (tzw. lokalnych)
            else:
                if qualifier and qualifier_value:
                    s_list[-1].qualifiers[qualifier] = qualifier_value
                if reference_property and reference_property != 'NONE' and reference_value:
                    s_list[-1].references[reference_property] = reference_value

        return s_list

    def get_global(self) -> dict:
        """get_global - pobiera z arkusza definicje referencji globalnych
           (możliwa tylko jedna referencja globalna dla każdego arkusza,
           jeżeli jest więcej to uwzględni tylko ostatnią). Referencje globalne
           dotyczą w praktyce tylko arkuszy P_statements i Q_statements
        """
        global GLOBAL_REFERENCE

        for row in self.globals.iter_rows(2, self.globals.max_row):
            g_sheet = row[self.globals_columns["Sheet"]].value
            g_property = row[self.globals_columns["Reference_property"]].value
            g_value = row[self.globals_columns["Reference_value"]].value
            if g_value.startswith("http") and g_value.endswith("/"):
                g_value = g_value[:-1]
            GLOBAL_REFERENCE[g_sheet] = (g_property, g_value)


class WDHProperty:
    """Klasa dla właściwości (property)"""

    def __init__(
        self,
        label_en: str = "",
        description_en: str = "",
        datatype: str = "",
        label_pl: str = "",
        description_pl: str = "",
        wiki_id: str = "",
        inverse_property: str = "",
    ):
        self.label_en = label_en
        self.description_en = description_en
        self.datatype = datatype
        self.label_pl = label_pl
        self.description_pl = description_pl
        self.wiki_id = wiki_id
        self.inverse_property = inverse_property

    @property
    def label_en(self) -> str:
        """get label_en"""
        return self._label_en

    @label_en.setter
    def label_en(self, value: str):
        """set label_en"""
        if value:
            self._label_en = value.strip()
        else:
            self._label_en = ""

    @property
    def description_en(self) -> str:
        """get description_en"""
        return self._description_en

    @description_en.setter
    def description_en(self, value: str):
        """set description_en"""
        if value:
            self._description_en = value.strip()
        else:
            self._description_en = ""

    @property
    def datatype(self) -> str:
        """get datatype"""
        return self._datatype

    @datatype.setter
    def datatype(self, value: str):
        """set datatype"""
        if value:
            if value == "item":
                value = "wikibase-item"
            elif value == "property":
                value = "wikibase-property"
            elif value == "external identifier":
                value = "external-id"
            elif value == "URL":
                value = "url"
            elif value == "monolingual text":
                value = "monolingualtext"
            elif value == "geographic coordinates":
                value = "globe-coordinate"
            elif value == "point in time":
                value = "time"

            self._datatype = value.strip()
        else:
            self._datatype = ""

    @property
    def label_pl(self) -> str:
        """get label_pl"""
        return self._label_pl

    @label_pl.setter
    def label_pl(self, value: str):
        """set label_pl"""
        if value:
            self._label_pl = value.strip()
        else:
            self._label_pl = ""

    @property
    def description_pl(self) -> str:
        """get description_pl"""
        return self._description_pl

    @description_pl.setter
    def description_pl(self, value: str):
        """set description_pl"""
        if value:
            self._description_pl = value.strip()
        else:
            self._description_pl = ""

    @property
    def wiki_id(self) -> str:
        """get wiki_id"""
        return self._wiki_id

    @wiki_id.setter
    def wiki_id(self, value: str):
        """set wiki_id"""
        if value:
            self._wiki_id = value.strip()
        else:
            self._wiki_id = ""

    @property
    def inverse_property(self) -> str:
        """get inverse_property"""
        return self._inverse_property

    @inverse_property.setter
    def inverse_property(self, value: str):
        """set inverse_property"""
        if value:
            self._inverse_property = value.strip()
        else:
            self._inverse_property = ""

    def write_to_wikibase(self):
        """zapis właściwości w instancji wikibase"""
        pass


class WDHStatementProperty:
    """Klasa dla deklaracji (statement) dla właściwości"""

    def __init__(
        self,
        label_en: str = "",
        statement_property: str = "",
        statement_value: str = "",
        reference_property: str = "",
        reference_value: str = "",
    ):
        self.label_en = label_en
        self.statement_property = statement_property
        self.statement_value = statement_value
        self.references = {}
        if reference_property and reference_value:
            self.references[reference_property.strip()] = reference_value.strip()
        self.sheet_name = ""
        self.additional_references = {}

    @property
    def label_en(self) -> str:
        """getter: label_en"""
        return self._label_en

    @label_en.setter
    def label_en(self, value: str):
        """setter: label_en"""
        if value:
            self._label_en = value.strip()
        else:
            self._label_en = ""

    @property
    def statement_property(self) -> str:
        """get statement_property"""
        return self._statement_property

    @statement_property.setter
    def statement_property(self, value: str):
        """set statement_property"""
        if value:
            self._statement_property = value.strip()
        else:
            self._statement_property = ""

    @property
    def statement_value(self):
        """get statement_value"""
        return self._statement_value

    @statement_value.setter
    def statement_value(self, value: str):
        """set statement_value"""
        if value:
            self._statement_value = value.strip()
        else:
            self._statement_value = ""

    # @property
    # def reference_property(self):
    #     """get reference_property"""
    #     return self._reference_property

    # @reference_property.setter
    # def reference_property(self, value: str):
    #     """set reference_property"""
    #     if value:
    #         self._reference_property = value.strip()
    #     else:
    #         self._reference_property = ""

    # @property
    # def reference_value(self):
    #     """gettet: reference_value"""
    #     return self._reference_value

    # @reference_value.setter
    # def reference_value(self, value: str):
    #     """setter: reference_value"""
    #     if value:
    #         self._reference_value = value.strip()
    #     else:
    #         self._reference_property = ""

    def write_to_wikibase(self):
        """zapis deklaracji w instancji wikibase"""
        pass


class WDHItem:
    """Klasa dla elementu (item)"""

    def __init__(
        self,
        label_en: str = "",
        description_en: str = "",
        label_pl: str = "",
        description_pl: str = "",
        wiki_id: str = "",
    ):
        self.label_en = label_en
        self.description_en = description_en
        self.label_pl = label_pl
        self.description_pl = description_pl
        self.wiki_id = wiki_id
        self.starts_at = ""
        self.ends_at = ""
        self.instance_of = ""
        self.purl_identifier = ""

    @property
    def label_en(self) -> str:
        """get label_en"""
        return self._label_en

    @label_en.setter
    def label_en(self, value: str):
        """set label_en"""
        if value:
            self._label_en = value.strip()
        else:
            self._label_en = ""

    @property
    def description_en(self) -> str:
        """get description_en"""
        return self._description_en

    @description_en.setter
    def description_en(self, value: str):
        """set description_en"""
        if value:
            self._description_en = value.strip()
        else:
            self._description_en = ""

    @property
    def label_pl(self) -> str:
        """get label_pl"""
        return self._label_pl

    @label_pl.setter
    def label_pl(self, value: str):
        """set label_pl"""
        if value:
            self._label_pl = value.strip()
        else:
            self._label_pl = ""

    @property
    def description_pl(self) -> str:
        """get description_pl"""
        return self._description_pl

    @description_pl.setter
    def description_pl(self, value: str):
        """set description_pl"""
        if value:
            self._description_pl = value.strip()
        else:
            self._description_pl = ""

    @property
    def wiki_id(self) -> str:
        """get wiki_id"""
        return self._wiki_id

    @wiki_id.setter
    def wiki_id(self, value: str):
        """set wiki_id"""
        if value:
            self._wiki_id = value.strip()
        else:
            self._wiki_id = ""

    @property
    def starts_at(self) -> str:
        """get StartsAt"""
        return self._starts_at

    @starts_at.setter
    def starts_at(self, value: str):
        """set starts_at
        format: +1501-00-00T00:00:00Z/9
        """
        pattern = r"\+\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z\/\d{1,2}"
        if value:
            value = value.strip()
            # jeżeli data podana z dniem będącym początkiem lub końcem roku
            # to chodziło o dokładność roczną
            if (len(value) == 10 and value.endswith("12-31")) or (
                len(value) == 10 and value.endswith("01-01")
            ):
                value = value[:4]

            if len(value) == 4:  # 1564
                self._starts_at = f"+{value}-00-00T00:00:00Z/9"
            elif len(value) == 7:  # 1564-10
                self._starts_at = f"+{value}-00T00:00:00Z/10"
            elif len(value) == 10:  # 1564-10-11
                self._starts_at = f"+{value}T00:00:00Z/11"
            else:
                match = re.search(pattern, value)
                if match:
                    self._starts_at = value
                else:
                    self._starts_at = ""
                    print(f"ERROR: nieznany format daty: {value}")
        else:
            self._starts_at = ""

    @property
    def ends_at(self) -> str:
        """get EndsAt"""
        return self._ends_at

    @ends_at.setter
    def ends_at(self, value: str):
        """set ends_at
        format: +1501-00-00T00:00:00Z/9
        """
        pattern = r"\+\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z\/\d{1,2}"
        if value:
            value = value.strip()
            # jeżeli data podana z dniem będącym początkiem lub końcem roku
            # to chodziło o dokładność roczną
            if (len(value) == 10 and value.endswith("12-31")) or (
                len(value) == 10 and value.endswith("01-01")
            ):
                value = value[:4]

            if len(value) == 4:  # 1564
                self._ends_at = f"+{value}-00-00T00:00:00Z/9"
            elif len(value) == 7:  # 1564-10
                self._ends_at = f"+{value}-00T00:00:00Z/10"
            elif len(value) == 10:  # 1564-10-11
                self._ends_at = f"+{value}T00:00:00Z/11"
            else:
                match = re.search(pattern, value)
                if match:
                    self._ends_at = value
                else:
                    self._ends_at = ""
                    print(f"ERROR: nieznany format daty: {value}")
        else:
            self._ends_at = ""

    @property
    def purl_identifier(self) -> str:
        """get Purl_identifier"""
        return self._purl_identifier

    @purl_identifier.setter
    def purl_identifier(self, value: str):
        """set Purl_identifier"""
        if value:
            self._purl_identifier = value.strip()
        else:
            self._purl_identifier = ""

    @property
    def instance_of(self) -> str:
        """get instance_of"""
        return self._instance_of

    @instance_of.setter
    def instance_of(self, value: str):
        """set instance_of"""
        if value:
            self._instance_of = value.strip()
        else:
            self._instance_of = ""

    def write_to_wikibase(self):
        """zapis elementu w instancji wikibase"""
        # jeżeli jest etykieta 'en'
        if self.label_en:
            search_item, search_id = element_search(
                self.label_en,
                "item",
                "en",
                description=self.description_en,
                strict=True,
                purl_id=self.purl_identifier,
            )
        # jeżeli brak etykiety 'en' ale jest 'pl'
        elif self.label_pl:
            print("Wyszukiwanie po polskiej etykiecie.")
            search_item, search_id = element_search(
                self.label_pl,
                "item",
                "pl",
                description=self.description_pl,
                strict=True,
            )

        item_is_changed = False
        # jeżeli znaleziono w wikibase
        if search_item:
            print(
                f"Item: '{self.label_en}' already exists: {search_id}, update mode enabled."
            )
            wd_item = wbi_core.ItemEngine(item_id=search_id)
            mode = "ZAKTUALIZOWANO: "
            # dla istniejących już elementów weryfikacja czy zmieniony opis
            if self.description_en:
                description_en = wd_item.get_description("en")
                if description_en == self.description_en:
                    print(
                        f'SKIP: element: {search_id} ({self.label_en}) posiada już dla języka: "en" opis: {self.description_en}'
                    )
                else:
                    wd_item.set_description(self.description_en, lang="en")
                    item_is_changed = True

            if self.description_pl:
                description_pl = wd_item.get_description("pl")
                if description_pl == self.description_pl:
                    print(
                        f'SKIP: element: {search_id} ({self.label_en}) posiada już dla języka: "pl" opis: {self.description_pl}'
                    )
                else:
                    wd_item.set_description(self.description_pl, lang="pl")
                    item_is_changed = True

        # jeżeli nie znaleziono w wikibase
        else:
            wd_item = wbi_core.ItemEngine(new_item=True)
            mode = "DODANO: "
            # tylko dla nowych jest ustawiania en i pl etykieta oraz opisy
            wd_item.set_label(self.label_en, lang="en")
            if self.description_en:
                wd_item.set_description(self.description_en, lang="en")
            wd_item.set_label(self.label_pl, lang="pl")
            if self.description_pl:
                wd_item.set_description(self.description_pl, lang="pl")

        wiki_dane = None
        if self.wiki_id:
            skip_wiki_id = False
            if wikibase_prop.wiki_id == "" or wikibase_prop.wiki_url == "":
                wikibase_prop.get_wiki_properties()
            if search_item:
                if has_statement(search_id, wikibase_prop.wiki_id, self.wiki_id):
                    print(
                        f"SKIP: element {search_id} ({self.label_en}) posiada deklarację: {wikibase_prop.wiki_id} o wartości: {self.wiki_id}"
                    )
                    skip_wiki_id = True

            if not skip_wiki_id:
                item_is_changed = True
                url = f"https://www.wikidata.org/wiki/{self.wiki_id}"
                references = [
                    [
                        wbi_datatype.Url(
                            value=url, prop_nr=wikibase_prop.wiki_url, is_reference=True
                        )
                    ]
                ]
                wiki_dane = wbi_datatype.ExternalID(
                    value=self.wiki_id,
                    prop_nr=wikibase_prop.wiki_id,
                    references=references,
                )

        # obsługa opcjonalnych kolumn StatsAt, EndsAt, Instance of, Purl_identifier
        wiki_starts = wiki_ends = wiki_instance = wiki_purl = None

        # StartsAt
        if self.starts_at:
            skip_starts_at = False
            res, start_qid = find_name_qid("starts at", "property")
            if res:
                if search_item:
                    if has_statement(search_id, start_qid, self.starts_at):
                        print(
                            f"SKIP: element {search_id} ({self.label_en}) posiada deklarację: {start_qid} o wartości: {self.starts_at}"
                        )
                        skip_starts_at = True

                if not skip_starts_at:
                    tmp = self.starts_at.split("/")
                    if len(tmp) == 2:
                        time_value = tmp[0]
                        precision = int(tmp[1])
                        wiki_starts = wbi_datatype.Time(
                            time_value,
                            prop_nr=start_qid,
                            precision=precision,
                            is_reference=False,
                            references=None,
                            is_qualifier=False,
                            qualifiers=None,
                        )
                        item_is_changed = True
                        if mode == 'ZAKTUALIZOWANO: ':
                            print(
                                f"DODANO deklarację: 'starts at' ({start_qid}) o wartości: {self.starts_at}"
                            )

            else:
                print(
                    "ERROR: nie znaleziono właściwości 'starts at' w instancji Wikibase."
                )

        # EndsAt
        if self.ends_at:
            skip_ends_at = False
            res, end_qid = find_name_qid("ends at", "property")
            if res:
                if search_item:
                    if has_statement(search_id, end_qid, self.ends_at):
                        print(
                            f"SKIP: element {search_id} ({self.label_en}) posiada deklarację: {end_qid} o wartości: {self.ends_at}"
                        )
                        skip_ends_at = True

                if not skip_ends_at:
                    tmp = self.ends_at.split("/")
                    if len(tmp) == 2:
                        time_value = tmp[0]
                        precision = int(tmp[1])
                        wiki_ends = wbi_datatype.Time(
                            time_value,
                            prop_nr=end_qid,
                            precision=precision,
                            is_reference=False,
                            references=None,
                            is_qualifier=False,
                            qualifiers=None,
                        )
                        item_is_changed = True
                        if mode == 'ZAKTUALIZOWANO: ':
                            print(
                                f"DODANO deklarację: 'ends at' ({end_qid}) o wartości: {self.ends_at}"
                            )

            else:
                print(
                    "ERROR: nie znaleziono właściwości 'ends at' w instancji Wkibase."
                )

        # Instance of
        if self.instance_of:
            skip_instance = False
            res, instance_value = find_name_qid(self.instance_of, "item", strict=True)
            if res:
                res, instance_qid = find_name_qid("instance of", "property")
                if res:
                    if search_item:
                        if has_statement(search_id, instance_qid, instance_value):
                            print(
                                f"SKIP: element {search_id} ({self.label_en}) posiada deklarację: {instance_qid} o wartości: {self.instance_of}"
                            )
                            skip_instance = True

                    if not skip_instance:
                        wiki_instance = wbi_datatype.ItemID(
                            value=instance_value,
                            prop_nr=instance_qid,
                            is_reference=False,
                            references=None,
                            is_qualifier=False,
                            qualifiers=None,
                        )
                        item_is_changed = True
                        if mode == 'ZAKTUALIZOWANO: ':
                            print(
                                f"DODANO deklarację: 'instance of' ({instance_qid}) o wartości: {instance_value}"
                            )
                else:
                    print(
                        "ERROR: nie znaleziono właściwości 'instance of' w instancji Wkibase."
                    )
            else:
                print(
                    f"ERROR: nie znaleziono symbolu Q dla wartości deklaracji instance_of: {instance_value}"
                )

        # Purl_identifier
        if self.purl_identifier:
            #print("Podano Purl")
            skip_purl_identifier = False
            res, purl_qid = find_name_qid("purl identifier", "property")
            if res:
                if search_item:
                    if has_statement(search_id, purl_qid, self.purl_identifier):
                        print(
                            f"SKIP: element {search_id} ({self.label_en}) posiada deklarację: {purl_qid} o wartości: {self.purl_identifier}"
                        )
                        skip_purl_identifier = True

                if not skip_purl_identifier:
                    wiki_purl = wbi_datatype.ExternalID(
                        value=self.purl_identifier, prop_nr=purl_qid, references=None
                    )
                    item_is_changed = True
                    if mode == 'ZAKTUALIZOWANO: ':
                        print(
                            f"DODANO deklarację: 'purl identifier' ({purl_qid}) o wartości: {self.purl_identifier}"
                        )

            else:
                print(
                    "ERROR: nie znaleziono właściwości 'purl identifier' w instancji Wkibase."
                )

        # zapis w Wikibase jeżeli nowy element lub zmiany dla elementu
        if not search_item or item_is_changed:
            try:
                if WIKIBASE_WRITE:
                    new_id = wd_item.write(login_instance, entity_type="item")
                else:
                    new_id = "TEST"

                if search_item:
                    new_id = search_id

                # zapis id nowego/modyfikowaneg item
                if self.label_en:
                    GLOBAL_ITEM[self.label_en + "/" + self.label_pl + f" ({self.purl_identifier})"] = new_id
                else:
                    GLOBAL_ITEM["-/" + self.label_pl + f" ({self.purl_identifier})"] = new_id

                # deklaracje dla elementu
                data = []
                if wiki_dane:
                    data.append(wiki_dane)
                if wiki_starts:
                    data.append(wiki_starts)
                if wiki_ends:
                    data.append(wiki_ends)
                if wiki_instance:
                    data.append(wiki_instance)
                if wiki_purl:
                    data.append(wiki_purl)

                if data and WIKIBASE_WRITE:
                    wd_statement = wbi_core.ItemEngine(
                        item_id=new_id, data=data, debug=False
                    )
                    wd_statement.write(login_instance, entity_type="item")

                if mode == 'ZAKTUALIZOWANO: ':
                    print(f"ZAKTUALIZOWANO element: {new_id} ({self.label_en}/{self.label_pl})")
                else:
                    print(f"DODANO element: {new_id} ({self.label_en}/{self.label_pl})")
            except (MWApiError, KeyError) as error_add_element:
                print(f"ERROR: {self.label_en} ({error_add_element.error_msg})")
        # jeżeli nie nowy element i nie ma zmian do zapisu
        else:
            if self.label_en:
                GLOBAL_ITEM[self.label_en + "/" + self.label_pl + f" ({self.purl_identifier})"] = search_id
            else:
                GLOBAL_ITEM["-/" + self.label_pl + f" ({self.purl_identifier})"] = search_id


class WDHStatementItem:
    """Klasa dla deklaracji (statement) dla elementów"""

    def __init__(
        self,
        label_en: str = "",
        statement_property: str = "",
        statement_value: str = "",
        qualifier: str = "",
        qualifier_value: str = "",
        reference_property: str = "",
        reference_value: str = "",
    ):
        self.label_en = label_en
        self.statement_property = statement_property
        self.statement_value = statement_value
        self.qualifiers = {}
        if qualifier and qualifier_value:
            self.qualifiers[qualifier.strip()] = qualifier_value.strip()
        self.sheet_name = ""
        self.references = {}
        # jeżeli identyfikator/nazwa właściwości referencji = NONE to znaczy że nie ma referencji lokalnej
        if reference_property and reference_property!= 'NONE' and reference_value:
            self.references[reference_property.strip()] = reference_value.strip()
        self.additional_references = {}

    @property
    def label_en(self) -> str:
        """getter: label_en"""
        return self._label_en

    @label_en.setter
    def label_en(self, value: str):
        """setter: label_en"""
        if value:
            self._label_en = value.strip()
        else:
            self._label_en = ""

    @property
    def statement_property(self) -> str:
        """get statement_property"""
        return self._statement_property

    @statement_property.setter
    def statement_property(self, value: str):
        """set statement_property"""
        if value:
            self._statement_property = value.strip()
        else:
            self._statement_property = ""

    @property
    def statement_value(self):
        """get statement_value"""
        return self._statement_value

    @statement_value.setter
    def statement_value(self, value: str):
        """set statement_value"""
        if value:
            if isinstance(value, int):
                value = str(value)
            elif isinstance(value, float):
                value = str(value)
                if value.endswith('.0'):
                    value = value[:-2]
            elif not isinstance(value, str):
                value = str(value)
            self._statement_value = value.strip()
        else:
            self._statement_value = ""

    def write_to_wikibase(self):
        """zapis deklaracji dla elementu w instancji wikibase
        także zapis aliasu, opisu, dodatkowej etykiety dla elementu - zależnie od wartości
        self.statement_property
        """
        # print("KWALIFIKATORY: ", self.qualifiers)
        is_ok, p_id = find_name_qid(self.label_en, "item")
        if not is_ok:
            print("ERROR: " + f"brak elementu -> {self.label_en}")
            return

        # jeżeli to alias?
        if self.statement_property in (
            "Apl",
            "Aen",
            "Ade",
            "Aru",
            "Aes",
            "Afr",
            "Alt",
            "Alv",
            "Aet",
            "Anl",
            "Ait",
            "Ala",
            "Ahu",
            "Apt",
            "Auk",
            "Acs",
            "Ask",
            "Asl",
            "Aro",
            "Asv",
            "Afi",
            "Ahe",
            "Aund"
        ):
            try:
                wd_item = wbi_core.ItemEngine(item_id=p_id)
                lang = self.statement_property[1:]
                aliasy = wd_item.get_aliases(lang=lang)
                if self.statement_value in aliasy:
                    print(
                        f"SKIP: element: '{p_id}' ({self.label_en}) już posiada alias: '{self.statement_value}' dla języka: {lang}."
                    )
                else:
                    lang_len = -1 * len(lang)
                    wd_item.set_aliases(
                        self.statement_value, lang=self.statement_property[lang_len:]
                    )
                    if WIKIBASE_WRITE:
                        wd_item.write(login_instance, entity_type="item")
                    print(
                        f"ALIAS ADDED, item {p_id} ({self.label_en}): {self.statement_property} -> {self.statement_value}"
                    )

                # aliasy dla elementów powinny od razu stawać się także deklaracjami właściwości
                # 'stated as', ale tylko jeżeli są zdefiniowane dla arkusza globalne referencje
                # w arkuszu Globals
                if self.additional_references:
                    is_ok, prop_id = find_name_qid("stated as", "property")
                    if not is_ok:
                        print(
                            "ERROR: w instancji Wikibase brak właściwości -> stated as",
                        )
                        return

                    lang_id = self.statement_property[1:]
                    p_value = f'{lang_id}:"{self.statement_value}"'

                    # kontrola czy istnieje deklaracja o takiej wartości
                    if has_statement(p_id, prop_id, value_to_check=p_value):
                        print(
                            f"SKIP: element: '{p_id}' ({self.label_en}) już posiada deklarację: '{prop_id}' o wartości: {p_value}."
                        )
                        # weryfikacja i uzupełnianie referencje z referencji globalnych
                        update_references(login_instance, wd_item, p_id, prop_id, p_value, self.additional_references)

                    else:
                        # wartości deklaracji 'stated as' są dołączane do istniejących, nie zastępują poprzednich!
                        st_data = create_statement_data(
                            prop_id,
                            p_value,
                            self.references,
                            None,
                            add_ref_dict=self.additional_references,
                            if_exists="APPEND",
                        )
                        if st_data:
                            try:
                                data = [st_data]
                                wd_statement = wbi_core.ItemEngine(
                                    item_id=p_id, data=data, debug=False
                                )
                                if WIKIBASE_WRITE:
                                    wd_statement.write(
                                        login_instance, entity_type="item"
                                    )

                                print(
                                    f"STATEMENT ADDED, {p_id} ({self.label_en}): {prop_id} -> {p_value}"
                                )
                            except (MWApiError, KeyError, ValueError):
                                print(
                                    f"ERROR, {p_id} ({self.label_en}): {prop_id} -> {p_value}"
                                )
                        else:
                            print(
                                f"INVALID DATA, {p_id} ({self.label_en}): {prop_id} -> {p_value}"
                            )

            except (MWApiError, KeyError, ValueError) as e:
                print(
                    f"ERROR: item {p_id} ({self.label_en}): {self.statement_property} -> {self.statement_value}, {e}"
                )

        # jeżeli to etykieta
        elif self.statement_property in (
            "Lpl",
            "Len",
            "Lde",
            "Lru",
            "Les",
            "Lfr",
            "Llt",
            "Llv",
            "Let",
            "Lnl",
            "Lit",
            "Lla",
            "Lhu",
            "Lpt",
            "Luk",
            "Lcs",
            "Lsk",
            "Lsl",
            "Lro",
            "Lsv",
            "Lfi",
            "Lhe",
        ):
            try:
                wd_item = wbi_core.ItemEngine(item_id=p_id)
                lang = self.statement_property[1:]
                current_label = wd_item.get_label(lang)
                if self.statement_value == current_label:
                    print(
                        f"SKIP: element: '{p_id}' ({self.label_en}) już posiada etykietę: '{self.statement_value}' dla języka: {lang}."
                    )
                else:
                    wd_item.set_label(
                        self.statement_value,
                        lang=self.statement_property[-2:],
                        if_exists="REPLACE",
                    )
                    if WIKIBASE_WRITE:
                        wd_item.write(login_instance, entity_type="item")
                    print(
                        f"LABEL ADDED/MODIFIED, item {p_id} ({self.label_en}): {self.statement_property} -> {self.statement_value}"
                    )

            except (MWApiError, KeyError, ValueError):
                print(
                    f"ERROR: item {p_id} ({self.label_en}): {self.statement_property} -> {self.statement_value}"
                )

        # jeżeli to opis (description)
        elif self.statement_property in (
            "Dpl",
            "Den",
            "Dde",
            "Dru",
            "Des",
            "Dfr",
            "Dlt",
            "Dlv",
            "Det",
            "Dnl",
            "Dit",
            "Dla",
            "Dhu",
            "Dpt",
            "Duk",
            "Dcs",
            "Dsk",
            "Dsl",
            "Dro",
            "Dsv",
            "Dfi",
            "Dhe",
        ):
            try:
                wd_item = wbi_core.ItemEngine(item_id=p_id)
                lang = self.statement_property[1:]
                current_desc = wd_item.get_description(lang)
                if self.statement_value == current_desc:
                    print(
                        f"SKIP: element: '{p_id}' ({self.label_en}) już posiada opis: '{self.statement_value}' dla języka: {lang}."
                    )
                else:
                    wd_item.set_description(
                        self.statement_value,
                        lang=self.statement_property[-2:],
                        if_exists="REPLACE",
                    )
                    if WIKIBASE_WRITE:
                        wd_item.write(login_instance, entity_type="item")
                    print(
                        f"DESCRIPTION ADDED/MODIFIED, item {p_id} ({self.label_en}): {self.statement_property} -> {self.statement_value}"
                    )

            except (MWApiError, KeyError, ValueError):
                print(
                    f"ERROR: item {p_id} ({self.label_en}): {self.statement_property} -> {self.statement_value}"
                )

        # jeżeli to deklaracja?
        else:
            is_ok, prop_id = find_name_qid(self.statement_property, "property")
            if not is_ok:
                print(
                    f"ERROR: w instancji wikibase brak właściwości -> {self.statement_property}"
                )
                return

            if self.qualifiers:
                # zmiana nazwy kwalifikatora na jego Q
                tmp = {}
                for q_key, value in self.qualifiers.items():
                    is_ok, qualifier_id = find_name_qid(q_key, "property")
                    if not is_ok:
                        print(
                            f"ERROR: w instancji Wikibase brak właściwości -> {q_key}",
                        )
                        return

                    tmp[qualifier_id] = value
                    # modyfikacja wartości jeżeli to typ time (point in time)
                    if get_property_type(qualifier_id) == "time":
                        tmp[qualifier_id] = prepare_datetime(value)

                self.qualifiers = tmp

            # tu obsługa specyficznych typów właściwości: item/property wartość
            # wprowadzana jako deklaracją powinna być symbolem P lub Q
            prop_type = get_property_type(prop_id)
            if prop_type == "wikibase-item":
                is_ok, p_value = find_name_qid(self.statement_value, "item")
                if not is_ok:
                    print(
                        f"ERROR: w instancji Wikibase brak elementu -> {self.statement_value} będącego wartością -> {self.statement_property}",
                    )
                    return
            elif prop_type == "wikibase-property":
                is_ok, p_value = find_name_qid(self.statement_value, "property")
                if not is_ok:
                    print(
                        f"ERROR: w instancji Wikibase brak właściwości -> {self.statement_value} będącej wartością -> {self.statement_property}",
                    )
                    return
            else:
                p_value = self.statement_value

            # tu podobna obsługa j.w. ale tym razem dla dla kwalifikatorów
            tmp = {}
            for key, value in self.qualifiers.items():
                qualifier_type = get_property_type(key)
                if qualifier_type == "wikibase-item":
                    is_ok, q_value = find_name_qid(value, "item")
                    if not is_ok:
                        print(
                            f"ERROR: w instancji Wikibase brak elementu -> {value} będącego wartością kwalifikatora -> {key}",
                        )
                        return
                elif qualifier_type == "wikibase-property":
                    is_ok, q_value = find_name_qid(value, "property")
                    if not is_ok:
                        print(
                            f"ERROR: brak właściwości -> {value} będącej wartością kwalifikatora -> {key}",
                        )
                        return
                else:
                    q_value = value

                tmp[key] = q_value

            self.qualifiers = tmp

            # jeżeli właściwość deklaracji jest zewnętrznym identyfikatorem to nie dodajemy referencji
            # z globalnych referencji (o ile są)
            if prop_type == "external-id" and self.additional_references:
                self.additional_references = None
                print(
                    f"Pominięto referencję globalną dla deklaracji: {p_id}->{prop_id} typu external-id."
                )

            # kontrola czy istnieje deklaracja o tej wartości
            if has_statement(p_id, prop_id, value_to_check=p_value):
                print(
                    f"SKIP!: element: '{p_id}' ({self.label_en}) już posiada deklarację: '{prop_id}' o wartości: {p_value}."
                )

                # weryfikacja czy deklaracja ma referencje z referencji globalnych i ewentualne uzupełnienie
                wd_item = wbi_core.ItemEngine(item_id=p_id)
                update_references(login_instance, wd_item, p_id, prop_id, p_value, self.additional_references)
                # weryfikacja czy deklaracja ma referencję lokalną i ewentualne uzupełnienie
                update_references(login_instance, wd_item, p_id, prop_id, p_value, self.references)

                # weryfikacja czy deklaracja ma wszystkie kwalifikatory, a jeżeli nie to
                # uzupełnianie kwalifikatorów
                if self.qualifiers:
                    q_list = get_qualifiers(wd_item, prop_id, p_value)
                    #print(q_list)
                    for qualifier_key, qualifier_value in self.qualifiers.items():
                        qw_exists = check_if_qw_exists(
                            q_list, qualifier_key, qualifier_value
                        )
                        if not qw_exists:
                            if WIKIBASE_WRITE:
                                clm_id = find_claim_id(wd_item, prop_id, p_value)
                                if clm_id:
                                    add_result = add_qualifier(
                                        login_instance,
                                        clm_id,
                                        qualifier_key,
                                        qualifier_value,
                                    )
                                    #print('Uzupełnianie kwalifikatorów', add_result)
                                else:
                                    print(
                                        f"ERROR: nie znaleziono GUID deklaracji {prop_id} o wartości {p_value}"
                                    )
                        else:
                            print('Kwalifikator istnieje:', qualifier_key, qualifier_value )
                else:
                    print('Brak kwalifikatorów do uzupełnienia.')

            # jeżeli nie ma deklaracji to jest dodawana nowa
            else:
                st_data = create_statement_data(
                    prop_id,
                    p_value,
                    self.references,
                    self.qualifiers,
                    add_ref_dict=self.additional_references,
                    if_exists="APPEND",
                )
                if st_data:
                    try:
                        data = [st_data]
                        wd_statement = wbi_core.ItemEngine(
                            item_id=p_id, data=data, debug=False
                        )
                        if WIKIBASE_WRITE:
                            wd_statement.write(login_instance, entity_type="item")
                        print(
                            f"STATEMENT ADDED, {p_id} ({self.label_en}): {prop_id} -> {self.statement_value}"
                        )
                    except (MWApiError, KeyError, ValueError) as err_wiki:
                        print(
                            f"ERROR, {p_id} ({self.label_en}): {prop_id} -> {self.statement_value} {err_wiki.error_msg}"
                        )
                else:
                    print(
                        f"INVALID DATA, {p_id} ({self.label_en}): {prop_id} -> {self.statement_value}"
                    )


# --- funkcje ---

#def printlog(value: str):
#    """ zapis w logu"""

def add_property(p_dane: WDHProperty) -> tuple:
    """
    funkcja dodaje nową właściwość
    zwraca tuple: (True/False, ID/ERROR)
    """

    # test czy właściwość już nie istnieje
    search_property, search_id = element_search(p_dane.label_en, "property", "en", strict=True)
    if search_property:
        print(
            f"Property: '{p_dane.label_en}' already exists: {search_id}, update mode."
        )
        wd_item = wbi_core.ItemEngine(item_id=search_id)
        # tryb aktualizacji, jeżeli były dane do zaktualizowania zmienna aktualizacja == True
        mode = "ZAKTUALIZOWANO właściwość:"
        aktualizacja = False
        description_en = wd_item.get_description("en")
        if description_en == p_dane.description_en:
            print(
                f'SKIP: właściwość: {search_id} ({p_dane.label_en}) posiada już opis w języku: "en" o wartości: {description_en}'
            )
        else:
            aktualizacja = True
            wd_item.set_description(p_dane.description_en, lang="en")

        description_pl = wd_item.get_description("pl")
        if description_pl == p_dane.description_pl:
            print(
                f'SKIP: właściwość: {search_id} ({p_dane.label_en}) posiada już opis w języku: "pl" o wartości: {description_pl}'
            )
        else:
            aktualizacja = True
            wd_item.set_description(p_dane.description_pl, lang="pl")
    else:
        #print("New property")
        wd_item = wbi_core.ItemEngine(new_item=True)
        mode = "DODANO właściwość:"
        # etykiety i opisy
        wd_item.set_label(p_dane.label_en, lang="en")
        wd_item.set_description(p_dane.description_en, lang="en")
        if p_dane.label_pl:
            wd_item.set_label(p_dane.label_pl, lang="pl")
        if p_dane.description_pl:
            wd_item.set_description(p_dane.description_pl, lang="pl")

    # Wikidata ID i Wikidata URL (reference URL)
    wiki_dane = None
    if p_dane.wiki_id:
        if wikibase_prop.wiki_id == "" or wikibase_prop.wiki_url == "":
            wikibase_prop.get_wiki_properties()

        if search_property and has_statement(
            search_id, wikibase_prop.wiki_id, p_dane.wiki_id
        ):
            print(
                f"SKIP: właściwość: {search_id} ({p_dane.label_en}) posiada już deklarację: {wikibase_prop.wiki_id} o wartości: {p_dane.wiki_id}"
            )
        else:
            aktualizacja = True
            url = f"https://www.wikidata.org/wiki/Property:{p_dane.wiki_id}"
            references = [
                [
                    wbi_datatype.Url(
                        value=url, prop_nr=wikibase_prop.wiki_url, is_reference=True
                    )
                ]
            ]
            wiki_dane = wbi_datatype.ExternalID(
                value=p_dane.wiki_id,
                prop_nr=wikibase_prop.wiki_id,
                references=references,
            )

    # odwrotność właściwości
    inverse_dane = None
    if p_dane.inverse_property:
        if wikibase_prop.inverse == "":
            wikibase_prop.get_wiki_properties()
        if search_property and has_statement(
            search_id, wikibase_prop.inverse, p_dane.inverse_property
        ):
            print(
                f"SKIP: właściwość: {search_id} ({p_dane.label_en}) posiada już deklarację: {wikibase_prop.inverse} o wartości: {p_dane.inverse_property}"
            )
        else:
            aktualizacja = True
            search_inverse, inv_pid = element_search(
                p_dane.inverse_property, "property", "en"
            )
            if search_inverse and wikibase_prop.inverse != "":
                inverse_dane = wbi_datatype.Property(
                    value=inv_pid, prop_nr=wikibase_prop.inverse
                )

    # typy danych dla property: 'string', 'wikibase-item', 'wikibase-property',
    # 'monolingualtext', 'external-id', 'quantity', 'time', 'geo-shape', 'url',
    # 'globe-coordinate'
    options = {"property_datatype": p_dane.datatype}

    try:
        if WIKIBASE_WRITE:
            p_new_id = wd_item.write(login_instance, entity_type="property", **options)
        else:
            p_new_id = "TEST"

        if search_property:
            p_new_id = search_id

        # zapis id nowej lub modyfikowanej właściwości
        GLOBAL_PROPERTY[p_dane.label_en + "/" + p_dane.label_pl] = p_new_id

        # deklaracje dla właściwości
        data = []
        if wiki_dane:
            data.append(wiki_dane)
        if inverse_dane:
            data.append(inverse_dane)

        if len(data) > 0:
            if WIKIBASE_WRITE:
                wd_statement = wbi_core.ItemEngine(
                    item_id=p_new_id, data=data, debug=False
                )
                wd_statement.write(login_instance, entity_type="property")

        # jeżeli dodano właściwość inverse_property do dla docelowej właściwości należy
        # dodać odwrotność: nową właściwość jako jej inverse_property
        if inverse_dane:
            if WIKIBASE_WRITE:
                inv_statement = WDHStatementProperty(
                    inv_pid, wikibase_prop.inverse, p_new_id
                )
                add_res, add_info = add_property_statement(inv_statement)
                if not add_res:
                    print(f"ADD_PROPERTY_STATEMENT: {add_info}")
        if mode.startswith('ZAKTUALIZOWANO'):
            if aktualizacja:
                add_result = (True, mode + " qid:" + p_new_id + f" ({p_dane.label_en})")
            else:
                add_result = (False, '')
        else:
            add_result = (True, mode + " qid:" + p_new_id + f" ({p_dane.label_en})")

    except (MWApiError, KeyError) as error_property:
        add_result = (False, f"ERROR: {error_property.error_msg}")

    return add_result


def find_name_qid(name: str, elem_type: str, strict: bool = False) -> tuple:
    """Funkcja sprawdza czy przekazany argument jest identyfikatorem właściwości/elementu
    jeżeli nie to szuka w wikibase właściwości/elementu o etykiecie (ang) równej argumentowi
    (jeżeli strict=True to dokładnie równej) i zwraca jej id
    """
    name = name.strip()
    output = (True, name)  # zakładamy, że w name jest id (np. P47)
    # ale jeżeli nie, to szukamy w wikibase

    # jeżeli szukana wartość name = 'somevalue' lub 'novalue' to zwraca True i wartość
    if name == "somevalue" or name == "novalue":
        return (True, name)

    if elem_type == "property":
        pattern = r"^P\d{1,9}$"
    elif elem_type == "item":
        pattern = r"^Q\d{1,9}$"

    match = re.search(pattern, name)
    if not match:
        # http://purl.org/ontohgis#administrative_system_1
        purl_pattern = r"https?:\/\/purl\.org\/"

        match = re.search(purl_pattern, name)
        # wyszukiwanie elementu z deklaracją 'purl identifier' o wartości równej
        # zmiennej name
        if match:
            f_result, purl_qid = find_name_qid("purl identifier", "property")
            if f_result:
                output = search_by_purl(purl_qid, name)
                if not output[0]:
                    output = (False, f"INVALID DATA, {elem_type}: {name}, {output[1]}")
            else:
                output = (False, f"ERROR: {purl_qid}")
        # zwykłe wyszukiwanie
        else:
            output = element_search(name, elem_type, "en", strict=strict)
            if not output[0]:
                output = (False, f"INVALID DATA, {elem_type}: {name}, {output[1]}")

    return output


def create_statement(
    prop: str,
    value: str,
    is_ref: bool = False,
    refs=None,
    is_qlf: bool = False,
    qlfs=None,
    if_exists: str = "REPLACE",
) -> Union[
    wbi_datatype.String,
    wbi_datatype.Property,
    wbi_datatype.ItemID,
    wbi_datatype.ExternalID,
    wbi_datatype.Url,
    wbi_datatype.Quantity,
    wbi_datatype.Time,
    wbi_datatype.GeoShape,
    wbi_datatype.GlobeCoordinate,
    wbi_datatype.MonolingualText,
]:
    """
    Funkcja tworzy obiekt będący deklaracją lub referencją lub kwalifikatorem
    """
    statement = None
    if value == "somevalue" or value == "novalue":
        snak_type = value
    else:
        snak_type = "value"

    res, property_nr = find_name_qid(prop, "property")
    if res:
        property_type = get_property_type(property_nr)
        if property_type == "string":
            if snak_type != "value":
                value = None
            statement = wbi_datatype.String(
                value=value,
                prop_nr=property_nr,
                is_reference=is_ref,
                references=refs,
                is_qualifier=is_qlf,
                qualifiers=qlfs,
                if_exists=if_exists,
                snak_type=snak_type,
            )
        elif property_type == "wikibase-item":
            if snak_type != "value":
                value_id = None
                res = True
            else:
                res, value_id = find_name_qid(value, "item")

            if res:
                statement = wbi_datatype.ItemID(
                    value=value_id,
                    prop_nr=property_nr,
                    is_reference=is_ref,
                    references=refs,
                    is_qualifier=is_qlf,
                    qualifiers=qlfs,
                    if_exists=if_exists,
                    snak_type=snak_type,
                )
        elif property_type == "wikibase-property":
            if snak_type != "value":
                value_id = None
                res = True
            else:
                res, value_id = find_name_qid(value, "property")

            if res:
                statement = wbi_datatype.Property(
                    value=value_id,
                    prop_nr=property_nr,
                    is_reference=is_ref,
                    references=refs,
                    is_qualifier=is_qlf,
                    qualifiers=qlfs,
                    if_exists=if_exists,
                    snak_type=snak_type,
                )
        elif property_type == "external-id":
            if snak_type != "value":
                value = None
            statement = wbi_datatype.ExternalID(
                value=value,
                prop_nr=property_nr,
                is_reference=is_ref,
                references=refs,
                is_qualifier=is_qlf,
                qualifiers=qlfs,
                if_exists=if_exists,
                snak_type=snak_type,
            )
        elif property_type == "url":
            if snak_type != "value":
                value = None
            statement = wbi_datatype.Url(
                value=value,
                prop_nr=property_nr,
                is_reference=is_ref,
                references=refs,
                is_qualifier=is_qlf,
                qualifiers=qlfs,
                if_exists=if_exists,
                snak_type=snak_type,
            )
        elif property_type == "monolingualtext":
            # zakładając że wartość monolingualtext jest zapisana w formie:
            # pl:"To jest tekst w języku polskim", a jeżeli brak przedrostka z kodem
            # języka to przyjmujemy 'en'
            if snak_type != "value":
                value = None
            if value and (value[2] == ":" or value[3] == ":"):
                if value[3] == ":":
                    lang_code_len = 3
                else:
                    lang_code_len = 2
                # jeżeli nietypowy cudzysłów w wartości z arkusza xlsx
                if "”" in value:
                    value = value.replace("”", '"')
                prop_lang = value[:lang_code_len]  #
                if value.startswith(f'{prop_lang}:"'):
                    value = value[len(prop_lang)+2:-1]  # bez cudzysłowów
                elif value.startswith(f'{prop_lang}: "'):
                    value = value[len(prop_lang)+3:-1]  # bez cudzysłowów
                else:
                    print(
                        f"ERROR: błędna zawartość dla wartości typu monoligualtext ({prop})."
                    )
            else:
                prop_lang = "en"
            statement = wbi_datatype.MonolingualText(
                text=value,
                prop_nr=property_nr,
                language=prop_lang,
                is_reference=is_ref,
                references=refs,
                is_qualifier=is_qlf,
                qualifiers=qlfs,
                if_exists=if_exists,
                snak_type=snak_type,
            )
        elif property_type == "quantity":
            if snak_type != "value":
                value = None
            statement = wbi_datatype.Quantity(
                quantity=value,
                prop_nr=property_nr,
                is_reference=is_ref,
                references=refs,
                is_qualifier=is_qlf,
                qualifiers=qlfs,
                if_exists=if_exists,
                snak_type=snak_type,
            )
        elif property_type == "time":
            # czas w formacie +1539-12-08T00:00:00Z/11, po slashu precyzja daty zgodnie
            # ze standardami wikibase 11 - dzień, 9 - rok
            if snak_type != "value":
                value = None

            if value:
                value = value.strip()
                if len(value) == len('2023-05-15 00:00:00') and value.endswith('00:00:00'):
                    value = value[:10]

                if "/" in value:
                    tmp = value.split("/")
                elif len(value) == 4:
                    tmp = []
                    tmp.append(f"+{value}-00-00T00:00:00Z")
                    tmp.append("9")
                elif len(value) == 10:
                    tmp = []
                    tmp.append(f"+{value}T00:00:00Z")
                    tmp.append("11")
                else:
                    print(f'ERROR: błędny format daty: "{value}"')
            else:
                tmp = [None, 11]

            if len(tmp) == 2:
                time_value = tmp[0]
                precision = int(tmp[1])
                statement = wbi_datatype.Time(
                    time_value,
                    prop_nr=property_nr,
                    precision=precision,
                    is_reference=is_ref,
                    references=refs,
                    is_qualifier=is_qlf,
                    qualifiers=qlfs,
                    if_exists=if_exists,
                    snak_type=snak_type,
                )
            else:
                print(f"ERROR: invalid value for time type: {value}.")
        elif property_type == "geo-shape":
            # to chyba oczekuje nazwy pliku mapy w wikimedia commons, nam się nie przyda?
            if snak_type != "value":
                value = None
            statement = wbi_datatype.GeoShape(
                value,
                prop_nr=property_nr,
                is_reference=is_ref,
                references=refs,
                is_qualifier=is_qlf,
                qualifiers=qlfs,
                if_exists=if_exists,
                snak_type=snak_type,
            )
        elif property_type == "globe-coordinate":
            # oczekuje na zapis w formacie: 51.2,20.1 opcjonalnie jeszcze ,0.1
            # czyli latitude, longitude (jako liczby dziesiętne), oraz precyzja, domyślnie 0.1
            # domyślny glob = Earth, ale można zmienić na Marsa
            # https://www.wikidata.org/wiki/Help:Data_type/pl#Globe_coordinate
            if snak_type != "value":
                value = None

            if value:
                tmp = value.split(",")
            else:
                tmp = [None, None, None]

            try:
                latitude = float(tmp[0])
                longitude = float(tmp[1])
                if len(tmp) > 2:
                    precision = float(tmp[2])
                else:
                    precision = 0.001666666666667
            except ValueError:
                print(f"ERROR: invalid value for globe-coordinate type: {value}.")
            else:
                statement = wbi_datatype.GlobeCoordinate(
                    latitude,
                    longitude,
                    precision,
                    prop_nr=property_nr,
                    is_reference=is_ref,
                    references=refs,
                    is_qualifier=is_qlf,
                    qualifiers=qlfs,
                    if_exists=if_exists,
                    snak_type=snak_type,
                )

    return statement


def create_references(
    ref_dict: dict, additional_ref_dict: dict = None, if_exists: str = "REPLACE"
) -> list:
    """Funkcja tworzy referencje z przekazanego słownika referencji, opcjonalnie
    może zostać przekazany drugi słownik referencji np. globalnych wówczas
    utworzny zostanie drugi blok referencji
    """
    if ref_dict:
        statements = []
        for key, value in ref_dict.items():
            statement = create_statement(
                key, value, is_ref=True, refs=None, if_exists=if_exists
            )
            if statement:
                statements.append(statement)
        new_references = [statements]
    else:
        new_references = None

    if additional_ref_dict:
        statements = []
        for key, value in additional_ref_dict.items():
            statement = create_statement(
                key, value, is_ref=True, refs=None, if_exists=if_exists
            )
            if statement:
                statements.append(statement)

        if new_references:
            new_references.append(statements)
        else:
            new_references = [statements]

    return new_references


def create_qualifiers(qlf_dict: dict, if_exists: str = "REPLACE") -> list:
    """Funkcja tworzy kwalifikatory"""
    new_qualifiers = []
    for key, value in qlf_dict.items():
        statement = create_statement(
            key, value, is_qlf=True, qlfs=None, if_exists=if_exists
        )
        if statement:
            new_qualifiers.append(statement)

    if len(new_qualifiers) == 0:
        new_qualifiers = None

    return new_qualifiers


def create_statement_data(
    prop: str,
    value: str,
    reference_dict: dict,
    qualifier_dict: dict,
    add_ref_dict: dict = None,
    if_exists: str = "REPLACE",
) -> Union[
    wbi_datatype.String,
    wbi_datatype.Property,
    wbi_datatype.ItemID,
    wbi_datatype.ExternalID,
    wbi_datatype.Url,
    wbi_datatype.Quantity,
    wbi_datatype.Time,
    wbi_datatype.GeoShape,
    wbi_datatype.GlobeCoordinate,
    wbi_datatype.MonolingualText,
]:
    """
    Funkcja tworzy dane deklaracji z opcjonalnymy referencjami
    """
    # referencje i kwalifikatory z domyślną wartością if_exists = 'REPLACE'
    references = None
    if reference_dict or add_ref_dict:
        references = create_references(reference_dict, add_ref_dict)

    qualifiers = None
    if qualifier_dict:
        qualifiers = create_qualifiers(qualifier_dict)

    output_data = create_statement(
        prop,
        value,
        is_ref=False,
        refs=references,
        is_qlf=False,
        qlfs=qualifiers,
        if_exists=if_exists,
    )

    return output_data


def add_property_statement(s_item: WDHStatementProperty) -> tuple:
    """
    Funkcja dodaje deklaracje (statement) do właściwości
    Parametry:
        s_item - obiekt z deklaracją
    """
    # weryfikacja czy istnieje właściwość do której chcemy dodać deklarację
    is_ok, p_id = find_name_qid(s_item.label_en, "property", strict=True)
    if not is_ok:
        return (False, p_id)

    # jeżeli to alias?
    if s_item.statement_property in (
        "Apl",
        "Aen",
        "Ade",
        "Aru",
        "Aes",
        "Afr",
        "Alt",
        "Alv",
        "Aet",
        "Anl",
        "Ait",
        "Ala",
        "Ahu",
        "Apt",
        "Auk",
        "Acs",
        "Ask",
        "Asl",
        "Aro",
        "Asv",
        "Afi",
        "Ahe",
    ):
        try:
            wd_item = wbi_core.ItemEngine(item_id=p_id)
            lang = s_item.statement_property[1:]
            aliasy = wd_item.get_aliases(lang=lang)
            if s_item.statement_value in aliasy:
                add_result = (
                    False,
                    f"SKIP: właściwość: '{p_id} ({s_item.label_en})' już posiada alias: '{s_item.statement_value}' dla języka: {lang}.",
                )
            else:
                wd_item.set_aliases(
                    s_item.statement_value, lang=s_item.statement_property[-2:]
                )
                if WIKIBASE_WRITE:
                    wd_item.write(login_instance, entity_type="property")
                add_result = (
                    True,
                    f"ALIAS ADDED, właściwość: {p_id} ({s_item.label_en}): {s_item.statement_property} -> {s_item.statement_value}",
                )

        except (MWApiError, KeyError, ValueError) as error_alias:
            add_result = (
                False,
                f"ERROR: item {p_id} ({s_item.label_en}) {s_item.statement_property} -> {s_item.statement_value}, błąd: {error_alias.error_msg}",
            )

    # jeżeli to etykieta (ale nie można zmienić  etykiety pl/en!)
    elif s_item.statement_property in (
        "Lde",
        "Lru",
        "Les",
        "Lfr",
        "Llt",
        "Llv",
        "Let",
        "Lnl",
        "Lit",
        "Lla",
        "Lhu",
        "Lpt",
        "Luk",
        "Lcs",
        "Lsk",
        "Lsl",
        "Lro",
        "Lsv",
        "Lfi",
        "Lhe",
    ):
        try:
            wd_item = wbi_core.ItemEngine(item_id=p_id)
            lang = s_item.statement_property[1:]
            current_label = wd_item.get_label(lang)
            if s_item.statement_value == current_label:
                add_result = (
                    False,
                    f"SKIP: właściwość: '{p_id} ({s_item.label_en})' już posiada etykietę: '{s_item.statement_value}' dla języka: {lang}.",
                )
            else:
                wd_item.set_label(
                    s_item.statement_value,
                    lang=s_item.statement_property[-2:],
                    if_exists="REPLACE",
                )
                if WIKIBASE_WRITE:
                    wd_item.write(login_instance, entity_type="property")
                add_result = (
                    True,
                    f"LABEL ADDED/MODIFIED, właściwość {p_id} ({s_item.label_en}): {s_item.statement_property} -> {s_item.statement_value}",
                )

        except (MWApiError, KeyError, ValueError) as error_label:
            add_result = (
                False,
                f"ERROR: item {p_id} ({s_item.label_en}) {s_item.statement_property} -> {s_item.statement_value}, błąd: {error_label.error_msg}",
            )

    # jeżeli to opis (description)
    elif s_item.statement_property in (
        "Dpl",
        "Den",
        "Dde",
        "Dru",
        "Des",
        "Dfr",
        "Dlt",
        "Dlv",
        "Det",
        "Dnl",
        "Dit",
        "Dla",
        "Dhu",
        "Dpt",
        "Duk",
        "Dcs",
        "Dsk",
        "Dsl",
        "Dro",
        "Dsv",
        "Dfi",
        "Dhe",
    ):
        try:
            wd_item = wbi_core.ItemEngine(item_id=p_id)
            lang = s_item.statement_property[1:]
            current_desc = wd_item.get_description(lang)
            if s_item.statement_value == current_desc:
                add_result = (
                    False,
                    f"SKIP: właściwość: '{p_id}' ({s_item.label_en}) już posiada opis: '{s_item.statement_value}' dla języka: {lang}.",
                )
            else:
                wd_item.set_description(
                    s_item.statement_value,
                    lang=s_item.statement_property[-2:],
                    if_exists="REPLACE",
                )
                if WIKIBASE_WRITE:
                    wd_item.write(login_instance, entity_type="property")
                add_result = (
                    True,
                    f"DESCRIPTION ADDED/MODIFIED, item {p_id}: {s_item.statement_property} -> {s_item.statement_value}",
                )

        except (MWApiError, KeyError, ValueError) as error_desc:
            add_result = (
                False,
                f"ERROR: właściwość {p_id} ({s_item.label_en}) {s_item.statement_property} -> {s_item.statement_value}, błąd: {error_desc.error_msg}",
            )

    # jeżeli to deklaracja (statement) dla właściwości
    else:

        # weryfikacja czy istnieje właściwość która ma być deklaracją
        is_ok, prop_id = find_name_qid(
            s_item.statement_property, "property", strict=True
        )
        if not is_ok:
            return (False, prop_id)

        # tu obsługa specyficznych typów właściwości: item/property wartość
        # wprowadzana jako deklaracją powinna być symbolem P lub Q
        prop_type = get_property_type(prop_id)
        if prop_type == "wikibase-item":
            is_ok, value = find_name_qid(s_item.statement_value, "item")
            if not is_ok:
                return (False, value)
        elif prop_type == "wikibase-property":
            is_ok, value = find_name_qid(
                s_item.statement_value, "property", strict=True
            )
            if not is_ok:
                return (False, value)
        else:
            value = s_item.statement_value

        # kontrola czy istnieje deklaracja o takiej wartości
        if has_statement(p_id, prop_id, value_to_check=value):
            return (
                False,
                f"SKIP: właściwość: '{p_id}' ({s_item.label_en}) already has a statement: '{prop_id} with value: {value}'.",
            )

        # jeżeli właściwość jest zewnętrznym identyfiktorem to nie dodajemy referencji
        # z globalnych referencji
        if prop_type == "external-id":
            s_item.additional_references = None
            print(
                f"Pominięto referencję globalną dla deklaracji: {p_id}->{prop_id} typu external-id."
            )

        st_data = create_statement_data(
            s_item.statement_property,
            value,
            s_item.references,
            qualifier_dict=None,
            add_ref_dict=s_item.additional_references,
            if_exists="APPEND",
        )
        if st_data:
            try:
                data = [st_data]
                if WIKIBASE_WRITE:
                    wd_statement = wbi_core.ItemEngine(
                        item_id=p_id, data=data, debug=False
                    )
                    wd_statement.write(login_instance, entity_type="property")
                add_result = (
                    True,
                    f"STATEMENT ADDED, {p_id}: {prop_id} -> {s_item.statement_value}",
                )
            except (MWApiError, KeyError, ValueError) as error_statement:
                add_result = (
                    False,
                    f"ERROR, {p_id}: {prop_id} -> {s_item.statement_value}, błąd: {error_statement.error_msg}",
                )
        else:
            add_result = (
                False,
                f"INVALID DATA, {p_id}: {prop_id} -> {s_item.statement_value}",
            )

    return add_result


# def get_property_type(p_id: str) -> str:
#     """Funkcja zwraca typ właściwości na podstawie jej identyfikatora"""
#     params = {"action": "wbgetentities", "ids": p_id, "props": "datatype"}

#     search_results = mediawiki_api_call_helper(
#         data=params,
#         login=None,
#         mediawiki_api_url=None,
#         user_agent=None,
#         allow_anonymous=True,
#     )
#     data_type = None
#     if search_results:
#         data_type = search_results["entities"][p_id]["datatype"]

#     return data_type


def prepare_datetime(t_value: str) -> str:
    """Modyfikuje format zapisu daty do akceptowalnego przez wikibase"""
    t_value = t_value.strip()

    if len(t_value) == 4:
        t_value = f"+{t_value}-00-00T00:00:00Z/9"
    elif len(t_value) == 7:  # 1564-10
        t_value = f"+{t_value}-00T00:00:00Z/10"
    elif len(t_value) == 10:  # 1564-10-11
        t_value = f"+{t_value}T00:00:00Z/11"

    return t_value


def add_qualifier(login_data, claim_id: str, prop_nr: str, prop_value: str) -> bool:
    """dodaje kwalifikator do deklaracji, obecnie obsługuje tylko 'value',
    obsługa 'somevalue' do zrobienia
    """
    add_result = False
    prop_type = get_property_type(prop_nr)

    # token
    params = {"action": "query", "meta": "tokens"}

    try:
        results = mediawiki_api_call_helper(
            data=params,
            login=login_data,
            mediawiki_api_url=None,
            user_agent=None,
            allow_anonymous=False,
        )
        token = results["query"]["tokens"]["csrftoken"]
    except MWApiError as wbsetqualifier_error:
        print("Error add qualifier - token:", wbsetqualifier_error)
        return False

    snak_type = "value"

    if prop_type == "monolingualtext":
        snak = {"text": prop_value[4:-1], "language": prop_value[:2]}
    elif prop_type == "quantity":
        if not prop_value.startswith("-"):  # liczba dodatnia/ujemne
            prop_value = "+" + prop_value
        snak = {"amount": prop_value, "unit": "1"}
    elif prop_type == "string":
        # [{'snaktype': 'value', 'property': 'P232', 'datavalue': {'value': '17', 'type': 'string'}
        snak = prop_value
    elif prop_type == "wikibase-item":
        numeric_id = int(prop_value[1:])
        snak = {"entity-type": "item", "numeric-id": numeric_id, "id": prop_value}
    elif prop_type == "time":
        # print(prop_value)
        prop_value = prepare_datetime(prop_value)
        tmp_value = prop_value.split("/")
        time_value = tmp_value[0]
        time_precision = int(tmp_value[1])
        snak = {
            "time": time_value,
            "precision": time_precision,
            "before": 0,
            "after": 0,
            "timezone": 0,
            "calendarmodel": "http://www.wikidata.org/entity/Q1985727",
        }
    elif prop_type == "globe-coordinate":
        tmp_value = prop_value.split(",")
        latitude = float(tmp_value[0])
        longitude = float(tmp_value[1])
        snak = {
            "latitude": latitude,
            "longitude": longitude,
            "precision": 0.01,
            "globe": "http://www.wikidata.org/entity/Q2",
        }

    snak_encoded = json.dumps(snak)
    # print(snak_encoded)

    params = {
        "action": "wbsetqualifier",
        "claim": claim_id,
        "snaktype": snak_type,
        "property": prop_nr,
        "value": snak_encoded,
        "token": token,
        "bot": True,
    }

    try:
        results = mediawiki_api_call_helper(
            data=params,
            login=login_data,
            mediawiki_api_url=None,
            user_agent=None,
            allow_anonymous=False,
        )
        if results["success"] == 1:
            add_result = True
    except MWApiError as wbsetqualifier_error:
        print(
            f"Error add qualifier:\n claim_id: {claim_id}, prop_nr: {prop_nr}, snak: {snak_encoded}\n",
            wbsetqualifier_error,
        )

    return add_result


def add_reference(login_data, claim_id: str, prop_nr: str, prop_value: str) -> bool:
    """dodaje odnośnik do deklaracji"""
    add_result = False

    # token
    params = {"action": "query", "meta": "tokens"}

    try:
        results = mediawiki_api_call_helper(
            data=params,
            login=login_data,
            mediawiki_api_url=None,
            user_agent=None,
            allow_anonymous=False,
        )
        token = results["query"]["tokens"]["csrftoken"]
    except MWApiError as wbsetreference_error:
        print("Error (add reference - token):", wbsetreference_error)
        return False

    snak_type = "value"
    snak = {
        prop_nr: [
            {
                "snaktype": snak_type,
                "property": prop_nr,
                "datavalue": {"type": "string", "value": prop_value},
            }
        ]
    }
    snak_encoded = json.dumps(snak)

    params = {
        "action": "wbsetreference",
        "statement": claim_id,
        "snaks": snak_encoded,
        "token": token,
        "bot": True,
    }

    try:
        results = mediawiki_api_call_helper(
            data=params,
            login=login_data,
            mediawiki_api_url=None,
            user_agent=None,
            allow_anonymous=False,
        )
        if results["success"] == 1:
            add_result = True
    except MWApiError as wbsetreference_error:
        print(f"Error add reference - snak: \n{snak_encoded}\n{wbsetreference_error}")

    return add_result


def find_claim_id(wd_item_test, stat_prop_qid: str, stat_prop_value: str):
    """
    zwraca guid deklaracji lub pusty string
    """
    if not isinstance(stat_prop_value, str):
        stat_prop_value = str(stat_prop_value)

    # jeżeli wartość typu monolingualtext ale ze zbędną dodatkową spacją
    if stat_prop_value[2:5] == ': "':
        stat_prop_value = stat_prop_value[:2] + ':"' + stat_prop_value[5:]

    # jeżeli nieznana lub brak wartości
    if stat_prop_value == "somevalue" or stat_prop_value == "novalue":
        # print(stat_prop_value)
        stat_prop_value = None

    result_id = ""
    for statement in wd_item_test.statements:
        # print(statement)
        statement_value = statement.get_value()
        statement_prop_nr = statement.get_prop_nr()
        statement_type = statement.data_type
        statement_value = statement_value_fix(statement_value, statement_type)
        # print(stat_prop_qid, statement_value, stat_prop_value, statement_type)

        # czy znaleziono poszukiwaną deklarację
        if statement_prop_nr == stat_prop_qid and statement_value == stat_prop_value:
            result_id = statement.get_id()
            break

    return result_id


def verify_reference(
    wd_item_test,
    stat_prop_qid: str,
    stat_prop_value: str,
    g_ref_qid: str,
    g_ref_value: str,
):
    """weryfikacja czy globalna referencja jest przypisana do deklaracji"""
    ref_exists = False

    if stat_prop_value in ("somevalue", "novalue"):
        stat_prop_value = None
    elif not isinstance(stat_prop_value, str):
        stat_prop_value = str(stat_prop_value)

    # pętla po deklaracjach elementu
    for statement in wd_item_test.statements:
        statement_value = statement.get_value()
        statement_prop_nr = statement.get_prop_nr()
        statement_type = statement.data_type
        statement_value = statement_value_fix(statement_value, statement_type)

        # czy znaleziono poszukiwaną deklarację
        if statement_prop_nr == stat_prop_qid and statement_value == stat_prop_value:
            tmp_references = statement.get_references()
            for t_ref_blok in tmp_references:
                # print(t_ref_blok[0].get_prop_nr(), ' - ', g_ref_qid)
                # print(t_ref_blok[0].get_value(), ' - ', g_ref_value)
                if (
                    t_ref_blok[0].get_prop_nr() == g_ref_qid
                    and t_ref_blok[0].get_value() == g_ref_value
                ):
                    ref_exists = True
                    break

    return ref_exists


def monolingual_text_fix(text_value: str) -> str:
    """korekta wartości tesktowej jeżeli to wygląda na monolingual text"""
    if len(text_value) > 3:
        if text_value[2] == ":" and "”" in text_value:
            # jeżeli nietypowy cudzysłów w wartości z arkusza xlsx
            text_value = text_value.replace("”", '"')
        # jeżeli wartość typu monolingualtext ale ze zbędną dodatkową spacją
        if text_value[2:5] == ': "':
            text_value = text_value[:2] + ':"' + text_value[5:]

    return text_value


# def statement_value_fix(s_value, s_type) -> str:
#     """poprawia wartość pobraną z deklaracji właściwości"""
#     if s_value is None:
#         return s_value

#     if s_type == "monolingualtext":
#         s_value = s_value[1] + ':"' + s_value[0] + '"'
#     elif s_type == "quantity":
#         if isinstance(s_value, tuple):
#             s_value = s_value[0].replace("+", "").replace("-", "")
#         else:
#             s_value = str(s_value)
#     elif s_type == "globe-coordinate":
#         if isinstance(s_value, tuple):
#             s_value = str(s_value[0]) + "," + str(s_value[1])
#         else:
#             s_value = str(s_value)
#     elif s_type == "wikibase-item":
#         if isinstance(s_value, int):
#             s_value = str(s_value)
#         if not s_value.startswith("Q"):
#             s_value = "Q" + s_value
#     elif s_type == "time":
#         if isinstance(s_value, tuple):
#             s_value_time = s_value[0]
#             if s_value_time is None:
#                 s_value = None
#             elif isinstance(s_value_time, str):
#                 s_value = s_value_time + "/" + str(s_value[3])
#             else:
#                 print(f"ERROR: wartość typu time: {s_value}")
#     else:
#         if not isinstance(s_value, str):
#             s_value = str(s_value)

#     return s_value


def get_qualifiers(element: wbi_core.ItemEngine, prop_id, prop_value) -> list:
    """zwraca słownik z kwalifikatorami dla deklaracji"""
    qualifiers = []

    if prop_value in ("somevalue", "novalue"):
        prop_value = None

    for statement in element.statements:
        statement_value = statement.get_value()
        statement_type = statement.data_type
        statement_value = statement_value_fix(statement_value, statement_type)

        if statement.get_prop_nr() == prop_id and statement_value == prop_value:
            tmp_json = statement.get_json_representation()
            if "qualifiers" in tmp_json:
                lista = tmp_json["qualifiers"].keys()
                for q_key in lista:
                    for q_item in tmp_json["qualifiers"][q_key]:
                        # print(q_item)
                        qualifier_property = q_item["property"]
                        if "datavalue" not in q_item:
                            continue
                        qualifier_type = q_item["datavalue"]["type"]
                        if qualifier_type == "string":
                            qualifier_value = q_item["datavalue"]["value"]
                        elif qualifier_type == "monolingualtext":
                            tmp = q_item["datavalue"]["value"]
                            qualifier_value = tmp["language"] + ':"' + tmp["text"] + '"'
                        elif qualifier_type == "quantity":
                            tmp = q_item["datavalue"]["value"]
                            qualifier_value = tmp["amount"].replace("+", "")
                        elif qualifier_type == "wikibase-entityid":
                            qualifier_value = q_item["datavalue"]["value"]["id"]
                        elif qualifier_type == "time":
                            qualifier_value = q_item["datavalue"]["value"]["time"]
                            q_precision = q_item["datavalue"]["value"]["precision"]
                            qualifier_value += "/" + str(q_precision)
                        elif qualifier_type == "globecoordinate":
                            tmp = q_item["datavalue"]["value"]
                            qualifier_value = (
                                str(tmp["latitude"]) + "," + str(tmp["longitude"])
                            )

                        qualifiers.append((qualifier_property, qualifier_value))

    return qualifiers


def check_if_qw_exists(q_list, qualifier_property, qualifier_value) -> bool:
    """Funkcja weryfikuje czy podany kwalifikator - property i value jest
    w przekazanej liście kwalifikatorów bieżącej deklaracji
    """
    qw_result = False
    for t_prop, t_value in q_list:
        # print(t_prop, t_value, ' = ', qualifier_property, qualifier_value)
        if t_prop == qualifier_property and t_value == qualifier_value:
            qw_result = True
            break

    return qw_result


def update_references(my_login, my_item, p_id: str, prop_id: str, p_value: str, additional_references: dict):
    """
    weryfikacja czy deklaracja ma referencje z przekazanego słownika,
    a jeżeli nie to próba uzupełnienia
    """
    # jeżeli przekazany słownik jest pusty
    if not additional_references:
        return

    for (add_ref_prop, add_ref_value,) in additional_references.items():
        is_ok, add_ref_qid = find_name_qid(add_ref_prop, "property")
        test_ref_exists = verify_reference(
            my_item, prop_id, p_value, add_ref_qid, add_ref_value
        )
        # jeźeli brak to próba dodania referencji
        if not test_ref_exists:
            print(
                f"Nie znaleziono referencji: {add_ref_qid} ({add_ref_prop}) o wartości: {add_ref_value} w deklaracji {prop_id} dla elementu {p_id}"
            )
            if WIKIBASE_WRITE:
                clm_id = find_claim_id(my_item, prop_id, p_value)
                if clm_id:
                    if add_reference(
                        my_login,
                        clm_id,
                        add_ref_qid,
                        add_ref_value,
                    ):
                        print(
                            f"REFERENCE: do deklaracji {prop_id} (o wartości {p_value}) dodano referencję: {add_ref_qid} ({add_ref_prop}) o wartości {add_ref_value}"
                        )
                else:
                    print(
                        f"ERROR: nie znaleziono GUID deklaracji {prop_id} o wartości {p_value}"
                    )


def has_statement(pid_to_check: str, claim_to_check: str, value_to_check: str = ""):
    """
    Funkcja weryfikuje czy właściwość (property) lub element (item) ma już
    taką deklarację (statement), opcjonalnie - z podaną wartością
    """
    has_claim = False
    wb_prop = wbi_core.ItemEngine(item_id=pid_to_check)
    data_prop = wb_prop.get_json_representation()
    claims = data_prop["claims"]
    if claim_to_check in claims:
        if not value_to_check:
            has_claim = True
        else:
            lista = claims[claim_to_check]
            for item in lista:
                # print(item)
                if value_to_check in ("somevalue", "novalue"):
                    value = item["mainsnak"]["snaktype"]
                else:
                    value_json = item["mainsnak"]["datavalue"]["value"]

                    if (
                        "type" in item["mainsnak"]["datavalue"]
                        and item["mainsnak"]["datavalue"]["type"] == "string"
                    ):
                        value = value_json
                    elif "text" in value_json and "language" in value_json:
                        # jeżeli nietypowy cudzysłów w wartości z arkusza xlsx
                        if "”" in value_to_check:
                            value_to_check = value_to_check.replace("”", '"')
                        # jeżeli zbędna spacja w wartości monoligualtext
                        if value_to_check[2] == ":" and ': "' in value_to_check:
                            value_to_check = value_to_check.replace(': "', ':"')
                        value = f"{value_json['language']}:\"{value_json['text']}\""
                    elif "entity-type" in value_json:
                        value = value_json["id"]
                    elif "latitude" in value_json:
                        value = f"{value_json['latitude']},{value_json['longitude']}"
                    elif "time" in value_json:
                        value = f"{value_json['time']}/{value_json['precision']}"
                    elif "amount" in value_json:
                        value = value_json["amount"]
                        if value.startswith("+"):
                            value = value[1:]
                    else:
                        value = "???"  # jeszcze nie obsługiwany typ - do weryfikacji
                        print(f"{item} (jeszcze nie obsługiwany typ) - {value_to_check}")

                if value == value_to_check:
                    has_claim = True
                    break

    return has_claim


def create_inverse_statement(my_login_instance, qid: str, main_property: str, inverse_property: str, references: dict = None):
    """ funkcja sprawdza czy istnieje odwrotna właściwość dla elementu
     będącego wartością głównej właściwości i jeżeli nie taką tworzy
     """
    if has_statement(qid, main_property):
        print(f"Znaleziono właściwość {main_property}")
        wd_item = wbi_core.ItemEngine(item_id=qid)

        for statement in wd_item.statements:
            statement_prop = statement.get_prop_nr()

            if statement_prop == main_property:
                statement_value = statement.get_value()
                statement_type = get_property_type(statement_prop)
                if statement_type == 'wikibase-item':
                    statement_value = f'Q{statement_value}'

                # weryfikacja czy element będący wartością właściwości main_property
                # ma deklarację inverse_property z wartością Q badanego elementu
                # a jeżeli nie to uzupełnienie
                if not has_statement(statement_value, inverse_property, value_to_check=qid):
                    print(f"Element docelowy właściwości {main_property} nie posiada właściwości {inverse_property}, trwa uzupełnianie...")
                    data = []
                    statement = create_statement_data(
                        inverse_property,
                        qid,
                        references,
                        None,
                        add_ref_dict=None,
                        if_exists="APPEND",
                    )
                    if statement:
                        data.append(statement)

                    if data:
                        if WIKIBASE_WRITE:
                            try:
                                wd_item_update = wbi_core.ItemEngine(item_id=statement_value, data=data, debug=False)
                                wd_item_update.write(my_login_instance, entity_type='item')
                                print(f"Dodano do elementu {statement_value} deklarację: {inverse_property} -> {qid}")
                            except (MWApiError, KeyError, ValueError) as write_error:
                                print(f"ERROR: podczas dodawania do elementu {statement_value}: {inverse_property} -> {qid}: {write_error.error_msg}")
                        else:
                            print(f"Przygotowano dodanie do elementu {statement_value} deklaracji: {inverse_property} -> {qid}")


# --------------------------------- MAIN ---------------------------------------
if __name__ == "__main__":
    # pomiar czasu wykonania
    start_time = time.time()

    # login i hasło ze zmiennych środowiskowych
    env_path = Path(".") / ".env"
    load_dotenv(dotenv_path=env_path)

    # OAuth
    WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
    WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
    WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
    WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

    login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                         consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                         access_token=WIKIDARIAH_ACCESS_TOKEN,
                                         access_secret=WIKIDARIAH_ACCESS_SECRET,
                                         token_renew_period=14400)

    # podstawowe właściwości Wikibase
    wikibase_prop = BasicProp()

    # ustalenie nr podstawowych property (jeżeli są, jeżeli będą dodawane podczas
    # pracy skryptu, wartości zostaną podczytane pred pierwszym użyciem)
    wikibase_prop.get_wiki_properties()

    # dane z arkusza XLSX, wg ścieżki przekazanej argumentem z linii komend
    # lub według pliku konfiguracyjnego yaml, w którym zawarta jest ścieżka
    # oraz mapowanie nazw kolumn rezeczywistych <-> oczekiwanych
    # jeżeli argumentu nie przekazano, skrypt szuka pliku 'data/arkusz_import.xlsx'
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    else:
        filename = "/home/piotr/ihpan/wikihub_skrypty/data/autorzy_ontohgis_fix.xlsx"

    plik_xlsx = WDHSpreadsheet(filename)
    plik_xlsx.open()

    # globalne referencje
    plik_xlsx.get_global()

    # właściwośći
    dane = plik_xlsx.get_property_list()
    dane_count = len(dane)
    for index, wb_property in enumerate(dane, start=1):
        print(f"PROPERTY ({index}/{dane_count}): {wb_property.label_en}")
        result, info = add_property(wb_property)
        if result or info:
            print(f"{info}, {result}")

    # dodatkowe deklaracje dla właściwości
    dane = plik_xlsx.get_statement_list()
    dane_count = len(dane)
    for index, stm in enumerate(dane, start=1):
        print(
            f"PROPERTY (statements) ({index}/{dane_count}): {stm.label_en}, STATEMENT: {stm.statement_property}, VALUE: {stm.statement_value}"
        )
        result, info = add_property_statement(stm)
        print(f"result {info}")

    # elementy 'strukturalne' ('definicyjne')
    dane = plik_xlsx.get_item_list()
    unique_item_en = []
    unique_item_pl = []
    unique_error = False
    for wb_item in dane:
        # angielski
        if wb_item.label_en or wb_item.description_en:
            lbl_desc_en = wb_item.label_en + "|" + wb_item.description_en
            if lbl_desc_en in unique_item_en:
                print("ERROR: etykieta i opis w języku ang powtarzają się: ", lbl_desc_en)
                unique_error = True
            else:
                unique_item_en.append(lbl_desc_en)

        # polski
        if wb_item.label_pl or wb_item.description_pl:
            lbl_desc_pl = wb_item.label_pl + "|" + wb_item.description_pl
            if lbl_desc_pl in unique_item_pl:
                print("ERROR: etykieta i opis w języku pl powtarzają się: ", lbl_desc_pl)
                unique_error = True
            else:
                unique_item_pl.append(lbl_desc_pl)

    if unique_error:
        sys.exit(1)

    dane_count = len(dane)
    for index, wb_item in enumerate(dane, start=1):
        print(f"ITEM ({index}/{dane_count}): {wb_item.label_en}/{wb_item.label_pl}")
        wb_item.write_to_wikibase()

    # dodatkowe deklaracje dla elementów strukturalnych/definicyjnych
    dane = plik_xlsx.get_item_statement_list()
    dane_count = len(dane)
    for index, stm in enumerate(dane, start=1):
        print(
            f"ITEM ({index}/{dane_count}): {stm.label_en}, STATEMENT: {stm.statement_property}, VALUE: {stm.statement_value}"
        )
        stm.write_to_wikibase()

    # zapis list przetwarzanych właściwości i elementów
    with open("property_list.html", "w", encoding="utf-8") as f:
        f.write(
            '<html>\n<head>\n<meta charset="UTF-8">\n<title>Lista właściwości</title>\n</head>\n<body>\n<h2>Lista dodanych/uaktualnionych właściwości</h2>\n<p>\n'
        )
        numer = 1
        for property_label, property_qid in GLOBAL_PROPERTY.items():
            f.write(
                f'{numer}. {property_label} = <a href="https://prunus-208.man.poznan.pl/wiki/Property:{property_qid}">{property_qid}</a><br>\n'
            )
            numer += 1
        f.write("</body></html>\n")

    with open("item_list.html", "w", encoding="utf-8") as f:
        f.write(
            '<html>\n<head>\n<meta charset="UTF-8">\n<title>Lista elementów</title>\n</head>\n<body>\n<h2>Lista dodanych/uaktualnionych elementów</h2>\n<p>\n'
        )
        numer = 1
        for item_label, item_qid in GLOBAL_ITEM.items():
            f.write(
                f'{numer}. {item_label} = <a href="https://prunus-208.man.poznan.pl/wiki/Item:{item_qid}">{item_qid}</a><br>\n'
            )
            numer += 1
        f.write("</p></body></html>\n")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

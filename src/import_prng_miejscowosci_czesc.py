""" import miejscowosci z pliku PRNG_egzonimy_czesc.xlsx """
import os
import sys
import time
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikidariahtools import find_name_qid, element_search_adv, get_coord, write_or_exit
from property_import import create_statement_data, has_statement


# adresy wikibase
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

WIKIBASE_WRITE = True

# standardowe właściwości i elementy
ok, p_instance_of = find_name_qid('instance of', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'instance of' w instancji Wikibase")
    sys.exit(1)
ok, p_stated_as = find_name_qid('stated as', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'stated as' w instancji Wikibase")
    sys.exit(1)
ok, p_coordinate = find_name_qid('coordinate location', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'coordinate location' w instancji Wikibase")
    sys.exit(1)
ok, p_reference_url = find_name_qid('reference URL', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'reference URL' w instancji Wikibase")
    sys.exit(1)
ok, p_inflectional_form = find_name_qid('inflectional form', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'inflectional form' w instancji Wikibase")
    sys.exit(1)
ok, p_locative_form = find_name_qid('locative form', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'locative form' w instancji Wikibase")
    sys.exit(1)
ok, p_adjective_form = find_name_qid('adjective form', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'adjective form' w instancji Wikibase")
    sys.exit(1)
ok, p_located_in_string = find_name_qid('located in (string)', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'located in (string)' w instancji Wikibase")
    sys.exit(1)
ok, p_located_in_country = find_name_qid('located in country', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'located in country' w instancji Wikibase")
    sys.exit(1)
ok, p_located_in = find_name_qid('located in', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'located in' w instancji Wikibase")
    sys.exit(1)
ok, p_id_sdi = find_name_qid('id SDI', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'id SDI' w instancji Wikibase")
    sys.exit(1)
ok, p_part_of = find_name_qid('part of', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'part of' w instancji Wikibase")
    sys.exit(1)
ok, p_retrieved = find_name_qid('retrieved', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'retrieved' w instancji Wikibase")
    sys.exit(1)
ok, p_point_in_time = find_name_qid('point in time', 'property', strict=True)
if not ok:
    print("ERROR: brak właściwości 'point in time' w instancji Wikibase")
    sys.exit(1)

# elementy definicyjne
# symbol QID elementu definicyjnego 'administrative unit', w wersji testowej: 'Q79096'
ok, q_administrative_unit = find_name_qid('administrative unit', 'item', strict=True)
if not ok:
    print("ERROR: brak elementu 'administrative unit' w instancji Wikibase")
    sys.exit(1)

# symbol QID elementu definicyjnego 'other territories', w wersji testowej: 'Q86555'
ok, q_other_territories = find_name_qid('other territories', 'item', strict=True)
if not ok:
    print("ERROR: brak elementu 'other territories' w instancji Wikibase")
    sys.exit(1)

# symbol QID elementu definicyjnego 'country', w wersji testowej: 'Q86557'
ok, q_country = find_name_qid('country', 'item', strict=True)
if not ok:
    print("ERROR: brak elementu 'country' w instancji Wikibase")
    sys.exit(1)

# symbol QID elementu definicyjnego 'dependent territory', w wersji testowej 'Q86554'
ok, q_dependent_territory = find_name_qid('dependent territory', 'item', strict=True)
if not ok:
    print("ERROR: brak elementu 'dependent territory' w instancji Wikibase")
    sys.exit(1)

# symbol QID elementu definicyjnego 'region', w wersji testowej 'Q86556'
ok, q_region = find_name_qid('region', 'item', strict=True)
if not ok:
    print("ERROR: brak elementu 'region' w instancji Wikibase")
    sys.exit(1)

# symbol QID elementu definicyjnego 'human settlement', w wersji testowej 'Q543'
ok, q_human_settlement = find_name_qid('human settlement', 'item', strict=True)
if not ok:
    print("ERROR: brak elementu 'human settlement' w instancji Wikibase")
    sys.exit(1)

# kraje w których leżą części miejscowości - niezbędne do wyszukania miejscowości
# nadrzędnych z pola InfoDod
kraje = {}
kraje['Azerbejdżan'] = 'Q362471'
kraje['Białoruś'] = 'Q362427'
kraje['Bułgaria'] = 'Q362429'
kraje['Chiny'] = 'Q362476'
kraje['Czechy'] = 'Q362432'
kraje['Egipt'] = 'Q362524'
kraje['Estonia'] = 'Q362434'
kraje['Francja'] = 'Q362436'
kraje['Izrael'] = 'Q362484'
kraje['Litwa'] = 'Q362444'
kraje['Łotwa'] = 'Q362446'
kraje['Maroko'] = 'Q362544'
kraje['Rosja'] = 'Q362454'
kraje['Słowacja'] = 'Q362458'
kraje['Sri Lanka'] = 'Q362505'
kraje['Tanzania'] = 'Q362561'
kraje['Ukraina'] = 'Q362462'
kraje['Uzbekistan'] = 'Q362512'
kraje['Węgry'] = 'Q362464'
kraje['Włochy'] = 'Q362466'

# wspólna referencja dla wszystkich deklaracji z PRNG
references = {}
references[p_reference_url] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'
#references[p_reference_url] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WMS/WorldGeographicalNames'
references[p_retrieved] = '2022-09-14'

# kwalifikator z punktem czasowym
qualifiers = {}
qualifiers[p_point_in_time] = '+2022-00-00T00:00:00Z/9' # rok 2022


def get_label_en(qid: str) -> str:
    """ zwraca angielską etykietę dla podanego QID """

    wb_item_test = wbi_core.ItemEngine(item_id=qid)
    result = wb_item_test.get_label('en')

    return result


# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':

    # logowanie do instancji wikibase
    login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                        consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                        access_token=WIKIDARIAH_ACCESS_TOKEN,
                                        access_secret=WIKIDARIAH_ACCESS_SECRET,
                                        token_renew_period=14400)

    xlsx_input = '/home/piotr/ihpan/wikihub_skrypty/data_prng/PRNG_egzonimy_czesc.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["PRNG_egzonimy_czesc"]

    # kolumny: nazwaGlow, odmianaNGD, odmianaNGM, odmianaNGP, nazwaDluga, odmianaNDD,
    # odmianaNDM, odmianaNDP, elementRoz, odmianaNOD, odmianaNOM, odmianaNOP, polozenieT, idiip,
    # wspGeograf, + instance of: “administrative unit”

    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    for index, row in enumerate(ws.iter_rows(2, ws.max_row), start=1):
        # wczytanie danych z xlsx
        #if index <= 53:
        #    continue
        nazwa = row[col_names['nazwaGlown']].value
        if not nazwa:
            continue

        label_pl = label_en = nazwa

        nazwa_hist = row[col_names['nazwaHist']].value
        nazwa_obocz = row[col_names['nazwaObocz']].value

        odmiana_ngd = row[col_names['odmianaNGD']].value
        odmiana_ngm = row[col_names['odmianaNGM']].value
        odmiana_ngp = row[col_names['odmianaNGP']].value

        odmiana_nod = row[col_names['odmianaNOD']].value
        odmiana_nom = row[col_names['odmianaNOM']].value
        odmiana_nop = row[col_names['odmianaNOP']].value

        inform_dod = row[col_names['informDod']].value

        idiip = row[col_names['idiip']].value
        polozenie_t = row[col_names['Polozeniet']].value
        polozenie_t_en = row[col_names['Polozeniet_en']].value
        wsp_geo = row[col_names['wspGeograf']].value

        description_pl = 'część miejscowości, osiedla ludzkiego'
        description_en = 'part of settlement'

        # przygotowanie struktur wikibase
        data = []
        aliasy = []

        # instance of
        statement = create_statement_data(p_instance_of, q_human_settlement,
                                          None,
                                          qualifiers,
                                          add_ref_dict=None)
        if statement:
            data.append(statement)

        # współrzędne geograficzne
        if wsp_geo:
            coordinate = get_coord(wsp_geo)
            statement = create_statement_data(p_coordinate, coordinate,
                                              None,
                                              qualifiers,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)

        # id SDI
        if idiip:
            statement = create_statement_data(p_id_sdi, idiip,
                                              None,
                                              qualifiers,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)

        # part of z pola informDod
        part_qid = ''
        if inform_dod:
            parameters = [(p_instance_of, q_human_settlement),
                          (p_located_in_country, kraje[polozenie_t])]
            ok, part_qid = element_search_adv(inform_dod, 'pl', parameters, max_results_to_verify=200)
            if ok:
                # tylko jeżeli ma identyfikator id IIP
                if has_statement(part_qid, p_id_sdi):
                    statement = create_statement_data(p_part_of, part_qid,
                                                      None,
                                                      qualifiers,
                                                      add_ref_dict=references)
                    if statement:
                        data.append(statement)
            else:
                print(f'ERROR: nie znaleziono miejscowości nadrzędnej {inform_dod} dla miejscowości {nazwa}.')

        # nazwaGlow
        local_qualifiers = qualifiers.copy()
        if odmiana_ngd:
            local_qualifiers[p_inflectional_form] = odmiana_ngd
        if odmiana_ngm:
            local_qualifiers[p_locative_form] = odmiana_ngm
        if odmiana_ngp:
            local_qualifiers[p_adjective_form] = odmiana_ngp
        statement = create_statement_data(p_stated_as, f'pl:"{nazwa}"',
                                          None,
                                          local_qualifiers,
                                          add_ref_dict=references)
        if statement:
            data.append(statement)
        del local_qualifiers

        # nazwaHist
        if nazwa_hist and nazwa_hist != nazwa:
            aliasy.append(nazwa_hist)
            statement = create_statement_data(p_stated_as, f'pl:"{nazwa_hist}"',
                                              None,
                                              qualifiers,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)

        # nazwaObocz
        if nazwa_obocz and nazwa_obocz not in (nazwa, nazwa_hist):
            aliasy.append(nazwa_obocz)
            local_qualifiers = qualifiers.copy()
            if odmiana_nod:
                local_qualifiers[p_inflectional_form] = odmiana_nod
            if odmiana_nom:
                local_qualifiers[p_locative_form] = odmiana_nom
            if odmiana_nop:
                local_qualifiers[p_adjective_form] = odmiana_nop
            statement = create_statement_data(p_stated_as, f'pl:"{nazwa_obocz}"',
                                              None,
                                              local_qualifiers,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)
            del local_qualifiers

        #name_description = polozenie_t
        # polozenieT
        if polozenie_t:
            # czy istnieje w wikibase element o takiej nazwie będący instancją 'country'?
            parameters = [(p_instance_of ,q_country)]
            ok, country_qid = element_search_adv(polozenie_t, 'pl', parameters)
            if country_qid:
                #name_description = get_label_en(country_qid)
                statement = create_statement_data(p_located_in_country, country_qid,
                                                  None,
                                                  qualifiers,
                                                  add_ref_dict=references)
                if statement:
                    data.append(statement)
            else:
                # być może to nie państwo a terytorium zależne?
                parameters = [(p_instance_of ,q_dependent_territory)]
                ok, dependent_territory_qid = element_search_adv(polozenie_t, 'pl', parameters)
                if dependent_territory_qid:
                    #name_description = get_label_en(dependent_territory_qid)
                    statement = create_statement_data(p_located_in, dependent_territory_qid,
                                                      None,
                                                      qualifiers,
                                                      add_ref_dict=references)
                    if statement:
                        data.append(statement)
                else:
                    # być może to jednak region?
                    parameters = [(p_instance_of ,q_region)]
                    ok, region_qid = element_search_adv(polozenie_t, 'pl', parameters)
                    if region_qid:
                        #name_description = get_label_en(region_qid)
                        statement = create_statement_data(p_located_in, dependent_territory_qid,
                                                          None,
                                                          qualifiers,
                                                          add_ref_dict=references)
                        if statement:
                            data.append(statement)
                    else:
                        # jeżeli ani country, ani dependent territory an region to zapisujemy
                        # we właściwości 'located_in_(string)'
                        #name_description = polozenie_t
                        statement = create_statement_data(p_located_in_string, polozenie_t,
                                                          None,
                                                          qualifiers,
                                                          add_ref_dict=references)
                        if statement:
                            data.append(statement)

        # etykiety, description, aliasy
        wb_item = wbi_core.ItemEngine(new_item=True, data=data)
        wb_item.set_label(label_en, lang='en')
        wb_item.set_label(label_pl,lang='pl')

        # description musi się różnić, gdyż mogą być identyczne jednostki
        # administracyjne w różnych państwach
        # w ang. wersji także z polozenie_t - w sumie i tak nie ma angielskich tłumaczeń...
        # gdyby były to można by pobrać ze zmiennej name_description
        wb_item.set_description(f'{description_en} ({polozenie_t_en} [{wsp_geo}])', 'en')
        wb_item.set_description(f'{description_pl} ({polozenie_t} [{wsp_geo}])', 'pl')

        if aliasy:
            for value_alias in aliasy:
                wb_item.set_aliases(value_alias, 'pl')

        # wyszukiwanie po etykiecie
        parameters = [(p_instance_of, q_human_settlement)]
        ok, item_id = element_search_adv(label_en, 'en', parameters, f'{description_en} ({polozenie_t_en} [{wsp_geo}])')

        if not ok:
            if WIKIBASE_WRITE:
                #new_id = wb_item.write(login_instance, bot_account=True, entity_type='item')
                new_id = write_or_exit(login_instance, wb_item, None)
                if new_id:
                    print(f'{index}/{ws.max_row - 1} Dodano nowy element: {label_en} / {label_pl} = {new_id} - part of {part_qid}')
            else:
                new_id = 'TEST'
                print(f"{index}/{ws.max_row - 1} Przygotowano dodanie elementu - {label_en} / {label_pl}  = {new_id} - part of {part_qid}")
        else:
            print(f'{index}/{ws.max_row - 1} Element: {label_en} / {label_pl} już istnieje: {item_id}')

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

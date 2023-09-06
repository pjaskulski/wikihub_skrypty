""" Tworzenie xlsx do importu na podstawie xlsx z danymi z PRNG (Kraina, region)  """
import openpyxl

# stałe - zmienić dla docelowej!!!
Q_REGION = 'Q233971'
P_STATED_AS = 'P195'
P_INFLECTIONAL_FORM = 'P282'
P_LOCATIVE_FORM = 'P281'
P_ADJECTIVE_FORM = 'P280'
P_LOCATED_IN = 'P291'
P_LOCATED_IN_COUNTRY = 'P292'
P_COORDINATE_LOCATION = 'P48'
P_ID_SDI = 'P289'
P_REFERENCE_URL = 'P182'
P_POINT_IN_TIME = 'P485'

reference_value = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'

xlsx_input = '../data_prng/PRNG_egzonimy_region_source.xlsx'
wb = openpyxl.load_workbook(xlsx_input)

xlsx_blank = '../data_prng/blank.xlsx'
wb_out = openpyxl.load_workbook(xlsx_blank)

ws = wb["PRNG_egzonimy_region"]

# kolumny: idiip, nazwaGlown, informDod, nazwaObocz, nazwaHist, Polozeniet, odmianaNGD,
# odmianaNGM, odmianaNGP, odmianaNOD, odmianaNOM, odmianaNOP, WGS84

col_names = {}
nr_col = 0
for column in ws.iter_cols(1, ws.max_column):
    col_names[column[0].value] = nr_col
    nr_col += 1

q_list = []
q_statement = []

for index, row in enumerate(ws.iter_rows(2, ws.max_row), start=1):
    nazwa = row[col_names['nazwaGlown']].value
    if not nazwa:
        continue
    nazwa_hist = row[col_names['nazwaHist']].value
    nazwa_obocz = row[col_names['nazwaObocz']].value

    inform_dod = row[col_names['informDod']].value
    idiip = row[col_names['idiip']].value

    odmiana_ngd = row[col_names['odmianaNGD']].value
    odmiana_ngm = row[col_names['odmianaNGM']].value
    odmiana_ngp = row[col_names['odmianaNGP']].value
    odmiana_nod = row[col_names['odmianaNOD']].value
    odmiana_nom = row[col_names['odmianaNOM']].value
    odmiana_nop = row[col_names['odmianaNOP']].value

    polozenie_t = row[col_names['Polozeniet']].value
    wgs84 = row[col_names['WGS84']].value
    wsp_geo = row[col_names['wspGeograf']].value

    if nazwa == 'Sahel' or nazwa == 'Nowa Anglia':
        nazwa_main = f'{nazwa} ({polozenie_t})'
    elif nazwa in ['Luksemburg', 'Śląsk', 'Landy', 'Elida', 'Lauda', 'Dolne Łużyce', 'Hanower', 'Kraina']:
        nazwa_main = f'{nazwa} (region)'
    else:
        nazwa_main = nazwa

    description_pl = f'region/kraina w: {polozenie_t}'
    description_en = f'region in: {polozenie_t}'

    # Label_en, Label_pl, Description_en, Description_pl, Wiki_id, StartsAt, EndsAt, Instance of
    list_item = [nazwa_main, nazwa_main, description_en, description_pl, '', '', '', Q_REGION]
    q_list.append(list_item)

    # Label_en, P, Value, Qualifier, Qualifier_value, Reference_property, Reference_value
    if odmiana_ngd:
        statement_item = [nazwa_main, P_STATED_AS, f'pl:"{nazwa}"',
                          P_INFLECTIONAL_FORM, odmiana_ngd,
                          P_REFERENCE_URL, reference_value]
        q_statement.append(statement_item)
        # point in time
        statement_item = ['', '', '', P_POINT_IN_TIME, '+2022-00-00T00:00:00Z/9', '', '']
        q_statement.append(statement_item)
    else:
        statement_item = [nazwa_main, P_STATED_AS, f'pl:"{nazwa}"', P_POINT_IN_TIME, '+2022-00-00T00:00:00Z/9',
                          P_REFERENCE_URL, reference_value]
        q_statement.append(statement_item)

    # locative form
    if odmiana_ngm:
        statement_item = ['', '', '', P_LOCATIVE_FORM, odmiana_ngm, '', '']
        q_statement.append(statement_item)
    # adjective form
    if odmiana_ngp:
        statement_item = ['', '', '', P_ADJECTIVE_FORM, odmiana_ngp, '', '']
        q_statement.append(statement_item)

    if nazwa_obocz:
        statement_item = [nazwa_main, 'Apl', nazwa_obocz, '', '', '', '']
        q_statement.append(statement_item)

        # inflectional form
        if odmiana_nod:
            statement_item = [nazwa_main, P_STATED_AS, f'pl:"{nazwa_obocz}"', P_INFLECTIONAL_FORM, odmiana_nod,
                              P_REFERENCE_URL, reference_value]
            q_statement.append(statement_item)
            # point in time
            statement_item = ['', '', '', P_POINT_IN_TIME, '+2022-00-00T00:00:00Z/9', '', '']
            q_statement.append(statement_item)

        # locative form
        if odmiana_nom:
            statement_item = ['', '', '', P_LOCATIVE_FORM, odmiana_nom, '', '']
            q_statement.append(statement_item)
        # adjective form
        if odmiana_nop:
            statement_item = ['', '', '', P_LOCATIVE_FORM, odmiana_nop, '', '']
            q_statement.append(statement_item)

    if nazwa_hist:
        statement_item = [nazwa_main, P_STATED_AS, f'pl:"{nazwa_hist}"', P_POINT_IN_TIME, '+2022-00-00T00:00:00Z/9',
                          P_REFERENCE_URL, reference_value]
        q_statement.append(statement_item)

        statement_item = [nazwa_main, 'Apl', nazwa_hist, '', '', '', '']
        q_statement.append(statement_item)

    if inform_dod and inform_dod.startswith('również'):
        tmp = inform_dod[7:].strip()
        tmp_tab = tmp.split(',')
        for tmp_tab_item in tmp_tab:
            tmp_tab_item = tmp_tab_item.strip()
            if tmp_tab_item:
                statement_item = [nazwa_main, P_STATED_AS, f'pl:"{tmp_tab_item}"', P_POINT_IN_TIME, '+2022-00-00T00:00:00Z/9',
                            P_REFERENCE_URL, reference_value]
                q_statement.append(statement_item)

                statement_item = [nazwa_main, 'Apl', tmp_tab_item, '', '', '', '']
                q_statement.append(statement_item)

    # if wgs84:
    #     coordinate = wgs84.replace('Point', '').replace('(', '').replace(')', '').strip().replace(' ', ',')
    #     statement_item = [nazwa_main, P_COORDINATE_LOCATION, coordinate, '', '',
    #                       P_REFERENCE_URL, reference_value]
    #     q_statement.append(statement_item)

    if wsp_geo:
        # 56°30'00" N, 23°30'00" E
        tmp_tab = wsp_geo.split(',')
        char = "'"
        latitude = tmp_tab[0].split(char)[0].replace('°','.')
        stopnie = float(latitude.split('.')[0])
        minuty = float(latitude.split('.')[1])/60.0
        latitude = str(stopnie + minuty)

        tmp_tab[1] = tmp_tab[1].strip()
        longitude = tmp_tab[1].split(char)[0].replace('°','.')
        stopnie = float(longitude.split('.')[0])
        minuty = float(longitude.split('.')[1])/60.0
        longitude = str(stopnie + minuty)

        coordinate = f'{latitude},{longitude}'
        statement_item = [nazwa_main, P_COORDINATE_LOCATION, coordinate, P_POINT_IN_TIME, '+2022-00-00T00:00:00Z/9',
                          P_REFERENCE_URL, reference_value]
        q_statement.append(statement_item)

    if idiip:
        statement_item = [nazwa_main, P_ID_SDI, idiip, P_POINT_IN_TIME, '+2022-00-00T00:00:00Z/9', '', '']
        q_statement.append(statement_item)


# zapis do wyjściowego arkusza
q_list_sheet = wb_out['Q_list']
q_statement_sheet = wb_out['Q_statements']

for item in q_list:
    q_list_sheet.append(item)

for item in q_statement:
    q_statement_sheet.append(item)

xlsx_output = '../data_prng/PRNG_egzonimy_region.xlsx'
wb_out.save(xlsx_output)

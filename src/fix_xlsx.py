""" test """

import os
from pathlib import Path
import openpyxl
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator import wbi_core
from dotenv import load_dotenv

# adresy
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# brak ustawienia tych wartości w wikibase powoduje ostrzeżenia, ale skrypt działa
#wbi_config['PROPERTY_CONSTRAINT_PID'] = 'Pxxx'
#wbi_config['DISTINCT_VALUES_CONSTRAINT_QID'] = 'Qxxx'

WIKIBASE_WRITE = False

# login i hasło ze zmiennych środowiskowych (plik .env w folderze ze źródłami)
env_path = Path('.') / '.env'
load_dotenv(dotenv_path=env_path)
BOT_LOGIN = os.environ.get('WIKIDARIAH_USER')
BOT_PASSWORD = os.environ.get('WIKIDARIAH_PWD')

login_instance = wbi_login.Login(user=BOT_LOGIN, pwd=BOT_PASSWORD)


purl = {}

with open("temp.txt", "r", encoding="utf-8") as f:
    lines = f.readlines()

for line in lines:
    line = line.strip()
    tmp = line.split(" = ")
    purl[tmp[0]] = tmp[1]

wb = openpyxl.load_workbook("../data/07_administrative_types_google.xlsx")

# Q_statements
ws = wb["Q_statements"]
max = ws.max_row

col_names = {}
nr_col = 0
for column in ws.iter_cols(1, ws.max_column):
    col_names[column[0].value] = nr_col
    nr_col += 1

for index, row in enumerate(ws.rows, start=1):
    purl_id = row[col_names['Purl identifier']].value
    property = row[col_names['P']].value
    value = row[col_names['Value']].value
    lista = ['part of', 'equivalent to', 'subclass of', 'composed of']

    if purl_id:
        purl_id = purl_id.strip()
    if property:
        property = property.strip()
    if value:
        value = value.strip()

    if purl_id and purl_id in purl:
        qid = purl[purl_id]
        ws.cell(row=index, column=col_names['Label_en'] + 1).value = qid
        wb_qid = wbi_core.ItemEngine(item_id=qid)
        label = wb_qid.get_label('pl')
        ws.cell(row=index, column=col_names['Label_txt'] + 1).value = label

    if value and (property in lista) and ('http://purl.org/ontohgis#administrative_type' in value):
        if value in purl:
            qid = purl[value]
            ws.cell(row=index, column=col_names['Value'] + 1).value = qid
            ws.cell(row=index, column=col_names['Value'] + 1).hyperlink = ''
            wb_purl = wbi_core.ItemEngine(item_id=qid)
            label = wb_purl.get_label('pl')
            ws.cell(row=index, column=col_names['Value_purl'] + 1).value = label


# Q_list
ws = wb["Q_list"]
max = ws.max_row

col_names = {}
nr_col = 0
for column in ws.iter_cols(1, ws.max_column):
    col_names[column[0].value] = nr_col
    nr_col += 1

for index, row in enumerate(ws.rows, start=1):
    purl_id = row[col_names['Purl identifier']].value

    if purl_id:
        purl_id = purl_id.strip()

    if purl_id and purl_id in purl:
        qid = purl[purl_id]
        ws.cell(row=index, column=col_names['QID'] + 1).value = qid


wb.save("../data/07_administrative_types_google_copy.xlsx")

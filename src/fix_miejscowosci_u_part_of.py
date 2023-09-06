""" uzupełnienie danych 'part of' i 'has part or parts' miejscowosci
    z pliku miejscowosciU_QID.xlsx, plik zawiera dane z PRNG z dodaną kolumną z QID
    do szybkiego wyszukiwania dane zapisane są także w bazie sqlite
"""
import os
import sys
import time
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import get_properties, create_connection, search_sql
from property_import import create_statement_data


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# dane do logowania ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

# znacznik czy zapisywać zmiany w Wikibase, czy tylko test
WIKIBASE_WRITE = True


# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':
    # standardowe właściwości
    print('Przygotowanie właściwości...')
    properties = get_properties(['SIMC place ID', 'prng id', 'reference URL', 'retrieved',
                                'part of', 'has part or parts',
                                ])

    # wspólna referencja dla wszystkich deklaracji z PRG
    references = {}
    references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'
    references[properties['retrieved']] = '2022-09-23'

    # kwalifikator z punktem czasowym
    qualifiers = {}
    qualifiers[properties['point in time']] = '+2022-00-00T00:00:00Z/9' # rok 2022

    # mapowanie rodzajów
    rodzaje = {}
    rodzaje['część kolonii'] = 'kolonia'
    rodzaje['część miasta'] = 'miasto'
    rodzaje['część osady'] = 'osada'
    rodzaje['część wsi'] = 'wieś'
    rodzaje['kolonia wsi'] = 'wieś'
    rodzaje['osada kolonii'] = 'kolonia'
    rodzaje['osada osady'] = 'osada'
    rodzaje['osada wsi'] = 'wieś'
    rodzaje['osiedle wsi'] = 'wieś'
    rodzaje['przysiółek kolonii'] = 'kolonia'
    rodzaje['przysiółek osady'] = 'osada'
    rodzaje['przysiółek wsi'] = 'wieś'
    rodzaje['osada leśna wsi'] = 'wieś'
    rodzaje['kolonia kolonii'] = 'kolonia'
    rodzaje['kolonia osady'] = 'osada'


    # logowanie do instancji wikibase
    login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                     consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                     access_token=WIKIDARIAH_ACCESS_TOKEN,
                                     access_secret=WIKIDARIAH_ACCESS_SECRET,
                                     token_renew_period=14400)

    xlsx_input = '../data_prng/miejscowosciU_QID.xlsx'
    wb = openpyxl.load_workbook(xlsx_input, read_only=False)
    ws = wb["miejscowosciU"]

    miejscowosci_path = '../data_prng/miejscowosci_u.sqlite'
    db_m = create_connection(miejscowosci_path)

    raport_path = '../data_prng/miejscowosci_part_of_errors.txt'

    # nazwy kolumn w arkuszu
    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    # przetwarzanie kolejnych wierszy arkusza
    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1
        if index <= 97842:
            continue

        # wczytanie danych z xlsx
        row_qid = row[col_names['QID']].value
        row_nazwamiejs = row[col_names['NAZWAMIEJS']].value
        row_gmina = row[col_names['GMINA']].value
        row_rodzaj = row[col_names['RODZAJOBIE']].value
        row_powiat = row[col_names['POWIAT']].value
        row_wojewodztw = row[col_names['WOJEWODZTW']].value
        row_wgs84 = row[col_names['WGS84']].value
        row_longitude = row_latitude = 0
        if row_wgs84:
            row_wgs84 = row_wgs84.replace('Point', '').replace('(', '').replace(')','').strip()
            tmp = row_wgs84.split(' ')
            row_longitude = float(tmp[0])
            row_latitude = float(tmp[1])

        if not row_qid:
            continue

        # przygotowanie struktur wikibase
        data = []
        r_data = []

        if not row_nazwamiejs:
            continue

        part_of_qid = ''

        # nazwa, województwo, powiat, rodzaj miejscowości i współrzędne
        sql = f"""
            SELECT NAZWAGLOWN, QID, RODZAJOBIE, WGS84 FROM miejscowosciU_QID
            WHERE NAZWAGLOWN = "{row_nazwamiejs.strip()}"
            AND WOJEWODZTW = "{row_wojewodztw.strip()}"
            AND POWIAT = "{row_powiat.strip()}"
            AND GMINA = "{row_gmina.strip()}"
            AND RODZAJOBIE = "{rodzaje[row_rodzaj]}"
            """
        part_of_qid = search_sql(db_m, sql, row_latitude, row_longitude)

        # nazwa, województwo, powiat, rodzaj miejscowości i współrzędne
        if not part_of_qid:
            sql = f"""
                SELECT NAZWAGLOWN, QID, RODZAJOBIE, WGS84 FROM miejscowosciU_QID
                WHERE NAZWAGLOWN = "{row_nazwamiejs.strip()}"
                AND WOJEWODZTW = "{row_wojewodztw.strip()}"
                AND POWIAT = "{row_powiat.strip()}"
                AND RODZAJOBIE = "{rodzaje[row_rodzaj]}"
            """
            part_of_qid = search_sql(db_m, sql, row_latitude, row_longitude)

        # nazwa, województwo, rodzaj miejscowości i współrzędne
        if not part_of_qid:
            sql = f"""
                SELECT NAZWAGLOWN, QID, RODZAJOBIE, WGS84 FROM miejscowosciU_QID
                WHERE NAZWAGLOWN = "{row_nazwamiejs.strip()}"
                AND WOJEWODZTW = "{row_wojewodztw.strip()}"
                AND RODZAJOBIE = "{rodzaje[row_rodzaj]}"
            """
            part_of_qid = search_sql(db_m, sql, row_latitude, row_longitude)

        # tylko nazwa i rodzaj miejscowości + współrzędne
        if not part_of_qid:
            sql = f"""
                SELECT NAZWAGLOWN, QID, RODZAJOBIE, WGS84 FROM miejscowosciU_QID
                WHERE NAZWAGLOWN = "{row_nazwamiejs.strip()}"
                AND RODZAJOBIE = "{rodzaje[row_rodzaj]}"
            """
            part_of_qid = search_sql(db_m, sql, row_latitude, row_longitude)

        # utworzenie deklaracji
        if part_of_qid:
            statement = create_statement_data(properties['part of'], part_of_qid,
                             None, qualifier_dict=qualifiers, add_ref_dict=references)
            if statement:
                data.append(statement)
            # odwrotność
            r_statement = create_statement_data(properties['has part or parts'], row_qid,
                            None, qualifier_dict=qualifiers, add_ref_dict=references, if_exists='APPEND')
            if r_statement:
                r_data.append(r_statement)
        else:
            print(f'ERROR: {row_qid} - brak miejscowości nadrzędnej {row_nazwamiejs}')
            with open(raport_path, 'a', encoding='utf-8') as f_raport:
                f_raport.write(f'ERROR: {row_qid} - brak miejscowości nadrzędnej {row_nazwamiejs}\n')

        # jeżeli nie ma nic do uzupełnienia
        if not data:
            continue

        if WIKIBASE_WRITE:
            test = 1
            while True:
                try:
                    wb_item = wbi_core.ItemEngine(item_id=row_qid, data=data)
                    wb_item.write(login_instance, bot_account=True, entity_type='item')
                    if part_of_qid and r_data:
                        wb_r_item = wbi_core.ItemEngine(item_id=part_of_qid, data=r_data)
                        wb_r_item.write(login_instance, bot_account=True, entity_type='item')

                    print(f"{index}/{max_row - 1} Uzupełniono właściwości 'part of' i 'has part or parts'")
                    break
                except MWApiError as wbdelreference_error:
                    err_code = wbdelreference_error.error_msg['error']['code']
                    message = wbdelreference_error.error_msg['error']['info']
                    print(f'ERROR: {err_code}, {message}')
                    # jeżeli jest to problem z tokenem to próba odświeżenia tokena i powtórzenie
                    # zapisu, ale tylko raz, w razie powtórnego błędu bad token, skrypt kończy pracę
                    if err_code in ['assertuserfailed', 'badtoken']:
                        if test == 1:
                            print('Generate edit credentials...')
                            login_instance.generate_edit_credentials()
                            test += 1
                            continue
                    sys.exit(1)
        else:
           print(f"{index}/{max_row - 1} Przygotowano uzupełnienie właściwości 'part of' i 'has part or parts'.")

    wb.close()

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

""" import powiatów z pliku powiaty.xlsx z danymi z PRG"""
import os
import time
import sys
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import element_search_adv, search_by_purl, get_properties, get_elements
from property_import import create_statement_data, has_statement


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

# czy zapisywać dane w wikibase, czy tylko test
WIKIBASE_WRITE = False

# standardowe właściwości
properties = get_properties(['instance of', 'stated as', 'reference URL', 'retrieved',
                            'id SDI', 'part of', 'has part or parts', 'TERYT'])

# elementy definicyjne
elements = get_elements(['administrative unit', 'http://purl.org/ontohgis#administrative_type_46'])


# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':

    # wspólna referencja dla wszystkich deklaracji z PRG
    references = {}
    references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGIK/PRG/WFS/AdministrativeBoundaries'
    references[properties['retrieved']] = '2022-09-05'

    # wspólna referencja dla wszystkich deklaracji z ontohgis
    onto_references = {}
    onto_references[properties['reference URL']] = 'https://ontohgis.pl'

    # logowanie do instancji wikibase
    if WIKIBASE_WRITE:
        login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                         consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                         access_token=WIKIDARIAH_ACCESS_TOKEN,
                                         access_secret=WIKIDARIAH_ACCESS_SECRET,
                                         token_renew_period=14400)

    xlsx_input = '../data_prng/powiaty.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["powiaty"]

    # nazwy kolumn z xlsx
    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    parts = {}

    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1
        # wczytanie danych z xlsx
        nazwa = row[col_names['JPT_NAZWA_']].value
        if not nazwa:
            continue

        label_pl = label_en = nazwa

        teryt = row[col_names['JPT_KOD_JE']].value
        idiip = row[col_names['IIP_IDENTY']].value
        part_of = teryt[:2]

        description_pl = 'powiat - współczesna jednostka administracyjna'
        description_en = 'powiat - współczesna jednostka administracyjna'

        # przygotowanie struktur wikibase
        data = []
        aliasy = []

        # instance of
        q_powiat = elements['http://purl.org/ontohgis#administrative_type_46']
        statement = create_statement_data(properties['instance of'], q_powiat, None, None, add_ref_dict=onto_references)
        if statement:
            data.append(statement)

        # id SDI
        if idiip:
            statement = create_statement_data(properties['id SDI'], idiip, None, None, add_ref_dict=references)
            if statement:
                data.append(statement)

        # TERYT
        if teryt:
            statement = create_statement_data(properties['TERYT'], teryt, None, None, add_ref_dict=references)
            if statement:
                data.append(statement)

        # JPT_NAZWA_
        statement = create_statement_data(properties['stated as'], f'pl:"{nazwa}"', None, None, add_ref_dict=references)
        if statement:
            data.append(statement)

        # part of i has part or parts
        ok, woj_qid = search_by_purl(properties['TERYT'], part_of)
        if ok:
            statement = create_statement_data(properties['part of'], woj_qid, None, None, add_ref_dict=references)
            if statement:
                data.append(statement)
        else:
            woj_qid = ''

        # etykiety, description, aliasy
        wb_item = wbi_core.ItemEngine(new_item=True, data=data)
        wb_item.set_label(label_en, lang='en')
        wb_item.set_label(label_pl,lang='pl')

        # description
        wb_woj = wbi_core.ItemEngine(item_id=woj_qid)
        woj_label_pl = wb_woj.get_label('pl')

        wb_item.set_description(f"{description_en} ({woj_label_pl})", 'en')
        wb_item.set_description(f"{description_pl} ({woj_label_pl})", 'pl')

        if aliasy:
            for value_alias in aliasy:
                wb_item.set_aliases(value_alias, 'pl')

        # wyszukiwanie po etykiecie, właściwości instance of oraz po opisie
        parameters = [(properties['instance of'], q_powiat)]
        ok, item_id = element_search_adv(label_en, 'en', parameters, description_en)

        if not ok:
            if WIKIBASE_WRITE:
                test = 1
                while True:
                    try:
                        new_id = wb_item.write(login_instance, bot_account=True, entity_type='item')
                        print(f'{index}/{ws.max_row - 1} Dodano nowy element: {label_en} / {label_pl} = {new_id}')
                        if woj_qid:
                            if woj_qid in parts:
                                parts[woj_qid].append(new_id)
                            else:
                                parts[woj_qid] = [new_id]
                        break
                    except MWApiError as wb_error:
                        err_code = wb_error.error_msg['error']['code']
                        message = wb_error.error_msg['error']['info']
                        print(f'ERROR: {err_code}, {message}')
                        # jeżeli jest to problem z tokenem to próba odświeżenia tokena i powtórzenie
                        # zapisu, ale tylko raz, w razie powtórnego błędu bad token, skrypt kończy pracę
                        if err_code in ['assertuserfailed', 'badtoken']:
                            if test == 1:
                                print('Generate edit credentials...')
                                login_instance.generate_edit_credentials()
                                test += 1
                                continue
                        sys.exit(1)
            else:
                new_id = 'TEST'
                print(f"{index}/{ws.max_row - 1} Przygotowano dodanie elementu - {label_en} / {label_pl}  = {new_id}")
                if woj_qid:
                    if woj_qid in parts:
                        parts[woj_qid].append(new_id)
                    else:
                        parts[woj_qid] = [new_id]
        else:
            print(f'{index}/{ws.max_row - 1} Element: {label_en} / {label_pl} już istnieje: {item_id}')

    # uzupełnienie właściwości 'has part or parts' dla województw
    if parts:
        print("\nUzupełnienie 'has part or parts' dla województw.\n")

        for item_qid, part_qids in parts.items():
            for part_qid in part_qids:
                if not has_statement(item_qid, properties['has part or parts'], part_qid):
                    if WIKIBASE_WRITE:
                        data = []
                        statement = create_statement_data(properties['has part or parts'], part_qid, references, None, if_exists='APPEND')
                        if statement:
                            data.append(statement)
                    else:
                        print(f"Przygotowano dodanie do elementu {item_qid} właściwości {properties['has part or parts']} o wartości {part_qid}")

            if data:
                if WIKIBASE_WRITE:
                    test = 1
                    while True:
                        try:
                            wd_item = wbi_core.ItemEngine(item_id=item_qid, data=data, debug=False)
                            wd_item.write(login_instance, entity_type='item')
                            print(f"Do elementu {item_qid} dodano właściwości {properties['has part or parts']}.")
                            break
                        except MWApiError as wb_error:
                            err_code = wb_error.error_msg['error']['code']
                            message = wb_error.error_msg['error']['info']
                            print(f'ERROR: {err_code}, {message}')
                            # jeżeli jest to problem z tokenem to próba odświeżenia tokena i powtórzenie
                            # zapisu, ale tylko raz, w razie powtórnego błędu bad token, skrypt kończy pracę
                            if err_code in ['assertuserfailed', 'badtoken']:
                                if test == 1:
                                    print('Generate edit credentials...')
                                    login_instance.generate_edit_credentials()
                                    test += 1
                                    continue
                            sys.exit(1)
            else:
                print(f"ERROR: {item_qid}, wystąpił problem z przygotowaniem danych dla właściwości 'has part or parts'")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

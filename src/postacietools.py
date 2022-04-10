""" moduł """
import re
import sys
import roman as romenum
from openpyxl import load_workbook
from wyjatki_postacie import WYJATKI
from wyjatki_postacie import ETYKIETY_WYJATKI
from wikidariahtools import text_clear, short_names_in_autor

class FigureName:
    """ obsługa imion i nazwisk postaci historycznych """

    def __init__(self, f_name: str) -> None:
        self.name = f_name.strip()
        self.org_name = self.name
        self.imie = ''
        self.imie2 = ''
        self.imie3 = ''
        self.imie4 = ''
        self.nazwisko = ''
        self.nazwisko2 = ''
        self.birth_name = ''
        self.name_etykieta = ''
        self.get_name()
        self.postac_etykieta()
        self.postac_etykieta_korekty()


    def _get_name_simple(self, value: str) -> tuple:
        """ _get name simple """
        p_nazwisko = p_imie = p_imie2 = ''

        tmp = value.split(" ")
        if len(tmp) == 1:
            p_imie = tmp[0].strip()
        elif len(tmp) == 2:
            if tmp[0][0].isupper() and tmp[1][0].isupper():
                p_nazwisko = tmp[0].strip()
                p_imie = tmp[1].strip()
        elif len(tmp) == 3:
            if tmp[0][0].isupper() and tmp[1][0].isupper() and tmp[2][0].isupper():
                p_nazwisko = tmp[0].strip()
                if (tmp[1].endswith('ski') or tmp[1].endswith('icz')
                    or tmp[1].endswith('ska') or tmp[1].endswith('iczowa')
                    or tmp[1].endswith('zic')):
                    p_imie = tmp[2]
                elif tmp[2].endswith('wic') or tmp[2].endswith('yc'): # Januszowski Jan Łazarzowic
                    p_imie = tmp[1]
                else:
                    p_imie = tmp[1].strip()
                    p_imie2 = tmp[2].strip()

        return p_nazwisko, p_imie, p_imie2


    def get_name(self):
        """ get_name """

        lista = ['(młodszy)', '(starszy)', '(Młodszy)', '(Starszy)',
                'młodszy', 'starszy', 'Młodszy', 'Starszy', 'junior', 'senior',
                'właśc.', 'jr']
        for item in lista:
            if item in self.name:
                self.name = self.name.replace(item, '').strip()

        # jeżeli postać jest w wyjątkach to funkcja zwraca wartości z tablicy wyjątków
        if self.name in WYJATKI:
            if 'imie' in WYJATKI[self.name]:
                self.imie = WYJATKI[self.name]['imie'].strip()
            if 'imie2' in WYJATKI[self.name]:
                self.imie2 = WYJATKI[self.name]['imie2'].strip()
            if 'imie3' in WYJATKI[self.name]:
                self.imie3 = WYJATKI[self.name]['imie3'].strip() 
            if 'imie4' in WYJATKI[self.name]:
                self.imie4 = WYJATKI[self.name]['imie4'].strip()       
            if 'nazwisko' in WYJATKI[self.name]:
                self.nazwisko = WYJATKI[self.name]['nazwisko'].strip()
            if 'nazwisko2' in WYJATKI[self.name]:
                self.nazwisko2 = WYJATKI[self.name]['nazwisko2'].strip()

            return

        tmp = self.name.strip().split(" ")

        # liczby rzymskie w przypadku władców - nie są imionami i nazwiskami
        # zazwyczaj za taką liczbą jest przydomek, więc pomijam
        roman = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X',
                'XI', 'XII', 'XIII', 'XIV', 'XV', 'XVI', 'XVII', 'XVIII', 'XIX', 'XX']

        is_roman = False
        for i, tmp_value in enumerate(tmp):
            tmp_value = tmp_value.strip()
            if tmp_value in roman:
                is_roman = True
            if is_roman:
                tmp[i] = ''

        tmp = [item for item in tmp if item.strip() != '']

        if len(tmp) == 1:
            # czy to imię czy nazwisko? Dodać słownik typowych imion? na razie wszystkie
            # pojedyncze traktowane są jak imiona, chyba że kończy się na 'ski'
            p_word = tmp[0].strip()
            if p_word.endswith('ski') or p_word.endswith('ska') or p_word.endswith('cki'):
                self.nazwisko = p_word
            else:
                self.imie = p_word

        elif len(tmp) == 2:
            if tmp[0][0].isupper() and tmp[1][0].isupper():
                self.nazwisko = tmp[0].strip()
                self.imie = tmp[1].strip()

        elif len(tmp) == 3:
            if tmp[0][0].isupper() and tmp[1][0].isupper() and tmp[2][0].isupper():
                self.nazwisko = tmp[0].strip()
                if (tmp[1].endswith('ski') or tmp[1].endswith('icz')
                    or tmp[1].endswith('ska') or tmp[1].endswith('iczowa')
                    or tmp[1].endswith('zic')):
                    self.imie = tmp[2]
                elif tmp[2].endswith('wic') or tmp[2].endswith('yc'): # Januszowski Jan Łazarzowic
                    self.imie = tmp[1]
                else:
                    self.imie = tmp[1]
                    self.imie2 = tmp[2]
            else:
                if ' z ' in self.name or ' ze ' in self.name: # Szymon ze Stawu
                    self.imie = tmp[0].strip()
                if (' zw. ' in self.name or 'zwany' in self.name or 'zapewne' in self.name
                    or ' z ' in self.name or ' ze ' in self.name or ' w zak. ' in self.name
                    or ' zak. ' in self.name or ' syn ' in self.name):
                    pos = self.name.find(' zw. ')
                    pos1 = self.name.find(' zwany')
                    if pos1 > 0 and (pos == -1 or pos1 < pos):
                        pos = pos1
                    pos1 = self.name.find(' zapewne ')
                    if pos1 > 0 and (pos == -1 or pos1 < pos):
                        pos = pos1     
                    pos1 = self.name.find(' z ')
                    if pos1 > 0 and (pos == -1 or pos1 < pos):
                        pos = pos1
                    pos1 = self.name.find(' ze ')
                    if pos1 > 0 and (pos == -1 or pos1 < pos):
                        pos = pos1
                    pos1 = self.name.find(' w zak. ')
                    if pos1 > 0 and (pos == -1 or pos1 < pos):
                        pos = pos1
                    pos1 = self.name.find(' zal. ')
                    if pos1 > 0 and (pos == -1 or pos1 < pos):
                        pos = pos1
                    pos1 = self.name.find(' syn ')
                    if pos1 > 0 and (pos == -1 or pos1 < pos):
                        pos = pos1
                    if pos != -1:
                        self.name = self.name[:pos].strip().replace(',', '')
                        self.nazwisko, self.imie, self.imie2 = self._get_name_simple(self.name)
                elif tmp[1] == 'de':  # Camelin de Jan
                    self.nazwisko = tmp[1] + ' ' + tmp[0].strip()
                    self.imie = tmp[2]
                elif tmp[-1] == 'de':  # Girard Filip de
                    self.nazwisko = tmp[-1] + ' ' + tmp[0].strip()
                    self.imie = tmp[1]
                elif tmp[-1] == 'von': # Kempen Eggert von
                    self.nazwisko = tmp[-1] + ' ' + tmp[0].strip()
                    self.imie = tmp[1]
                elif tmp[1] == 'del':  # Pace del Luca
                    self.nazwisko = tmp[1] + ' ' + tmp[0].strip()
                    self.imie = tmp[2]

        else:
            if ' de ' in self.name:
                self.imie = tmp[-1].strip()
                self.nazwisko = ' '.join(tmp[:-1])
            elif (' z ' in self.name or ' ze ' in self.name or ' zw. ' in self.name
                  or 'syn' in self.name):
                # Boner Seweryn z Balic, Abraham ben Joszijahu z Trok, Łukasz z Nowego Miasta
                pos = self.name.find(' z ')
                pos1 = self.name.find(' ze ')
                if pos1 > 0 and (pos == -1 or pos1 < pos):
                    pos = pos1
                pos1 = self.name.find(' zw. ')
                if pos1 > 0 and (pos == -1 or pos1 < pos):
                    pos = pos1
                pos1 = self.name.find(' syn ')
                if pos1 > 0 and (pos == -1 or pos1 < pos):
                    pos = pos1

                if pos != -1:
                    self.name = self.name[:pos].strip()
                    tmp = self.name.split(" ")
                    self.nazwisko, self.imie, self.imie2 = self._get_name_simple(self.name)
            elif tmp[-1] == 'de':  # Caraccioli Ludwik Antoni de
                self.nazwisko = tmp[-1] + ' ' + tmp[0].strip()
                self.imie = tmp[1]
                self.imie2 = tmp[2]
            elif 'van der' in self.name and len(tmp) == 4:
                self.nazwisko = 'van der' + ' ' + tmp[0].strip()
                self.imie = tmp[-1].strip()
            elif 'w zakonie' in self.name or 'w zak.' in self.name or ' zak. ' in self.name:
                pos = self.name.find(' w zak')
                pos1 = self.name.find(' zak. ')
                if pos1 > 0 and (pos == -1 or pos1 < pos):
                    pos = pos1
                if pos != -1:
                    self.name = self.name[:pos].strip().replace(',', '')
                    self.nazwisko, self.imie, self.imie2 = self._get_name_simple(self.name)
            else:
                if len(tmp) == 4:
                    self.nazwisko = tmp[0]
                    if (tmp[1].endswith('ski') or tmp[1].endswith('icz') 
                        or tmp[1].endswith('ska') or tmp[1].endswith('iczowa')):
                        self.nazwisko2 = tmp[1]
                        self.imie = tmp[2]
                        self.imie2 = tmp[3]
                    else:
                        self.imie = tmp[1]
                        self.imie2 = tmp[2]
                        self.imie3 = tmp[3]

        # zakładam że otczewstwo to nie imię
        if self.imie2.endswith('icz') or self.imie2.endswith('cki'):
            self.imie2 = ''
        if self.imie3.endswith('icz') or self.imie3.endswith('cki'):
            self.imie3 = ''

        # jeżeli imię zaczyna się z małej litery, to błędnie rozpoznano i to nie jest imię
        if self.imie and self.imie[0].islower():
            self.imie = ''
        if self.imie2 and self.imie2[0].islower():
            self.imie2 = ''
        if self.imie3 and self.imie3[0].islower():
            self.imie3 = ''
        if self.imie4 and self.imie4[0].islower():
            self.imie4 = ''

        if not self.imie and not self.nazwisko:
            print(f'PROBLEM, nie ropoznano imienia i nazwiska: {self.org_name}')


    def _double_space(self, value:str) -> str:
        """ usuwa podwójne spacje z przekazanego tekstu """
        return ' '.join(value.strip().split())


    def postac_etykieta(self):
        """ ustala etykietę dla postaci (imiona nazwiska)
            imie_1 .. imie_4 - kolejne imiona postaci
            nazwisko_1, nazwisko_2 - kolejne nazwiska postaci
            konstrukcja etykiety powinna uwzględniać przestawienia kolejności
            imienia i nazwiska (w indeksie BB jest odwrotnie) oraz ewentualne 
            imiona zakonnych itp.
        """
        self.name_etykieta = f'{self.imie} {self.imie2} {self.imie3} {self.imie4} {self.nazwisko2} {self.nazwisko}'
        self.name_etykieta = self.name_etykieta.strip()
        self.name_etykieta = self._double_space(self.name_etykieta)


    def postac_etykieta_korekty(self):
        """ funkcja koryguje w razie potrzeby etykietę dla postaci
            zwraca:
            poprawioną etykietę, imię po narodzinach (dla zakonników/zakonnic)
        """
        if ' zwany ' in self.org_name:
            t_mark = ' zwany '
        elif ' zw. ' in self.org_name:
            t_mark = ' zw. '
        elif ' z ' in self.org_name:
            t_mark = ' z '
        elif ' ze ' in self.org_name:
            t_mark = ' ze '
        elif ' h.' in self.org_name:
            t_mark = ' h.'
        else:
            t_mark = ''

        if t_mark:
            pos = self.org_name.find(t_mark)
            toponimik = self.org_name[pos:]
            self.name_etykieta = self.name_etykieta + toponimik

        if 'starszy' in self.org_name and 'starszy' not in self.name_etykieta:
            self.name_etykieta += ' ' + 'starszy'
        if 'Starszy' in self.org_name and 'Starszy' not in self.name_etykieta:
            self.name_etykieta += ' ' + 'Starszy'
        if 'młodszy' in self.org_name and 'młodszy' not in self.name_etykieta:
            self.name_etykieta += ' ' + 'młodszy'
        if 'Młodszy' in self.org_name and 'Młodszy' not in self.name_etykieta:
            self.name_etykieta += ' ' + 'Młodszy'
        if 'junior' in self.org_name and 'junior' not in self.name_etykieta:
            self.name_etykieta += ' ' + 'junior'
        if 'senior' in self.org_name and 'senior' not in self.name_etykieta:
            self.name_etykieta += ' ' + 'senior'

        # imona zakonne
        if ' w zak. ' in self.org_name:
            z_mark = ' w zak.'
        elif ' w zakonie ' in self.org_name:
            z_mark = ' w zakonie '
        elif ' w zak ' in self.org_name:
            z_mark = ' w zak '
        elif ' zak. ' in self.org_name:
            z_mark = ' zak. '
        else:
            z_mark = ''

        if z_mark:
            pos = self.org_name.find(z_mark)
            pos2 = pos + len(z_mark)
            self.birth_name = self.name_etykieta
            # zakonimik = self.name[pos:]
            zakon_names = self.org_name[pos2:].strip()
            self.name_etykieta = f'{zakon_names} {self.nazwisko2} {self.nazwisko}'.strip()

        self.name_etykieta = self._double_space(self.name_etykieta)

        # dla postaci typu 'Szneur Zalman ben Baruch' na razie tak jak w oryginale
        if ' ben ' in self.org_name:
            self.name_etykieta = self.org_name
        if ' Ben ' in self.org_name:
            self.name_etykieta = self.org_name

        # dla postaci typu 'Salomon syn Joela' na razie tak jak w oryginale
        if ' syn ' in self.org_name:
            self.name_etykieta = self.org_name

        # dla władców tak jak w oryginale
        is_king = False
        roman_num = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X',
            'XI', 'XII', 'XIII', 'XIV', 'XV', 'XVI', 'XVII', 'XVIII', 'XIX', 'XX']
        for r_item in roman_num:
            if f' {r_item} ' in self.org_name or self.org_name.endswith(f' {r_item}'):
                is_king = True
                break
        if is_king:
            self.name_etykieta = self.org_name

        # konstrukcja typu Mikołaj z Jaroszowa zw. Kornicz lub Siestrzeniec
        # zostaje bez zmian w etykiecie
        if ' z ' in self.org_name and ' zw. ' in self.org_name:
            self.name_etykieta = self.org_name
        if ' z ' in self.org_name and ' zwany ' in self.org_name:
            self.name_etykieta = self.org_name

        # jeżeli nie było rozpoznanego imienia tylko przydomek/przezwisko
        if self.name_etykieta.startswith('z '):
            self.name_etykieta = self.org_name

        # super wyjątki nie podpadające gdzie indziej
        self.name_etykieta = ETYKIETY_WYJATKI.get(self.org_name, self.name_etykieta)


class DateBDF:
    """ obługa daty urodzenia, śmierci lub flourit """

    P_DATE_OF_BIRTH = 'P7'
    P_DATE_OF_DEATH = 'P8'
    P_EARLIEST_DATE = 'P38'
    P_LATEST_DATE = 'P39'
    P_FLORUIT = 'P54'
    Q_CIRCA = 'Q37979'
    P_SOURCING_CIRCUMSTANCES = 'P189'
    P_REFINE_DATE = 'P190'
    Q_FIRST_HALF = 'Q40688'
    Q_SECOND_HALF = 'Q41336'
    Q_BEGINNING_OF = 'Q41337'
    Q_MIDDLE_OF = 'Q41338'
    Q_END_OF = 'Q41339'
    Q_FIRST_QUARTER = 'Q49427'

    def __init__(self, text:str, typ:str = '') -> None:
        """ init """
        self.text_org = text
        self.text = text.strip().lower()
        self.type = typ
        self.date = ''
        self.date_2 = ''
        self.about = False
        self.between = False
        self.or_date = False
        self.turn = False
        self.before = False
        self.after = False
        self.certain = False
        self.first_half = False
        self.second_half = False
        self.beginning_of = False
        self.middle_of = False
        self.end_of = False
        self.first_quarter = False
        self.somevalue = False
        self.roman = self.roman_numeric()
        if not self.type:
            self.find_type()
        self.find_uncertainty()
        self.find_date()
        if not self.certain and (self.before or self.after or self.between):
            self.somevalue = True


    def find_type(self):
        """ ustala typ daty """
        if 'zm.' in self.text or 'zmarł' in self.text:
            self.type = 'D'
        elif 'um.' in self.text:
            self.type = 'D'
        elif 'ur.' in self.text:
            self.type = 'B'
        else:
            self.type = 'F'


    def find_uncertainty(self):
        """ ustala czy jest i jaka niepewność dla daty """

        tmp = self.text.replace('zm.', '').replace('um.', '').replace('ur.', '').strip()
        if tmp.isnumeric():
            self.certain = True

        if 'prawdopodobnie' in self.text:
            self.about = True

        if 'ok.' in self.text or 'około' in self.text or 'ok ' in self.text:
            self.about = True

        if 'między' in self.text or 'miedzy' in self.text:
            self.between = True

        if 'w okresie II wojny światowej' in self.text_org:
            self.between = True

        if 'lub nieco później' in self.text:
            self.after = True

        if 'w/przed' in self.text:
            self.before = True

        if 'przed lub w' in self.text:
            self.before = True

        if 'w lub po' in self.text or 'po lub w' in self.text:
            self.after = True

        if ('prawdopodobnie' in self.text or 'zapewne' in self.text 
            or 'rzekomo' in self.text):
            self.about = True

        if ('lub' in self.text and not 'po ' in self.text
               and not ' w ' in self.text and not 'przed ' in self.text
               and not 'nieco później' in self.text):
            self.or_date = True

        if 'przed ' in self.text:
            self.before = True

        if 'po ' in self.text:
            self.after = True

        if 'nie później niż' in self.text:
            self.before = True

        if 'najpóźniej' in self.text:
            self.before = True

        # niepewność
        if '?' in self.text:
            self.about = True

        # przełom lat
        match = re.search(r'\d{3,4}/\d{1,2}', self.text)
        if match:
            self.turn = True

        if '1 poł.' in self.text or 'i poł.' in self.text or '1. poł.' in self.text:
            self.first_half = True
        elif '2 poł.' in self.text or 'ii poł.' in self.text or '2. poł.' in self.text:
            self.second_half = True
        elif '1. ćwierć' in self.text:
            self.first_quarter = True
        elif 'pocz.' in self.text:
            self.beginning_of = True
        elif 'koniec' in self.text or 'w końcu' in self.text:
            self.end_of = True
        elif 'poł.' in self.text:
            self.middle_of = True


    def roman_numeric(self) -> bool:
        """ czy liczba rzymska oznaczająca wiek? 
            (ale nie część daty np. 3 V 1458)
        """
        # jeżeli to specyficzny zapis:
        if 'II wojny' in self.text_org:
            return False

        pattern_test = r'\d{1,2}\s+[IVX]{1,4}\s+\d{4}'
        match = re.search(pattern_test, self.text_org)
        if match:
            return False

        pattern = r'[IVX]{1,5}\s+w\.{0,1}'
        match = re.search(pattern, self.text_org)
        if not match:
            pattern = r'[IVX]{1,5}'
            match = re.search(pattern, self.text_org)

        return bool(match)

    def find_date(self):
        """ wyszukuje daty """
        # XVI w.
        if self.roman: 
            matches = [x.group() for x in re.finditer(r'[IVX]{1,5}', self.text_org)]
            if len(matches) == 1:
                self.date = str(romenum.fromRoman(matches[0]))
            elif len(matches) == 2:
                matches = [str(romenum.fromRoman(x)) for x in matches]
                self.date = matches[0]
                self.date_2 = matches[1]
                if '/' in self.text:
                    self.between = True
                    self.somevalue = True
        # 1523/4
        elif self.turn:
            match = re.search(r'\d{3,4}/\d{1,2}', self.text)
            if match:
                v_list = match.group().split('/')
                v_list1 = v_list[0].strip()
                v_list2 = v_list1[:len(v_list1)-len(v_list[1].strip())] + v_list[1].strip()
                self.date = v_list1
                self.date_2 = v_list2
        # zwykłe daty (jeszcze obsługa dat dziennych i miesięcznych do zrobienia)
        else:
            if 'w okresie II wojny światowej' in self.text_org:
                self.date = '1939'
                self.date_2 = '1945'
            else:
                # test czy to nie data dzienna
                pattern_test = r'\d{1,2}\s+[IVX]{1,4}\s+\d{4}'
                match = re.search(pattern_test, self.text_org)
                if match:
                    t_match = match.group().split(' ')
                    y = t_match[2]
                    m = str(romenum.fromRoman(t_match[1]))
                    d = t_match[0]
                    self.date = f'{y}-{m.zfill(2)}-{d.zfill(2)}'
                else:
                    pattern = r'\d{3,4}'
                    matches = [x.group() for x in re.finditer(pattern, self.text)]
                    if len(matches) == 1:
                        self.date = matches[0]
                    elif len(matches) > 1:
                        self.date = matches[0]
                        self.date_2 = matches[1]


    def _format_date(self, value: str) -> str:
        """ formatuje datę na sposób oczekiwany przez QuickStatements
            np. +1839-00-00T00:00:00Z/9
        """
        result = ''
        if len(value) == 4:                          # tylko rok
            result = f"+{value}-00-00T00:00:00Z/9"
        elif len(value) == 10:                       # dokłada data
            if value.endswith('00'):                 # jeżeli brak daty dziennej 
                result = f"+{value}T00:00:00Z/10"
            else:
                result = f"+{value}T00:00:00Z/11"
        elif len(value) == 2 and value.isnumeric():  # wiek
            result = f"+{str(int(value)-1)}01-00-00T00:00:00Z/7"
        elif len(value) == 1 and value.isnumeric():  # wiek np. X
            value = str(int(value)-1)
            result = f"+{value.zfill(2)}01-00-00T00:00:00Z/7"

        return result

    def prepare_qs(self, etykieta: str = '') -> str:
        """ drukuje zapisy QuickStatements"""
        # data urodzin
        print_date = print_date_2 = print_kw_date = print_kw_date_2 = ''
        if self.somevalue:
            print_date = 'somevalue'
            print_kw_date = self._format_date(self.date)
            print_kw_date_2 = self._format_date(self.date_2)
        else:
            print_date = self._format_date(self.date)
            if self.or_date or self.turn:
                print_date_2 = self._format_date(self.date_2)

        if self.type == 'B':
            print_type = self.P_DATE_OF_BIRTH
        elif self.type == 'D':
            print_type = self.P_DATE_OF_DEATH
        elif self.type == 'F':
            print_type = self.P_FLORUIT
        else:
            print('ERROR: nieokreślony typ daty.')

        if print_date == 'somevalue':
            qid = etykieta
        else:
            qid = 'LAST'

        line = f'{qid}\t{print_type}\t{print_date}'
        if self.about:
            line += f'\t{self.P_SOURCING_CIRCUMSTANCES}\t{self.Q_CIRCA}'
        if self.or_date:
            line += '\n'
            line += f'{qid}\t{print_type}\t{print_date_2}'
            if self.about:
                line += f'\t{self.P_SOURCING_CIRCUMSTANCES}\t{self.Q_CIRCA}'
        if self.turn:
            line += '\n'
            line += f'{qid}\t{print_type}\t{print_date_2}'
            if self.about:
                line += f'\t{self.P_SOURCING_CIRCUMSTANCES}\t{self.Q_CIRCA}'
        if self.after:
            line += f'\t{self.P_EARLIEST_DATE}\t{print_kw_date}'

        if self.before and self.after: # dla opisu: zm. w lub po 1458 a przed 1467
            line += f'\t{self.P_LATEST_DATE}\t{print_kw_date_2}'
        elif self.before:
            line += f'\t{self.P_LATEST_DATE}\t{print_kw_date}'

        if self.between:
            line += f'\t{self.P_EARLIEST_DATE}\t{print_kw_date}'
            line += f'\t{self.P_LATEST_DATE}\t{print_kw_date_2}'
        if self.beginning_of:
            line += f'\t{self.P_REFINE_DATE}\t{self.Q_BEGINNING_OF}'
        if self.middle_of:
            line += f'\t{self.P_REFINE_DATE}\t{self.Q_MIDDLE_OF}'
        if self.end_of:
            line += f'\t{self.P_REFINE_DATE}\t{self.Q_END_OF}'
        if self.first_half:
            line += f'\t{self.P_REFINE_DATE}\t{self.Q_FIRST_HALF}'
        if self.second_half:
            line += f'\t{self.P_REFINE_DATE}\t{self.Q_SECOND_HALF}'
        if self.first_quarter:
            line += f'\t{self.P_REFINE_DATE}\t{self.Q_FIRST_QUARTER}'

        line += '\n'
        return line


def ustal_etykiete_biogramu(value: str, title_value: str) -> str:
    """ ustala etykiete biogramu do wyszukiwania elementów biogramów
        w Wikibase
        value - zawartość nawiasu z danymi bibliograficznymi z indeksu BB
        title_value - tytuł biogramu
    """
    l_nawias = value.split(",")
    if len(l_nawias) != 4:
        print(f'ERROR: {l_nawias}')
        sys.exit(1)
    autor = text_clear(l_nawias[0])
    autor_in_title = short_names_in_autor(autor)
    tom = text_clear(l_nawias[1])
    tom = tom.replace("t.","").strip()
    strony = text_clear(l_nawias[3])

    if ";" in autor_in_title:
        autor_in_title = autor_in_title.replace(';', ',')

    return f"{autor_in_title}, {title_value}, w: PSB {tom}, {strony}"


def load_wyjatki(path: str) -> dict:
    """ load wyjatki - funkcja ładuje dane z arkusza xlsx z zapisanymi
        identyfikatorami VIAF ustalonymi ręcznie dla osób
        path - ściezka do pliku xlsx
    """
    result = {}

    try:
        work_book = load_workbook(path)
    except IOError:
        print(f"ERROR. Can't open and process file: {path}")
        sys.exit(1)

    sheet = work_book['Arkusz1']
    columns = {'POSTAĆ':0, 'VIAF':1}
    for current_row in sheet.iter_rows(2, sheet.max_row):
        u_osoba = current_row[columns['POSTAĆ']].value
        u_viaf = current_row[columns['VIAF']].value
        if u_osoba:
            u_osoba = u_osoba.strip()
        if u_viaf:
            u_viaf = u_viaf.strip()
        if u_osoba and u_viaf:
            result[u_osoba.strip()] = u_viaf.strip()
        else:
            break

    return result


def diff_date(one: str, two: str, tolerancja = 3) -> bool:
    """ badanie czy różnica lat nie jest zbyt duża """
    result = True
    if '-' in one:
        tmp = one.split('-')
        one = tmp[0]
    if '-' in two:
        tmp = two.split('-')
        two = tmp[0]

    if len(one) >=4 and len(two) >=4:
        # jeżeli obie daty są conajmniej roczne a różnica jest
        # wieksza od założonej tolerancji (domyślnie 3) to zapewne 
        # coś jest nie tak (np. znaleziono nie ten viaf id) 
        if abs(int(one[:4]) - int(two[:4])) > tolerancja:
            result = False
    return result

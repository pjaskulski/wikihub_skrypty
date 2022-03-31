""" autorzy.xlsx -> imiona i nazwiska do QuickStatements """

import sys
import os
import pickle
from time import sleep
from pathlib import Path
from openpyxl import load_workbook
from wikibaseintegrator.wbi_config import config as wbi_config
from wikidariahtools import element_search, gender_detector

# adresy
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

P_INSTANCE_OF = 'P47'
Q_MALE_NAME = 'Q987'
Q_FEMALE_NAME = 'Q988'
Q_FAMILY_NAME = 'Q34'

# listy na znalezione imiona i nazwiska
IMIONA = []
NAZWISKA = []

# słownik na określenie płci imienia 
NAME_GENDER = {}

# słowniki na QID imion i nazwisk
IMIONA_QID = {}
NAZWISKA_QID = {}

WYJATKI_IMIONA = {'Dwornik Gutowska Ewa':
                    {'imie':'Ewa', 'nazwisko':'Dwornik', 'nazwisko2':'Gutowska'},
                  'Adamczyk Prengel Irena':
                    {'imie':'Irena', 'nazwisko':'Adamczyk', 'nazwisko2':'Prengel'},
                  'Gozdawa Gołębiowski Jan':
                    {'imie':'Jan','nazwisko':'Gozdawa','nazwisko2':'Gołębiowski'},
                  'Odrowąż Pieniążek Janusz':
                    {'imie':'Janusz', 'nazwisko':'Odrowąż', 'nazwisko2':'Pieniążek'},
                  'Marciniak Jadwiga Puchała':
                    {'imie':'Jadwiga','nazwisko':'Marciniak','nazwisko2':'Puchała'},
                  'Krause Ignacy J.T.':
                    {'imie':'Ignacy', 'nazwisko':'Krause'}
                }

MALE_FEMALE_NAME = ['Maria', 'Anna']

LOAD_DICT = True
SAVE_DICT = True

class Autor:
    """" klasa Autor """

    def __init__(self, p_etykieta: str = '', alias: str = '', p_imie: str = '',
                 p_nazwisko: str = ''):
        """ init """
        self.etykieta = p_etykieta.strip()
        self._alias = alias
        self.imie = p_imie.strip()
        self.imie2 = ''
        self.nazwisko = p_nazwisko.strip()
        self.nazwisko2 = ''
        self.viaf = ''
        self.viaf_url = ''
        self.birth_date = ''
        self.death_date = ''

    @property
    def alias(self):
        """ alias """
        return self._alias

    @alias.setter
    def alias(self, value: str):
        """ alias """
        if value:
            value = ' '.join(value.strip().split())
            lista = value.split("|")
            lista = [change_name_forname(item) for item in lista]
            self._alias = lista
        else:
            self._alias = []


def is_inicial(value: str) -> bool:
    """ sprawdza czy przekazany tekst jest inicjałem imienia """
    result = False
    if len(value) == 2 and value[0].isupper() and value.endswith("."):
        result = True

    return result


def change_name_forname(name: str) -> str:
    """ change_name_forname"""
    name_parts = name.split(" ")
    if len(name_parts) == 2 and name_parts[0][0].isupper() and name_parts[1][0].isupper():
        result_name = name_parts[1] + " " + name_parts[0]
    elif (len(name_parts) == 3 and name_parts[0][0].isupper() and name_parts[1][0].isupper()
              and name_parts[2][0].isupper()):
        result_name = name_parts[1] + " " + name_parts[2] + " " + name_parts[0]
    else:
        print(f"ERROR: {name}")
    return result_name


if __name__ == "__main__":
    xlsx_path = Path('.').parent / 'data/autorzy.xlsx'
    output_imiona = Path('.').parent / 'out/autorzy_imiona.qs'
    output_nazwiska = Path('.').parent / 'out/autorzy_nazwiska.qs'
    imiona_qid_pickle = Path('.').parent / 'out/imiona_qid.pickle'
    nazwiska_qid_pickle = Path('.').parent / 'out/nazwiska_qid.pickle'

    # odmrażanie słowników QID dla imion i nazwisk
    if LOAD_DICT:
        if os.path.isfile(imiona_qid_pickle):
            with open(imiona_qid_pickle, 'rb') as handle:
                IMIONA_QID = pickle.load(handle)
        if os.path.isfile(nazwiska_qid_pickle):
            with open(nazwiska_qid_pickle, 'rb') as handle:
                NAZWISKA_QID = pickle.load(handle)

    try:
        wb = load_workbook(xlsx_path)
    except IOError:
        print(f"ERROR. Can't open and process file: {xlsx_path}")
        sys.exit(1)

    ws = wb['Arkusz1']

    col_names = {'NAZWA WŁAŚCIWA':0, 'NAZWA WARIANTYWNA (znany też jako)':1, 'Drugie': 2}

    max_row = ws.max_row
    #max_row = 50
    for row in ws.iter_rows(2, max_row):
        osoba = row[col_names['NAZWA WŁAŚCIWA']].value
        #osoba_alias = row[col_names['NAZWA WARIANTYWNA (znany też jako)']].value
        osoba_drugie = row[col_names['Drugie']].value

        if osoba:
            osoba = osoba.strip()
            autor = Autor()
            osoba = ' '.join(osoba.strip().split()) # podwójne, wiodące i kończące spacje
            tmp = osoba.split(" ")

            if osoba in WYJATKI_IMIONA:
                autor.imie = WYJATKI_IMIONA[osoba]['imie']
                autor.nazwisko = WYJATKI_IMIONA[osoba]['nazwisko']
                if 'nazwisko2' in WYJATKI_IMIONA[osoba]:
                    autor.nazwisko2 = WYJATKI_IMIONA[osoba]['nazwisko2']
                autor.etykieta = autor.imie + ' ' + autor.nazwisko + ' ' + autor.nazwisko2
            if len(tmp) == 2 and tmp[0][0].isupper() and tmp[1][0].isupper():
                autor.etykieta = tmp[1] + " " + tmp[0]
                autor.imie = tmp[1]
                autor.nazwisko = tmp[0]
            elif "Szturm de Sztrem" in osoba:
                autor.etykieta = "Tadeusz Szturm de Sztrem"
                autor.imie = "Tadeusz"
                autor.nazwisko = "Szturm de Sztrem"
            elif "Kurde-Banowska" in osoba:
                autor.etykieta = "Hanna Kurde-Banowska Lutzowa"
                autor.imie = "Hanna"
                autor.nazwisko = "Kurde-Banowska Lutzowa"
            elif (len(tmp) == 3 and tmp[0][0].isupper() and tmp[1][0].isupper()
                    and tmp[2][0].isupper()):
                # zastąpienie inicjału drugim imieniem jeżeli znane
                if len(tmp[2]) == 2 and tmp[2].endswith('.') and osoba_drugie:
                    tmp[2] = osoba_drugie
                autor.etykieta = tmp[1] + " " + tmp[2] + " " + tmp[0]
                autor.imie = tmp[1]
                # drugie imię
                if len(tmp[2]) > 2:
                    autor.imie2 = tmp[2]
                autor.nazwisko = tmp[0]
            elif tmp[0].startswith('d’'):
                autor.etykieta = tmp[1] + " " + tmp[0]
                autor.imie = tmp[1]
                autor.nazwisko = tmp[0]
            else:
                print(f'ERROR: {osoba}')

            if "_" in autor.etykieta:
                autor.etykieta = autor.etykieta.replace("_", "-")

            if "_" in autor.nazwisko:
                autor.nazwisko = autor.nazwisko.replace("_", "-")

            # nie tworzymy elementów inicjałów imienia
            if autor.imie and not is_inicial(autor.imie) and autor.imie not in IMIONA:
                IMIONA.append(autor.imie)
            if autor.imie2 and not is_inicial(autor.imie2) and autor.imie2 not in IMIONA:
                IMIONA.append(autor.imie2)
            if autor.nazwisko and autor.nazwisko not in NAZWISKA:
                NAZWISKA.append(autor.nazwisko)
            if autor.nazwisko2 and autor.nazwisko2 not in NAZWISKA:
                NAZWISKA.append(autor.nazwisko2)

            print(f'Przetworzono: {autor.etykieta}')

    # weryfikacja imion w wikibase
    for imie in IMIONA:
        print(f'Weryfikacja imienia: {imie}')
        if imie in IMIONA_QID:
            ok = True
            qid = IMIONA_QID[imie]
        else:   
            sleep(0.03) # mały odstęp między poszukiwaniami
            gender = gender_detector(imie)
            ok, qid = element_search(imie, 'item', 'pl', description=gender)
            if ok:
                IMIONA_QID[imie] = qid

        if ok:
            print(f'Znaleziono: {imie} w Wikibase: {qid}.')
            IMIONA.remove(imie)

    # IMIONA = set(IMIONA) # zbiór zawiera tylko unikalne (kontrola jest też wyżej)

    # zapis imiona Quickstatements w pliku
    print('Zapis quickstatements dla imion...')
    with open(output_imiona, "w", encoding='utf-8') as f:
        for imie in sorted(IMIONA):
            #męskie czy żeńskie?
            f.write('CREATE\n')
            f.write(f'LAST\tLpl\t"{imie}"\n')
            f.write(f'LAST\tLen\t"{imie}"\n')
            gender = gender_detector(imie)
            if gender == 'imię męskie':
                f.write('LAST\tDpl\t"imię męskie"\n')
                f.write('LAST\tDen\t"male given name"\n')
                f.write(f'LAST\t{P_INSTANCE_OF}\t{Q_MALE_NAME}\n')
            elif gender == 'imię żeńskie':
                f.write('LAST\tDpl\t"imię żeńskie"\n')
                f.write('LAST\tDen\t"female given name"\n')
                f.write(f'LAST\t{P_INSTANCE_OF}\t{Q_FEMALE_NAME}\n')

            # wybrane imiona są także w wariantach 'męskich'
            if imie in MALE_FEMALE_NAME:
                ok, qid = element_search(imie, 'item', 'pl', description='imię męskie')
                if not ok:
                    f.write('CREATE\n')
                    f.write(f'LAST\tLpl\t"{imie}"\n')
                    f.write(f'LAST\tLen\t"{imie}"\n')
                    f.write('LAST\tDpl\t"imię męskie"\n')
                    f.write('LAST\tDen\t"male given name"\n')
                    f.write(f'LAST\t{P_INSTANCE_OF}\t{Q_MALE_NAME}\n')

    # weryfikacja nazwisk w wikibase
    for nazwisko in NAZWISKA:
        print(f'Weryfikacja nazwiska: {nazwisko}')
        if nazwisko in NAZWISKA_QID:
            ok = True
            qid = NAZWISKA_QID[imie]
        else:
            sleep(0.05) # mały odstęp między poszukiwaniami
            ok, qid = element_search(nazwisko, 'item', 'pl', description='nazwisko')
            if ok:
                NAZWISKA_QID[nazwisko] = qid

        if ok:
            print(f'Znaleziono: {nazwisko} w Wikibase: {qid}.')
            NAZWISKA.remove(nazwisko)

    # NAZWISKA = set(NAZWISKA) # zbiór zawiera tylko unikalne (kontrola jest też wyżej)

    # zapis nazwisk Quickstatements w pliku 
    print('Zapis quickstatements dla nazwisk...')
    with open(output_nazwiska, "w", encoding='utf-8') as f:
        for nazwisko in NAZWISKA:
            f.write('CREATE\n')
            f.write(f'LAST\tLpl\t"{nazwisko}"\n')
            f.write(f'LAST\tLen\t"{nazwisko}"\n')
            f.write('LAST\tDpl\t"nazwisko"\n')
            f.write('LAST\tDen\t"family name"\n')
            f.write(f'LAST\t{P_INSTANCE_OF}\t{Q_FAMILY_NAME}\n')

    # zamrażanie słowników imion i nazwisk znalezionych w wikibase 
    if SAVE_DICT:
        with open(imiona_qid_pickle, 'wb') as handle:
            pickle.dump(IMIONA_QID, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(nazwiska_qid_pickle, 'wb') as handle:
            pickle.dump(NAZWISKA_QID, handle, protocol=pickle.HIGHEST_PROTOCOL)

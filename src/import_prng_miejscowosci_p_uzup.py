""" import miejscowosci z pliku miejscowosciP.xlsx z danymi z PRG"""
import os
import os.path
import time
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import element_search_adv, get_properties, get_elements
from wikidariahtools import read_qid_from_text
from property_import import create_statement_data


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

WIKIBASE_WRITE = True

# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':
    # standardowe właściwości
    print('Przygotowanie właściwości...')
    properties = get_properties(['instance of', 'stated as', 'reference URL', 'retrieved',
                                'id SDI', 'part of', 'has part or parts', 'TERYT', 'settlement type',
                                'coordinate location', 'located in the administrative territorial entity',
                                'name status', 'inflectional ending', 'adjective form',
                                'located in the administrative territorial entity'
                                ])

    # elementy definicyjne
    print('Przygotowanie elementów definicyjnych...')
    elements = get_elements(['unofficial name', 'human settlement',
        'part of a colony', 'part of a city', 'part of a settlement', 'part of a village',
        'colony', 'colony of a colony', 'colony of a settlement', 'colony of a village',
        'city/town', 'settlement', 'settlement of a colony', 'forest settlement',
        'forest settlement of a village', 'settlement of a settlement', 'settlement of a village',
        'housing developments', 'housing estate of a village', 'hamlet', 'hamlet of a colony',
        'hamlet of a settlement', 'hamlet of a village', 'tourist shelter', 'village', "forester's lodge"
                            ])

    settlement_type_map = {}
    settlement_type_map['część kolonii'] = 'part of a colony'
    settlement_type_map['część miasta'] = 'part of a city'
    settlement_type_map['część osady'] = 'part of a settlement'
    settlement_type_map['część wsi'] = 'part of a village'
    settlement_type_map['kolonia'] = 'colony'
    settlement_type_map['kolonia kolonii'] = 'colony of a colony'
    settlement_type_map['kolonia osady'] = 'colony of a settlement'
    settlement_type_map['kolonia wsi'] = 'colony of a village'
    settlement_type_map['miasto'] = 'city/town'
    settlement_type_map['osada'] = 'settlement'
    settlement_type_map['osada kolonii'] = 'settlement of a colony'
    settlement_type_map['osada leśna'] = 'forest settlement'
    settlement_type_map['osada leśna wsi'] = 'forest settlement of a village'
    settlement_type_map['osada osady'] = 'settlement of a settlement'
    settlement_type_map['osada wsi'] = 'settlement of a village'
    settlement_type_map['osiedle'] = 'housing developments'
    settlement_type_map['osiedle wsi'] = 'housing estate of a village'
    settlement_type_map['przysiółek'] = 'hamlet'
    settlement_type_map['przysiółek kolonii'] = 'hamlet of a colony'
    settlement_type_map['przysiółek osady'] = 'hamlet of a settlement'
    settlement_type_map['przysiółek wsi'] = 'hamlet of a village'
    settlement_type_map['schronisko turystyczne'] = 'tourist shelter'
    settlement_type_map['wieś'] = 'village'
    settlement_type_map['leśniczówka'] = "forester's lodge"


    # wspólna referencja dla wszystkich deklaracji z PRG
    references = {}
    references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'
    references[properties['retrieved']] = '2022-09-23'

    # logowanie do instancji wikibase
    if WIKIBASE_WRITE:
        login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                         consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                         access_token=WIKIDARIAH_ACCESS_TOKEN,
                                         access_secret=WIKIDARIAH_ACCESS_SECRET,
                                         token_renew_period=14400)

    xlsx_input = '../data_prng/miejscowosciP.xlsx'
    qid_output = '../data_prng/miejscowosciP_qid.txt'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["miejscowosciP"]

    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1


    parts = {}

    lista_brakow = [
        16816, 16817, 16818, 16819, 16820, 16821, 16822, 16816, 16817,
        16818, 16819, 16820, 16821, 16822, 16823, 16824, 16825, 16826, 16827,
        16828, 16829, 16830, 16831, 16832, 16833, 16834, 16835, 16836, 16837,
        16838, 16839, 16840, 16841, 16842, 16843, 16844, 16845, 16846, 16847,
        16848, 16849, 16850, 16851, 16853, 16854, 16855, 16856, 16857, 16858,
        16859, 16860, 16861, 16862, 16863, 16864, 16865, 16866, 16867, 16868,
        16869, 16870, 16871, 16872, 16873, 16874, 16875, 16876, 21564, 22118,
        22160, 22440, 22447, 22448, 22449, 22469, 22471, 22472, 22574, 22579,
        22619, 22664, 22703
    ]

    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1
        if index not in lista_brakow:
            continue
        # wczytanie danych z xlsx
        nazwa = row[col_names['NAZWAGLOWN']].value
        if not nazwa:
            continue

        label_pl = label_en = nazwa

        dopelniacz = row[col_names['DOPELNIACZ']].value
        przymiotni = row[col_names['PRZYMIOTNI']].value
        nazwy_obocz = row[col_names['NAZWYOBOCZ']].value
        nazwy_dodat = row[col_names['NAZWYDODAT']].value
        nazwy_histo = row[col_names['NAZWYHISTO']].value
        nazwa_miejsc = row[col_names['NAZWAMIEJS']].value
        rodzajobie = row[col_names['RODZAJOBIE']].value
        wgs84 = row[col_names['WGS84']].value
        identyfi_2 = row[col_names['IDENTYFI_2']].value
        idiip = row[col_names['IDIIP']].value
        gmina = row[col_names['GMINA']].value
        if gmina:
            gmina = gmina.split('-gmina')[0]
        powiat = row[col_names['POWIAT']].value
        wojewodztw = row[col_names['WOJEWODZTW']].value

        rodzaje_czesci_miejscowosci = ['część wsi', 'przysiółek osady', 'kolonia wsi',
                                       'część miasta', 'część kolonii', 'przysiółek wsi',
                                       'część osady', 'kolonia wsi', 'osada kolonii',
                                       'osada osady', 'osada wsi', 'osiedle wsi',
                                       'przysiółek kolonii', 'przysiółek osady', 'osada leśna wsi',
                                       'kolonia koloni', 'kolonia osady']
        if rodzajobie in rodzaje_czesci_miejscowosci:
            description_pl = f'{rodzajobie}: {nazwa_miejsc} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'
            description_en = f'{rodzajobie}: {nazwa_miejsc} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'
        else:
            description_pl = f'{rodzajobie} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'
            description_en = f'{rodzajobie} (gmina: {gmina}, powiat: {powiat}, wojewódzwo: {wojewodztw})'

        # przygotowanie struktur wikibase
        data = []
        aliasy = {}

        # instance of
        statement = create_statement_data(properties['instance of'], elements['human settlement'],
                                          None, None, add_ref_dict=None)
        if statement:
            data.append(statement)

        # stated as NAZWA_GLOWN
        qualifiers = {}
        if dopelniacz:
            qualifiers[properties['inflectional ending']] = dopelniacz
        if przymiotni:
            qualifiers[properties['adjective form']] = przymiotni
        qualifiers[properties['name status']] = elements['unofficial name']

        statement = create_statement_data(properties['stated as'], f'pl:"{nazwa}"', None,
                                          qualifiers, add_ref_dict=references)
        if statement:
            data.append(statement)

        # NAZWY_OBOCZ
        if nazwy_obocz:
            tmp = nazwy_obocz.split(',')
            for tmp_item in tmp:
                tmp_item = tmp_item.strip()
                if 'pl' in aliasy:
                    aliasy['pl'].append(tmp_item)
                else:
                    aliasy['pl'] = [tmp_item]
                statement = create_statement_data(properties['stated as'], f'pl:"{tmp_item}"',
                    None, None, add_ref_dict=references)
                if statement:
                    data.append(statement)

        # NAZWYHISTO
        if nazwy_histo:
            tmp = nazwy_histo.split(',')
            for tmp_item in tmp:
                tmp_item = tmp_item.strip()
                if 'pl' in aliasy:
                    aliasy['pl'].append(tmp_item)
                else:
                    aliasy['pl'] = [tmp_item]
                statement = create_statement_data(properties['stated as'], f'pl:"{tmp_item}"',
                    None, None, add_ref_dict=references)
                if statement:
                    data.append(statement)

        # RODZAJOBIE
        statement = create_statement_data(properties['settlement type'], elements[settlement_type_map[rodzajobie]],
            None, None, add_ref_dict=references)
        if statement:
            data.append(statement)

        # WGS84 - Point (23.29833332 52.68194448)
        if wgs84:
            wgs84 = wgs84.replace('Point', '').replace('(', '').replace(')','').strip()
            tmp = wgs84.split(' ')
            longitude = tmp[0]
            latitude = tmp[1]
            coordinate = f'{latitude},{longitude}'
            statement = create_statement_data(properties['coordinate location'], coordinate,
                None, None, add_ref_dict=references)
            if statement:
                data.append(statement)

        # IDENTYFI_2
        if identyfi_2:
            parameters = [(properties['TERYT'], identyfi_2)]
            ok, gmina_qid = element_search_adv(gmina, 'en', parameters)
            if ok:
                statement = create_statement_data(properties['located in the administrative territorial entity'],
                    gmina_qid, None, None, add_ref_dict=references)
                if statement:
                    data.append(statement)

        # id SDI
        if idiip:
            statement = create_statement_data(properties['id SDI'], idiip, None, None, add_ref_dict=references)
            if statement:
                data.append(statement)

        description_en = f'{description_en} [{coordinate}]'
        description_pl = f'{description_pl} [{coordinate}]'

        # etykiety, description, aliasy
        wb_item = wbi_core.ItemEngine(new_item=True, data=data)
        wb_item.set_label(label_en, lang='en')
        wb_item.set_label(label_pl, lang='pl')

        wb_item.set_description(description_en, 'en')
        wb_item.set_description(description_pl, 'pl')

        if aliasy:
            for alias_lang, alias_value in aliasy.items():
                for alias_item in alias_value:
                    wb_item.set_aliases(alias_item, alias_lang)

        if WIKIBASE_WRITE:
            test = 1
            while True:
                try:
                    new_id = wb_item.write(login_instance, bot_account=True, entity_type='item')
                    print(f'{index}/{ws.max_row - 1} Dodano nowy element: {label_en} / {label_pl} = {new_id}')
                    with open(qid_output, 'a', encoding='utf-8') as f:
                        f.write(f"{index};{label_en};{new_id}\n")
                    break
                except MWApiError as wbdelreference_error:
                    err_code = wbdelreference_error.error_msg['error']['code']
                    message = wbdelreference_error.error_msg['error']['info']
                    if err_code in ['assertuserfailed', 'badtoken']:
                        if test == 1:
                            print('Generate edit credentials...')
                            login_instance.generate_edit_credentials()
                            test += 1
                            continue
                    else:
                        print(err_code, message)
                    sys.exit(1)
        else:
            # wyszukiwanie po etykiecie, właściwości instance of oraz po opisie
            parameters = [(properties['instance of'], elements['human settlement'])]
            ok, item_id = element_search_adv(label_en, 'en', parameters, description_en, max_results_to_verify=500)
            if not ok:
                new_id = 'TEST'
                print(f"{index}/{ws.max_row - 1} Przygotowano dodanie elementu - {label_en} / {label_pl}  = {new_id}")
            else:
                print(f'{index}/{ws.max_row - 1} Element: {label_en} / {label_pl} już istnieje: {item_id}')

        # czy to pomoże na zrywanie połączenia?
        time.sleep(0.05)

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

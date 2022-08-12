""" dane dla właściwości part of z bazy
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import psycopg2
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login


# adresy wikibase
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# lista elementów (typów jednostek administracyjnych)
administrative_types = ['Q79902', 'Q79903', 'Q79904', 'Q79905', 'Q79906', 'Q79907',
                     'Q79896', 'Q79897', 'Q79908', 'Q79909', 'Q79910', 'Q79911',
                     'Q79912', 'Q79913', 'Q79914', 'Q79915', 'Q79916', 'Q79917',
                     'Q79918', 'Q79919', 'Q79920', 'Q79921', 'Q79922', 'Q79923',
                     'Q79924', 'Q79925', 'Q79926', 'Q79927', 'Q79928', 'Q79929',
                     'Q79930', 'Q79931', 'Q79932', 'Q79933', 'Q79934', 'Q79935',
                     'Q79936', 'Q79937', 'Q79938', 'Q79939', 'Q79940', 'Q79941',
                     'Q79942', 'Q79943', 'Q79944', 'Q79945', 'Q79946', 'Q79947',
                     'Q79948', 'Q79949', 'Q79950', 'Q79951', 'Q79952', 'Q79953',
                     'Q79954', 'Q79955', 'Q79956', 'Q79957', 'Q79958', 'Q79959',
                     'Q79960', 'Q79961', 'Q79962', 'Q79963', 'Q79964', 'Q79965',
                     'Q79966', 'Q79967', 'Q79968', 'Q79969', 'Q79970', 'Q79971',
                     'Q79972', 'Q79973', 'Q79974', 'Q79975', 'Q79976', 'Q79977',
                     'Q79978', 'Q79979', 'Q79980', 'Q79981', 'Q79982', 'Q79983',
                     'Q79984', 'Q79985', 'Q79986', 'Q79987', 'Q79988', 'Q79989',
                     'Q79990', 'Q79991', 'Q79992', 'Q79993', 'Q79994', 'Q79995',
                     'Q79996', 'Q79997', 'Q79998', 'Q79999', 'Q80000', 'Q80001',
                     'Q80002', 'Q80003', 'Q80004', 'Q80005', 'Q80006', 'Q80007',
                     'Q80008', 'Q80009', 'Q80010', 'Q80011', 'Q80012', 'Q80013',
                     'Q80014', 'Q80015', 'Q80016', 'Q80017', 'Q80018', 'Q80019',
                     'Q80020', 'Q80021', 'Q80022', 'Q80023', 'Q80024', 'Q80025',
                     'Q80026', 'Q80027', 'Q80028', 'Q80029', 'Q80030', 'Q80031',
                     'Q80032', 'Q80033', 'Q80034', 'Q80035', 'Q80036', 'Q80037',
                     'Q80038', 'Q80039', 'Q80040', 'Q80041', 'Q80042', 'Q80043',
                     'Q80044', 'Q80045', 'Q80046', 'Q80047', 'Q80048', 'Q80049',
                     'Q80050', 'Q80051', 'Q80052', 'Q80053', 'Q80054', 'Q80055',
                     'Q80056', 'Q80057', 'Q80058', 'Q80059', 'Q80060', 'Q80061',
                     'Q80062', 'Q80063', 'Q80064', 'Q80065', 'Q80066', 'Q80067',
                     'Q80068', 'Q80069', 'Q80070', 'Q80071', 'Q80072', 'Q80073',
                     'Q80074', 'Q80075', 'Q80076', 'Q80077', 'Q80078', 'Q80079',
                     'Q80080', 'Q80081', 'Q80082', 'Q80083', 'Q80084', 'Q80085',
                     'Q80086', 'Q80087', 'Q80088', 'Q80089', 'Q80090', 'Q80091',
                     'Q80092', 'Q80093', 'Q80094', 'Q80095', 'Q80096', 'Q80097',
                     'Q80098', 'Q80099', 'Q80100', 'Q80101', 'Q80102', 'Q80103',
                     'Q80104', 'Q80105', 'Q80106', 'Q80107', 'Q80108', 'Q80109',
                     'Q80110', 'Q80111', 'Q80112', 'Q80113', 'Q80114', 'Q80115',
                     'Q80116', 'Q80117', 'Q80118', 'Q80119', 'Q80120', 'Q80121',
                     'Q80122', 'Q80123', 'Q80124', 'Q80125', 'Q80126', 'Q80127',
                     'Q80128', 'Q80129', 'Q80130', 'Q80131', 'Q80132', 'Q80133',
                     'Q80134', 'Q80135', 'Q80153', 'Q80154', 'Q80155', 'Q80156',
                     'Q80157', 'Q80158', 'Q80159', 'Q80160', 'Q80161', 'Q80162',
                     'Q80163', 'Q80164', 'Q80165', 'Q80166', 'Q80167', 'Q80168',
                     'Q80169', 'Q80170', 'Q80171', 'Q80172', 'Q80173', 'Q80174',
                     'Q80175', 'Q80176', 'Q80177', 'Q80178', 'Q80179', 'Q80180',
                     'Q80181', 'Q80182', 'Q80183', 'Q80184', 'Q80185', 'Q80186',
                     'Q80187', 'Q80188', 'Q80189', 'Q80190', 'Q80191', 'Q80192',
                     'Q80193', 'Q80194', 'Q80195', 'Q80196', 'Q80197', 'Q80198',
                     'Q80199', 'Q80200']

def create_sheet(my_workbook, sheet_name, my_columns):
    """ create sheet"""
    col_char = ['A','B','C','D','E','F','G','H','I','J']

    my_new_sheet = my_workbook.create_sheet(sheet_name)
    for index, column in enumerate(my_columns):
        cell_id = f'{col_char[index]}1'
        my_new_sheet[cell_id] = column
        my_new_sheet[cell_id].font = Font(bold=True)
        my_new_sheet[cell_id].alignment = Alignment(horizontal='center')

    return my_new_sheet


# --------------------------------- MAIN ---------------------------------------

if __name__ == "__main__":
    # login i hasło ze zmiennych środowiskowych
    env_path = Path(".") / ".env"
    load_dotenv(dotenv_path=env_path)

    DB_LOGIN = os.environ.get("DB_LOGIN")
    DB_PASSWORD = os.environ.get("DB_PASSWORD")
    DB_HOST = os.environ.get("DB_HOST")
    DB_PORT = os.environ.get("DB_PORT")
    DB_DATABASE = os.environ.get("DB_DATABASE")

    BOT_LOGIN = os.environ.get('WIKIDARIAH_USER')
    BOT_PASSWORD = os.environ.get('WIKIDARIAH_PWD')

    # logowanie do instancji wikibase
    login_instance = wbi_login.Login(user=BOT_LOGIN, pwd=BOT_PASSWORD)

    # tymczasowo - wczytanie słownika QID/Purl (odwrotnie)
    purl = {}
    purl_inv = {}

    with open("temp.txt", "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()
        tmp = line.split(" = ")
        purl[tmp[1]] = tmp[0]       # układ QID - Purl
        purl_inv[tmp[0]] = tmp[1]   # układ Purl - QID

    # połączenie do serwera DB
    conn = psycopg2.connect(
    database=DB_DATABASE,
    user=DB_LOGIN,
    password=DB_PASSWORD,
    host=DB_HOST,
    port= DB_PORT
    )

    # utworzenie kursora
    cursor = conn.cursor()

    wb = Workbook()
    columns = ['Label_en', 'Purl identifier', 'Label_txt', 'P', 'Value', 'Purl value', 'Value_txt', 'Qualifier', 'Qualifier_value']
    ws_q_s = create_sheet(wb, 'Q_statements', columns)
    columns = ['Label_en', 'Label_pl', 'Purl identifier', 'Description_en', 'Description_pl', 'Wiki_id', 'StartsAt', 'EndsAt', 'Instance of']
    ws_q_l = create_sheet(wb, 'Q_list', columns)
    columns = ['Label_en', 'Label_pl', 'Datatype', 'Description_en', 'Description_pl', 'Wiki_id', 'Inverse_property']
    ws_p_l = create_sheet(wb, 'P_list', columns)
    columns = ['Label_en', 'P', 'Value', 'Reference_property', 'Reference_value']
    ws_p_s = create_sheet(wb, 'P_statements', columns)

    for item in administrative_types:
        purl_id = purl[item]
        if not 'administrative_type' in purl_id:
            continue

        tmp = purl_id.split('_')
        search_id = int(tmp[-1])
        wb_item = wbi_core.ItemEngine(item_id=item)
        label_pl = wb_item.get_label('pl')

        # zapytanie zwraca identyfikator typu jednostki nadrzędnej
        sql = f"""
            SELECT "Identifiers", "LevelsLinkUp", "Types"
            FROM ontology."AdministrativeUnitTypesNamesDictionary"
            WHERE "Identifiers" = {search_id}
        """

        cursor.execute(sql)
        results = cursor.fetchall()

        tablica = []
        subclass = []
        for result in results:
            if result[1]:
                if result[2] == 'type':
                    tablica.append(int(result[1]))
                elif result[2] == 'subtype':
                    subclass.append(int(result[1]))

        for identifier in tablica:
            part_of_purl = f'http://purl.org/ontohgis#administrative_type_{identifier}'
            part_of_qid = purl_inv[part_of_purl]
            wb_value = wbi_core.ItemEngine(item_id=part_of_qid)
            value_label_pl = wb_value.get_label('pl')
            row = {
                1: item,
                2: purl_id,
                3: label_pl,
                4:'part of',
                5: part_of_qid,
                6: part_of_purl,
                7: value_label_pl
            }
            ws_q_s.append(row)

        for identifier in subclass:
            part_of_purl = f'http://purl.org/ontohgis#administrative_type_{identifier}'
            part_of_qid = purl_inv[part_of_purl]
            wb_value = wbi_core.ItemEngine(item_id=part_of_qid)
            value_label_pl = wb_value.get_label('pl')
            row = {
                1: item,
                2: purl_id,
                3: label_pl,
                4:'subclass of',
                5: part_of_qid,
                6: part_of_purl,
                7: value_label_pl
            }
            ws_q_s.append(row)

    # zapis pliku xlsx
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    wb.save('../data/administrative_types_part_of.xlsx')

    # zamykanie połączenia z DB
    conn.close()

""" autorzy.xlsx -> QuickStatements """

import sys
import os
import pickle
from pathlib import Path
from time import sleep
from urllib.parse import quote
import requests
from openpyxl import load_workbook
from wikidariahtools import format_date


P_INSTANCE_OF = 'P47'
Q_HUMAN = 'Q32'
P_IMIE = 'P3'
P_NAZWISKO = 'P4'
P_VIAF = 'P79'
P_REFERENCE_URL = 'S182'
P_DATE_OF_BIRTH = 'P7'
P_DATE_OF_DEATH = 'P8'

# słownik na znalezione identyfikatory
VIAF_ID = {}
VIAF_BIRTH = {}
VIAF_DEATH = {}
WERYFIKACJA_VIAF = {}
WYJATKI = {}

LOAD_DICT = True
SAVE_DICT = True


class Autor:
    """" klasa Autor """

    def __init__(self, p_etykieta: str = '', alias: str = '', p_imie: str = '',
                 p_nazwisko: str = ''):
        """ init """
        self.etykieta = p_etykieta.strip()
        self._alias = alias
        self.imie = p_imie.strip()
        self.imie2 = ''
        self.nazwisko = p_nazwisko.strip()
        self.viaf = ''
        self.viaf_url = ''
        self.birth_date = ''
        self.death_date = ''

    @property
    def alias(self):
        """ alias """
        return self._alias

    @alias.setter
    def alias(self, value: str):
        """ alias """
        if value:
            value = ' '.join(value.strip().split())
            lista = value.split("|")
            lista = [change_name_forname(item) for item in lista]
            self._alias = lista
        else:
            self._alias = []


def get_viaf_data(v_url: str) -> tuple:
    """ get_viaf_data """
    v_id = v_birth = v_death = ''
    response = requests.get(v_url + 'viaf.json')
    result = response.json()
    if 'viafID' in result:
        v_id = result['viafID']
    if 'birthDate' in result:
        v_birth = result['birthDate']
    if 'deathDate' in result:
        v_death = result['deathDate']

    return v_id, v_birth, v_death


def viaf_search(name: str) -> tuple:
    """ szukanie identyfikatora VIAF """
    info = id_url = birthDate = deathDate = ''
    result = False

    # jeżeli osoba jest w wyjątkach to pobieramy dane ze znanego adresu
    # lub od razu NOT FOUND
    if name in WYJATKI:
        if WYJATKI[name].strip() == 'BRAK':
            return False, "NOT_FOUND", '', '', ''

        info, birthDate, deathDate = get_viaf_data(WYJATKI[name])
        id_url = WYJATKI[name]
        VIAF_ID[name] = info
        if birthDate:
            VIAF_BIRTH[name] = birthDate
        if deathDate:
            VIAF_DEATH[name] = deathDate

        return True, info, id_url, birthDate, deathDate

    # jeżeli identyfikator jest już znany to nie ma potrzeby szukania
    # przez api
    if name in VIAF_ID:
        info = VIAF_ID[name]
        id_url = f"http://viaf.org/viaf/{info}/"

        if name in VIAF_BIRTH:
            birthDate = VIAF_BIRTH[name]

        if name in VIAF_DEATH:
            deathDate = VIAF_DEATH[name]

        return True, info, id_url, birthDate, deathDate

    identyfikatory = []
    urls = {}
    base = 'https://viaf.org/viaf/search'
    format_type = 'application/json'
    search_person = quote(f'"{name}"')
    adres = f'{base}?query=local.personalNames+=+{search_person}'
    adres+= f'&local.sources+=+"plwabn"&sortKeys=holdingscount&httpAccept={format_type}'

    # mały odstęp między poszukiwaniami
    sleep(0.05)

    try:
        response = requests.get(adres)
        result = response.json()
        if 'records' in result['searchRetrieveResponse']:
            rekordy = result['searchRetrieveResponse']['records']

            for rekord in rekordy:
                v_id = rekord['record']['recordData']['viafID']
                if v_id:
                    url = rekord['record']['recordData']['Document']['@about']
                    if isinstance(rekord['record']['recordData']['mainHeadings']['data'], list):
                        label = rekord['record']['recordData']['mainHeadings']['data'][0]['text']
                    elif isinstance(rekord['record']['recordData']['mainHeadings']['data'], dict):
                        label = rekord['record']['recordData']['mainHeadings']['data']['text']

                    if label:
                        label = label.replace(",", "")
                        l_name = name.split(" ")
                        find_items = True
                        for item_name in l_name:
                            if len(item_name) > 2 and not item_name in label:
                                find_items = False
                                break
                        if find_items:
                            identyfikatory.append(v_id)
                            urls[v_id] = url
                            if 'birthDate' in rekord['record']['recordData']:
                                birthDate = rekord['record']['recordData']['birthDate']

                            if 'deathDate' in rekord['record']['recordData']:
                                deathDate = rekord['record']['recordData']['deathDate']

                            VIAF_ID[name] = v_id  # zapis identyfikatora w słowniku

                            if birthDate:
                                VIAF_BIRTH[name] = birthDate
                            if deathDate:
                                VIAF_DEATH[name] = deathDate

                            break

    except requests.exceptions.RequestException as e_info:
        print(f'Name: {name} ERROR {e_info}')

    if len(identyfikatory) == 1:
        return True, identyfikatory[0], urls[identyfikatory[0]], birthDate, deathDate

    return False, "NOT FOUND", '', '', ''


def is_inicial(imie) -> bool:
    """ sprawdza czy przekazany tekst jest inicjałem imienia """
    result = False
    if len(imie) == 2 and imie[0].isupper() and imie.endswith("."):
        result = True

    return result


def change_name_forname(name: str) -> str:
    """ change_name_forname"""
    name_parts = name.split(" ")
    if len(name_parts) == 2 and name_parts[0][0].isupper() and name_parts[1][0].isupper():
        result_name = name_parts[1] + " " + name_parts[0]
    elif (len(name_parts) == 3 and name_parts[0][0].isupper() and name_parts[1][0].isupper()
              and name_parts[2][0].isupper()):
        result_name = name_parts[1] + " " + name_parts[2] + " " + name_parts[0]
    else:
        print(f"ERROR: {name}")
    return result_name


def load_wyjatki(path: str) -> dict:
    """ load wyjatki"""
    result = {}

    try:
        work_book = load_workbook(path)
    except IOError:
        print(f"ERROR. Can't open and process file: {path}")
        sys.exit(1)

    sheet = work_book['Arkusz1']
    columns = {'AUTOR':0, 'VIAF':1}
    for current_row in sheet.iter_rows(2, sheet.max_row):
        u_osoba = current_row[columns['AUTOR']].value
        u_viaf = current_row[columns['VIAF']].value
        u_osoba = u_osoba.strip()
        u_viaf = u_viaf.strip()
        if u_osoba and u_viaf:
            result[u_osoba.strip()] = u_viaf.strip()

    return result


if __name__ == "__main__":
    xlsx_path = Path('.').parent / 'data/autorzy.xlsx'
    uzup_path = Path('.').parent / 'data/autorzy_viaf_uzup.xlsx'
    output = Path('.').parent / 'out/autorzy.qs'
    log_path = Path('.').parent / 'out/autorzy_brak.log'
    autorzy_pickle = Path('.').parent / 'out/autorzy.pickle'
    birth_pickle = Path('.').parent / 'out/birth.pickle'
    death_pickle = Path('.').parent / 'out/death.pickle'
    html_path = Path('.').parent / 'out/autorzy_viaf.html'

    # odmrażanie słownika identyfikatorów VIAF
    if LOAD_DICT:
        if os.path.isfile(autorzy_pickle):
            with open(autorzy_pickle, 'rb') as handle:
                VIAF_ID = pickle.load(handle)
        if os.path.isfile(birth_pickle):
            with open(birth_pickle, 'rb') as handle:
                VIAF_BIRTH = pickle.load(handle)
        if os.path.isfile(death_pickle):
            with open(death_pickle, 'rb') as handle:
                VIAF_DEATH = pickle.load(handle)

    # wyjatki
    WYJATKI = load_wyjatki(uzup_path)

    try:
        wb = load_workbook(xlsx_path)
    except IOError:
        print(f"ERROR. Can't open and process file: {xlsx_path}")
        sys.exit(1)

    ws = wb['Arkusz1']

    col_names = {'NAZWA WŁAŚCIWA':0, 'NAZWA WARIANTYWNA (znany też jako)':1, 'Drugie': 2}

    autor_list = []
    with open(log_path, "w", encoding='utf-8') as f_log:
        max_row = ws.max_row
        #max_row = 50
        for row in ws.iter_rows(2, max_row):
            osoba = row[col_names['NAZWA WŁAŚCIWA']].value
            osoba_alias = row[col_names['NAZWA WARIANTYWNA (znany też jako)']].value
            osoba_drugie = row[col_names['Drugie']].value

            if osoba:
                autor = Autor()
                osoba = ' '.join(osoba.strip().split()) # podwójne, wiodące i kończące spacje
                tmp = osoba.split(" ")

                if len(tmp) == 2 and tmp[0][0].isupper() and tmp[1][0].isupper():
                    autor.etykieta = tmp[1] + " " + tmp[0]
                    autor.imie = tmp[1]
                    # jeżeli znamy tylko inicjał imienia to nie zakładamy Q
                    if len(autor.imie) == 2 and autor.imie.endswith("."):
                        continue
                    autor.nazwisko = tmp[0]
                elif "Szturm de Sztrem" in osoba:
                    autor.etykieta = "Tadeusz Szturm de Sztrem"
                    autor.imie = "Tadeusz"
                    autor.nazwisko = "Szturm de Sztrem"
                elif "Kurde-Banowska" in osoba:
                    autor.etykieta = "Hanna Kurde-Banowska Lutzowa"
                    autor.imie = "Hanna"
                    autor.nazwisko = "Kurde-Banowska Lutzowa"
                elif (len(tmp) == 3 and tmp[0][0].isupper() and tmp[1][0].isupper()
                        and tmp[2][0].isupper()):
                    # zastąpienie inicjału drugim imieniem
                    if len(tmp[2]) == 2 and tmp[2].endswith('.') and osoba_drugie:
                        tmp[2] = osoba_drugie
                    autor.etykieta = tmp[1] + " " + tmp[2] + " " + tmp[0]
                    autor.imie = tmp[1]
                    # drugie imię
                    if len(tmp[2]) > 2:
                        autor.imie2 = tmp[2]
                    autor.nazwisko = tmp[0]
                elif tmp[0].startswith('d’'):
                    autor.etykieta = tmp[1] + " " + tmp[0]
                    autor.imie = tmp[1]
                    autor.nazwisko = tmp[0]
                else:
                    print(f'ERROR: {osoba}')
                    f_log.write(f'ERROR: {osoba}\n')

                if "_" in autor.etykieta:
                    autor.etykieta = autor.etykieta.replace("_", "-")

                if "_" in autor.nazwisko:
                    autor.nazwisko = autor.nazwisko.replace("_", "-")

                if osoba_alias:
                    autor.alias = osoba_alias

                if "_" in autor.alias:
                    autor.alias = autor.alias.replace("_", " ")

                if autor.etykieta:
                    # szukanie VIAF
                    ok, wynik, wynik_url, birth_d, death_d = viaf_search(autor.etykieta)
                    if ok:
                        autor.viaf = wynik
                        autor.viaf_url = wynik_url
                        autor.birth_date = birth_d
                        autor.death_date = death_d
                        print(f'VIAF, {autor.etykieta}, {wynik}, {wynik_url}, {birth_d}, {death_d}')
                        WERYFIKACJA_VIAF[autor.etykieta] = wynik_url
                    else:
                        print(f'VIAF, {autor.etykieta}, {wynik}')
                        f_log.write(f'VIAF, {autor.etykieta}, {wynik}\n')

                # nie tworzymy elementów dla autorów znanych tylko z inicjału imienia
                if not is_inicial(autor.imie):
                    autor_list.append(autor)

    # zapis Quickstatements w pliku 
    with open(output, "w", encoding='utf-8') as f:
        for autor in autor_list:
            tmp_data_b = tmp_data_d = ''
            if autor.birth_date:
                if len(autor.birth_date) == 4:
                    tmp_data_b = autor.birth_date
                elif len(autor.birth_date) == 10:
                    tmp_data_b = autor.birth_date[:4]

            if autor.death_date:
                if len(autor.death_date) == 4:
                    tmp_data_d = autor.death_date
                elif len(autor.death_date) == 10:
                    tmp_data_d = autor.death_date[:4]

            do_opisu_pl = do_opisu_en = ''
            if tmp_data_b and tmp_data_d:
                do_opisu_pl = do_opisu_en = f"({tmp_data_b}-{tmp_data_d})"
            elif tmp_data_b and tmp_data_b == '1900':
                do_opisu_pl = "(XX w.)"
                do_opisu_en = '(20 c.)'
            elif tmp_data_b:    
                do_opisu_pl = do_opisu_en = f"({tmp_data_b}- )"
            elif tmp_data_d:
                do_opisu_pl = do_opisu_en = f"( -{tmp_data_d})"

            f.write('CREATE\n')

            f.write(f'LAST\tLpl\t"{autor.etykieta}"\n')
            f.write(f'LAST\tLen\t"{autor.etykieta}"\n')

            if do_opisu_pl:
                f.write(f'LAST\tDpl\t"{do_opisu_pl}"\n')
            if do_opisu_en:
                f.write(f'LAST\tDen\t"{do_opisu_en}"\n')

            f.write(f'LAST\t{P_INSTANCE_OF}\t{Q_HUMAN}\n')
            f.write(f'LAST\t{P_IMIE}\t"{autor.imie}"\n')
            if autor.imie2:
                f.write(f'LAST\t{P_IMIE}\t"{autor.imie2}"\n')
            f.write(f'LAST\t{P_NAZWISKO}\t"{autor.nazwisko}"\n')
            if autor.alias:
                for item in autor.alias:
                    f.write(f'LAST\tApl\t"{item}"\n')
                    f.write(f'LAST\tAen\t"{item}"\n')

            if autor.birth_date:
                if autor.birth_date == '1900' and (not autor.death_date or autor.death_date == '0'):
                    birth_date = "+1901-00-00T00:00:00Z/7"
                else:
                    birth_date = format_date(autor.birth_date)
                if birth_date:
                    f.write(f'LAST\t{P_DATE_OF_BIRTH}\t{birth_date}\n')
            if autor.death_date:
                death_date = format_date(autor.death_date)
                if death_date:
                    f.write(f'LAST\t{P_DATE_OF_DEATH}\t{death_date}\n')
            if autor.viaf and autor.viaf_url:
                f.write(f'LAST\t{P_VIAF}\t"{autor.viaf}"\t{P_REFERENCE_URL}\t"{autor.viaf_url}"\n')

    # zamrażanie słownika identyfikatów VIAF_ID 
    if SAVE_DICT:
        with open(autorzy_pickle, 'wb') as handle:
            pickle.dump(VIAF_ID, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(birth_pickle, 'wb') as handle:
            pickle.dump(VIAF_BIRTH, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(death_pickle, 'wb') as handle:
            pickle.dump(VIAF_DEATH, handle, protocol=pickle.HIGHEST_PROTOCOL)

    # zapis wyszukiwań VIAF w HTML dla łatwiejszej weryfikacji
    with open(html_path, "w", encoding='utf-8') as h:
        h.write('<html>\n')
        h.write('<head>\n')
        h.write('<meta charset="utf-8">\n')
        h.write('<title>Weryfikacja VIAF dla autorów biogramów</title>\n')
        h.write('</head>\n')
        h.write('<body>\n')
        h.write('<h2>Weryfikacja VIAF dla autorów biogramów</h2>\n')
        h.write('<table>\n')
        for key, val in WERYFIKACJA_VIAF.items():
            h.write(f'<tr><td>{key}</td><td><a href="{val}">{val}</td></tr>\n')
        h.write('</table>\n')
        h.write('</body>\n')
        h.write('</html>\n')

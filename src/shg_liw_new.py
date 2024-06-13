""" import słownika SHG Liw - osady historyczne - wymaga wikibaseintegrator 0.12 lub nowszej """
# pylint: disable=logging-fstring-interpolation
import os
import sys
import time
import re
import logging
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import WikibaseIntegrator, wbi_login
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator.wbi_enums import ActionIfExists
from wikibaseintegrator.wbi_helpers import execute_sparql_query
from wikibaseintegrator.datatypes import MonolingualText, String, Item, Time, URL
from wikibaseintegrator.wbi_enums import WikibaseDatePrecision


# adresy dla API Wikibase (instancja docelowa)
# wbi_config['MEDIAWIKI_API_URL'] = 'https://wikihum.lab.dariah.pl/api.php'
# wbi_config['SPARQL_ENDPOINT_URL'] = 'https://wikihum.lab.dariah.pl/bigdata/sparql'
# wbi_config['WIKIBASE_URL'] = 'https://wikihum.lab.dariah.pl'

# login i hasło ze zmiennych środowiskowych (instancja docelowa)
# env_path = Path(".") / ".env_wikihum"
# load_dotenv(dotenv_path=env_path)

# adresy dla API wikibase (instancja testowa)
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych (instancja testowa)
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

WIKIBASE_WRITE = True # czy zapis do wikibase czy tylko test

# instancja produkcyjna
#P_STATED_AS = 'P54'
#P_AHP_ID = 'P81'
#P_PRNG = 'P76'
#Q_SHG = 'Q' <- uzupełnić
#P_STATED_IN = 'P55'
#P_VOLUME = 'P60'
#P_PAGES = 'P36'
#P_ADD_INFO = 'P155'
#P_STARTED_AT = 'P17'
#P_ENDS_AT = 'P32'
#P_LOC_ADM = 'P70'
#P_INFO_STATUS = 'P26'
#P_LOC_CHURCH = 'P143'
#P_OWNERSHIP = 'P134'
#P_POINT_IN_TIME = 'P40'
#P_DESCRIBED_URL = 'P13'
#P_REFERENCE_URL ='P2'
# Q_DISTRICT_LIWSKI = 'Q175732'
# Q_DISTRICT_KAM = 'Q175722'
# Q_DISTRICT_CZERSKI = ''
# Q_DYSTRYKT_WARSZ = ''
# Q_DYSTRYKT_CZER = ''
# Q_ZIEMIA_CZERSKA = ''
# Q_ZIEMIA_LIWSKA = ''
# Q_CIRCA = 'Q18'
# Q_NOBLE_PROP = 'Q66'
# Q_CHURCH_PROP = 'Q63'
# Q_MONARCH_PROP = 'Q64'
# Q_TOWN_PROP = 'Q65'

# instancja testowa
P_STATED_AS = 'P505'
P_AHP_ID = 'P403'
P_PRNG = 'P487'
Q_SHG = 'Q405481'
P_STATED_IN = 'P506'
P_VOLUME = 'P518'
P_PAGES = 'P479'
P_ADD_INFO = 'P538'
P_STARTED_AT = 'P432' # earliest
P_ENDS_AT = 'P464'    # latest
P_POINT_IN_TIME = 'P485'
P_LOC_ADM = 'P470'
P_INFO_STATUS = 'P458'
P_LOC_CHURCH = 'P537'
P_OWNERSHIP = 'P515'
P_DESCRIBED_URL = 'P424'
P_REFERENCE_URL = 'P399'

Q_DISTRICT_LIWSKI = 'Q397633'
Q_DISTRICT_KAM = 'Q397623'
Q_DISTRICT_CZERSKI = 'Q397611'
Q_DYSTRYKT_WARSZ = 'Q405484'
Q_DYSTRYKT_CZER = 'Q405485'
Q_ZIEMIA_CZERSKA = 'Q405483'
Q_ZIEMIA_LIWSKA = 'Q405482'
Q_CIRCA = 'Q233831'
Q_NOBLE_PROP = 'Q233989'
Q_CHURCH_PROP = 'Q233986'
Q_MONARCH_PROP = 'Q233987'
Q_TOWN_PROP = 'Q233988'


# parafie instancja testowa
parafie_wiki = {
    "Czerwonka": "Q398341",
    "Dobre": "Q398415",
    "Grębków": "Q398661",
    "Jadowo": "Q398756",
    "Kałuszyno": "Q398850",
    "Korytnica": "Q399014",
    "Liw": "Q399235",
    "Liw Stary": "Q399235",
    "Mińsko": "Q399453",
    "Niwiski": "Q399574",
    "Oleksin": "Q399631",
    "Pniewnik": "Q399815",
    "Wierzbno": "Q400642",
    "Wiśniew": "Q400682",
    "w Liwie": "Q399235",
    "Wodynie": "Q400705"
}

# tworzenie obiektu loggera
file_log = Path('..') / 'log' / 'shg_liw_new.log'
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
log_format = logging.Formatter('%(asctime)s - %(message)s')
c_handler = logging.StreamHandler()
c_handler.setFormatter(log_format)
c_handler.setLevel(logging.DEBUG)
logger.addHandler(c_handler)
# zapis logów do pliku tylko jeżeli uruchomiono z zapisem do wiki
if WIKIBASE_WRITE or 1:
    f_handler = logging.FileHandler(file_log)
    f_handler.setFormatter(log_format)
    f_handler.setLevel(logging.INFO)
    logger.addHandler(f_handler)


# -------------------------- FUNKCJE -------------------------------------------
def tekst_data(value: str):
    """ ekstracja nazwy i dat """
    patterns = [r'a\.\s+\d{1,2}\s+[XVI]{2,4}\s+\d{4}\s+kop.\s+[XVI]{2,4}\s+w\.', # a. 8 XI 1542 kop. XIX w. Wyglądówko
                r'\d{4}\!\s+\[recte:\s+\d{4}\]', # 1569! [recte: 1570] Koriczkawola
                r'\d{4}\s+kop\.\s+[XVI\/]{2,7}\s+w\.', # 1496 kop. XIX/XX w. Wierzbno
                r'\d{4}\s+kop\.\s+\d{4}\s+zachowana\s+w\s+odpisie\s+z\s+\d{1}\.\s+poł\.\s+[XVI\/]{2,7}\s+w\.', # 1447 kop. 1667 zachowana w odpisie z 2. poł. XIX w. Polkowo
                r'\d{4}\s+kopia\s+[XVI\/]{2,7}\s+w\.', # kopia XX w. Polyko!
                r'\d{4}\s+kop\.\s+z\s+[XVI\/]{2,7}\s+w\.', # 1496 kop. z XIX/XX w. Wierzbno
                r'[ok\s\.]{0,}\d{4}\s+or\.',
                r'\d{4}\s+kop\.\s+[\d\-]+',
                r'po\s+\d{4}', # po 1578 Mroczki-Klekty i Mroczki-Łopuchy
                r'\d{4}\s+kop\.\s+[XVI]{2,3}\s+w\.', # 1545 kop. XX w. Polomya Rakowe Lanky
                r'\[\d{4}\]\s+kop\.\s+\d{4}-\d{2,4}', # [1419] kop. 1456-1459 Zavadi
                r'\[\d{4}\]\s+kop\.\s+\d{4}', # [1419] kop. 1456 Zavadi
                r'\d{4}\s+kop\.\s+ok\.\s+\d{4}\s+', # 1453 kop. ok. 1776 Krypy alias Wyczolkowo
                r'\d{4}\s+kop\.\s+z\s+\d{1}\.\s+poł\.\s+[XVI]{2,4}\s+w\.\s+', # 1515 kop. z 2. poł. XVI w. Wolia Prossewska
                r'\[\d{4}\s+a\.\s+\d{1,2}\s+[XVI]{2,4}\]\s+kop\.\s+[XVI]{2,3}\s+w\.', # [1542 a. 8 XI] kop. XIX w. Wyglądały
                r'\d{4}\s+kop\.\s+z\s+\d{1}\s+poł\.\s+[XVI]{2,3}\s+w\.', # 1513 kop. z 2 poł. XVI w. Moscziska
                r'\d{4}\s+or\.\?', # 1434 or.?
                r'\d{4}\s+kop\.\s+z\s+\d{1}\s+poł\.\s+[XVI]{2,4}\s+w\.', # 1468 kop. z 1. poł. XX w. Polycowo
                r'\d{4}\s+kop\.\s+\d{1}\s+poł\.\s+[XVI]{2,4}\s+w\.', # 1511 kop. 2 poł. XVI w. Mschadle
                r'\d{4}\s+kop\.\s+z\s+końca\s+[XVI]{2,4}\s+w\.', # 1536 kop. z końca XVI w. Zołkowo
                r'\d{4}\s+kop\.\s+z\s+[XVI]{2,4}\s+w\.' # 1536 kop. z XVI w. Niedzwiedzkierz
                r'\d{4}\s+or\.\?\s+kop\.\?', # 1434 or.? kop.? Sucha;
                r'[XVI]{2,4}\s+w\.\s+', # XVII w. Popiołek
                r'\[\d{4}-\d{1,4}\]', # [1426-27] Boyno
                r'\d{4},\s*\d{4}', # 1435,1452 Cirwonka;
                r'\d{4}\s+kop\.\s+[XVI]{2,4}', # 1556 kop. XIX Ryciołek
                r'\[a\.\s+\d{4}\]', # [a. 1452] Polycowo
                r'\d{4}-\d{2,4}\s+',
                r'\[ok\.\s+\d{4}\]\s+',
                r'\[\d{4}\]\s+',
                r'\d{4}'
                ]

    space_pattern = re.compile(r'\u00A0') # zdarzają się niełamliwe spacje
    value = space_pattern.sub(' ', value)

    daty = ''
    for pattern in patterns:
        match = re.search(pattern=pattern, string=value)
        if match:
            daty = match.group()
            break

    nazwa = value.replace(daty, '')
    return daty.strip(), nazwa.strip(), value


def get_month(value:str)->str:
    """ month """
    months = {"I":"01", "II":"02", "III":"03", "IV":"04", "V":"05", "VI":"06",
              "VII":"07", "VIII":"08", "IX":"09", "X":"10", "XI":"11", "XII":"12"}
    return months[value]


def search_by_unique_id(prop_id: str, id_value: str) -> tuple:
    """ wyszukiwanie elementu na podstawie wartości deklaracji będącej jednoznacznym
        identyfikatorem, zwraca krotkę (True/False, qid) """
    query = f'SELECT ?item WHERE {{ ?item wdt:{prop_id} "{id_value}". }} LIMIT 5'

    results = execute_sparql_query(query)
    output = []
    for result in results["results"]["bindings"]:
        output.append(result["item"]["value"])

    # wynik to lista adresów np. https://wikihum.lab.dariah.pl/entity/Q77881
    # lub https://prunus-208.man.poznan.pl/entity/Q277021 dla testowej
    if len(output) == 1:
        output_string = output[0].strip()
        if 'https' in output_string:
            if 'wikihum' in output_string:
                search_result = output_string.replace('https://wikihum.lab.dariah.pl/entity/', '')
            else:
                search_result = output_string.replace('https://prunus-208.man.poznan.pl/entity/', '')
        else:
            if 'wikihum' in output_string:
                search_result = output_string.replace('http://wikihum.lab.dariah.pl/entity/', '')
            else:
                search_result = output_string.replace('http://prunus-208.man.poznan.pl/entity/', '')
        return True, search_result

    return False, f'ERROR: brak lub niejednoznaczny wynik wyszukiwania (znaleziono: {len(output)}).'


# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':
    # pomiar czasu wykonania
    start_time = time.time()

    # logowanie do instancji wikibase
    login_instance = wbi_login.OAuth1(consumer_token=WIKIDARIAH_CONSUMER_TOKEN,
                                     consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                     access_token=WIKIDARIAH_ACCESS_TOKEN,
                                     access_secret=WIKIDARIAH_ACCESS_SECRET,
                                     token_renew_period=14400)

    # plik xlsx z listą miejscowości
    xlsx_input = Path('..') / 'data' / 'shg_liw.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["Miejscowości"]

    #  nazwy kolumn w arkuszu
    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    # SHG_ID, Nazwa, Własność, Strony, Odmianki, Powiaty, Powiaty_Daty, Parafie, Parafie_Daty, AHP, AHP Parafia,
    # Uwagi lub przyczyna odrzucenia, Uwagi AB

    # przetwarzanie kolejnych wierszy arkusza (w pierwszym są nazwy kolumn)
    max_row = ws.max_row
    start_row = 33
    max_row = 33
    licznik = start_row - 1
    for row in ws.iter_rows(start_row, max_row):
        # wczytanie danych z xlsx
        licznik += 1
        shg_id = row[col_names['SHG_ID']].value
        shg_id_http = ws.cell(row = licznik, column=1).hyperlink.target
        nazwa = str(row[col_names['Nazwa']].value).strip()
        odmianki = row[col_names['Odmianki']].value
        powiaty = str(row[col_names['Powiaty']].value).strip()
        powiaty_daty = str(row[col_names['Powiaty_Daty']].value).strip()
        parafie = str(row[col_names['Parafie']].value).strip()
        parafie_daty = str(row[col_names['Parafie_Daty']].value).strip()
        ahp_ok = str(row[col_names['AHP_OK']].value).strip()
        strona = str(row[col_names['Strona']].value).strip()
        wlasnosc = str(row[col_names['Własność']].value).strip()

        # pomijanie wierszy jeżeli nie ma ustaonego AHP ID
        if shg_id is None or ahp_ok is None or ahp_ok == '' or ahp_ok == 'None' or ahp_ok not in ('NEW'):
            # logger.info(f'Pominięto: shg = {shg_id} ahp = {ahp_ok}')
            continue

        wbi = WikibaseIntegrator(login=login_instance)
        wb_item = wbi.item.new()

        nazwa = nazwa.capitalize()
        wb_item.labels.set('pl', nazwa)
        wb_item.labels.set('en', nazwa)
        description_pl = f'osada historyczna (powiat {powiaty}, parafia: {parafie})'
        description_en = f'historical settlement (district {powiaty}, parish: {parafie})'
        wb_item.descriptions.set('pl', description_pl)
        wb_item.descriptions.set('en', description_en)

        qid = ''
        if WIKIBASE_WRITE:
            item = wb_item.write()
            qid = item.id

        print(f'{nazwa} QID: https://prunus-208.man.poznan.pl/wiki/Item:{qid}')

        # referencje
        references = [[Item(value=Q_SHG, prop_nr=P_STATED_IN),
            String(value=strona, prop_nr=P_PAGES),
            URL(value=shg_id_http, prop_nr=P_REFERENCE_URL)
        ]]

        # ODMIANKI
        lista_odmianek_dat = []
        if odmianki is not None:
            lista = odmianki.split(';')
            lista = [tekst_data(x.strip()) for x in lista]
            for year, name, odmianka in lista:
                if '[' in year:
                    year = year.replace('[', '')
                if ']' in year:
                    year = year.replace(']', '')

                if ',' in year:
                    tmp = year.split(',')
                    for t_year in tmp:
                        lista_odmianek_dat.append((t_year, name))
                elif '-' in year:
                    tmp = year.split('-')
                    tmp[1] = tmp[0][:2] + tmp[1]
                    for t_year in tmp:
                        lista_odmianek_dat.append((t_year, name))
                elif ',' in name:
                    tmp = name.split(',')
                    for t_name in tmp:
                        lista_odmianek_dat.append((year, t_name))
                else:
                    lista_odmianek_dat.append((year, name))

            for year, name in lista_odmianek_dat:
                add_info = ''
                status_info = ''
                property_year = P_POINT_IN_TIME
                if 'ok.' in year:
                    status_info = Q_CIRCA
                    year = year.replace('ok.', '').strip()
                elif year.startswith('po '):
                    property_year = P_STARTED_AT
                    year = year[3:]
                elif year.startswith('a. '):
                    property_year = P_ENDS_AT
                    year = year[3:]
                elif 'recte' in year:
                    pos = year.find('recte')
                    year = year[:pos].strip()

                if '!' in year:
                    year = year.replace('!', '').strip()

                pt = r'\d{1}\s+[XVI]{1,5}\s+\d{4}' # 8 XI 1542 kop. XIX w. Wyglądówko
                pt_match = re.search(pattern=pt, string=year)
                if pt_match:
                    pt_year = pt_match.group().strip()
                    tmp = pt_year.split(' ')
                    day = tmp[0].strip().zfill(2)
                    month = get_month(tmp[1].strip())
                    year = tmp[2].strip()
                    year = f'{year}-{month}-{day}'

                if 'or.' in year:
                    add_info = 'nazwa pochodzi z oryginalnego dokumentu'
                    year = year.replace('or.', '').strip()
                elif 'kop.' in year:
                    pos = year.find('kop.')
                    add_info = year[pos:]
                    add_info = add_info.replace('kop.', 'kopia')
                    year = year[:pos].strip()
                elif 'kopia ' in year:
                    pos = year.find('kopia ')
                    add_info = year[pos:]
                    year = year[:pos].strip()
                else:
                    year = year.strip()

                # kwalifikatory
                if year:
                    if len(year) == 4:
                        format_time = f'+{year}-00-00T00:00:00Z'
                    elif len(year) == 10:
                        format_time = f'+{year}T00:00:00Z'

                    qualifier = [Time(prop_nr=property_year,
                                      time=format_time,
                                      precision=WikibaseDatePrecision.YEAR)]
                else:
                    qualifier = [Time(prop_nr=P_ENDS_AT,
                                      time='+1600-00-00T00:00:00Z',
                                      precision=WikibaseDatePrecision.CENTURY)]

                if add_info:
                    qualifier.append(String(value=add_info, prop_nr=P_ADD_INFO))
                if status_info:
                    qualifier.append(Item(value=status_info, prop_nr=P_INFO_STATUS))

                name_claim = MonolingualText(text=name,
                                             language='pl',
                                             prop_nr=P_STATED_AS,
                                             references=references,
                                             qualifiers=qualifier)
                wb_item.aliases.set('pl', name)
                wb_item.claims.add(name_claim, action_if_exists=ActionIfExists.APPEND_OR_REPLACE)

            # zapis
            if WIKIBASE_WRITE:
                wb_item.write()

        # POWIATY
        if powiaty_daty is not None:
            pattern = r'\d{4}(-\d{2})?(\s+n\.)?'
            tmp = powiaty_daty.split(';')
            for t in tmp:
                p = re.compile(r'\u00A0') # zdarzają się niełamliwe spacje
                t = p.sub(' ', t)
                name = year = ''
                match = re.search(pattern=pattern, string=t)
                if match:
                    year = match.group().strip()
                    name = t.replace(year, '').strip()
                    if year.endswith('n.'):
                        year = year.replace('n.', '').strip()
                        prop = P_STARTED_AT
                    else:
                        prop = P_POINT_IN_TIME
                    format_year = f'+{year}-00-00T00:00:00Z'
                    precision=WikibaseDatePrecision.YEAR
                else:
                    name = t.strip()
                    prop = P_ENDS_AT
                    format_year = '+1600-00-00T00:00:00Z'
                    precision=WikibaseDatePrecision.CENTURY

                add_info = ''

                if 'pow. liwski' in name:
                    powiat = Q_DISTRICT_LIWSKI
                elif 'pow. liw.' in name:
                    powiat = Q_DISTRICT_LIWSKI
                elif 'pow. kamieniecki' in name:
                    powiat = Q_DISTRICT_KAM
                elif 'pow. czerski' in name:
                    powiat = Q_DISTRICT_CZERSKI
                elif 'dystr. czer' in name:
                    powiat = Q_DYSTRYKT_CZER
                elif 'dystr. warsz.' in name:
                    powiat = Q_DYSTRYKT_WARSZ
                elif 'z. czer.' in name:
                    powiat = Q_ZIEMIA_CZERSKA
                elif 'z. liw.' in name:
                    powiat = Q_ZIEMIA_LIWSKA
                else:
                    powiat = ''
                    add_info = name

                qualifier = [Time(prop_nr=prop,
                                  time=format_year,
                                  precision=precision)]
                if add_info:
                    qualifier.append(String(value=add_info, prop_nr=P_ADD_INFO))
                if '?' in name:
                    qualifier.append(Item(value=Q_CIRCA, prop_nr=P_INFO_STATUS))

                powiat_claim = Item(value=powiat,
                                    prop_nr=P_LOC_ADM,
                                    references=references,
                                    qualifiers=qualifier)
                wb_item.claims.add(powiat_claim, action_if_exists=ActionIfExists.APPEND_OR_REPLACE)

            # zapis
            if WIKIBASE_WRITE:
                wb_item.write()

        # PARAFIE
        if parafie_daty is not None:
            pattern = r'\[?\d{4}\]?(\s+[na]{1}\.)?'
            tmp = parafie_daty.split(';')
            for t in tmp:
                p = re.compile(r'\u00A0') # zdarzają się niełamliwe spacje
                t = p.sub(' ', t)
                name = year = ''
                match = re.search(pattern=pattern, string=t)
                if match:
                    year = match.group().strip()
                    name = t.replace(year, '').strip()
                    if year.endswith('n.'):
                        year = year.replace('n.', '').strip()
                        prop = P_STARTED_AT
                    elif year.endswith('a.'):
                        year = year.replace('a.', '').strip()
                        prop = P_ENDS_AT
                    else:
                        prop = P_POINT_IN_TIME

                    if '[' in year:
                        year = year.replace('[','')
                    if ']' in year:
                        year = year.replace(']','')

                    format_year = f'+{year}-00-00T00:00:00Z'
                    precision=WikibaseDatePrecision.YEAR
                else:
                    name = t.strip()
                    prop = P_ENDS_AT
                    format_year = '+1600-00-00T00:00:00Z'
                    precision=WikibaseDatePrecision.CENTURY

                qualifier = [Time(prop_nr=prop,
                                  time=format_year,
                                  precision=precision)]
                if '?' in name:
                    qualifier.append(Item(value=Q_CIRCA, prop_nr=P_INFO_STATUS))

                q_parafia = ''
                for key, value in parafie_wiki.items():
                    if key in name:
                        q_parafia = value
                        break

                if not q_parafia:
                    print(f'ERR: brak parafii {name}')
                    sys.exit(1)

                #print(f'Parafia: {name} {q_parafia} year: {format_year} prop: {prop}')

                parafia_claim = Item(value=q_parafia,
                                    prop_nr=P_LOC_CHURCH,
                                    references=references,
                                    qualifiers=qualifier)
                wb_item.claims.add(parafia_claim, action_if_exists=ActionIfExists.APPEND_OR_REPLACE)

            # zapis
            if WIKIBASE_WRITE:
                wb_item.write()

        # WŁASNOŚĆ
        if wlasnosc is not None:
            tmp = wlasnosc.split(',')
            for t in tmp:
                t = t.strip()
                add_info = ''
                if 'częściowo' in t:
                    add_info = 'częściowo'
                    t = t.replace('częściowo', '').strip()

                if t == 'szlachecka':
                    wl = Q_NOBLE_PROP
                elif t == 'duchowna':
                    wl = Q_CHURCH_PROP
                elif t == 'monarsza':
                    wl = Q_MONARCH_PROP

                prop = P_ENDS_AT
                format_year = '+1600-00-00T00:00:00Z'
                precision=WikibaseDatePrecision.CENTURY

                qualifier = [Time(prop_nr=prop,
                                  time=format_year,
                                  precision=precision)]
                if add_info:
                    qualifier.append(String(value=add_info, prop_nr=P_ADD_INFO))

                wlasnosc_claim = Item(value=wl,
                                    prop_nr=P_OWNERSHIP,
                                    references=references,
                                    qualifiers=qualifier)
                wb_item.claims.add(wlasnosc_claim, action_if_exists=ActionIfExists.APPEND_OR_REPLACE)

            # zapis
            if WIKIBASE_WRITE:
                wb_item.write()

        # DESCRIBED BY URL
        if shg_id_http is not None:
            url_claim = URL(value=shg_id_http, prop_nr=P_DESCRIBED_URL)
            wb_item.claims.add(url_claim, action_if_exists=ActionIfExists.APPEND_OR_REPLACE)

            # zapis
            if WIKIBASE_WRITE:
                wb_item.write()


    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

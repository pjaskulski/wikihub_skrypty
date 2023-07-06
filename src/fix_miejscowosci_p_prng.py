""" uzupełnienie danych PRNG/SIMC miejscowosci z pliku miejscowosciP_QID.xlsx"""
import os
import sys
import time
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import get_properties
from property_import import create_statement_data


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

WIKIBASE_WRITE = True

# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':
    # standardowe właściwości
    print('Przygotowanie właściwości...')
    properties = get_properties(['SIMC place ID', 'prng id', 'reference URL', 'retrieved',
                                'part of', 'has part or parts',
                                ])

    # wspólna referencja dla wszystkich deklaracji z PRG
    references = {}
    references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGiK/PRNG/WFS/GeographicalNames'
    references[properties['retrieved']] = '2022-09-23'

    print('Logowanie do wikibase...')
    # logowanie do instancji wikibase
    login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                     consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                     access_token=WIKIDARIAH_ACCESS_TOKEN,
                                     access_secret=WIKIDARIAH_ACCESS_SECRET,
                                     token_renew_period=14400)

    print('Wczytanie pliku xlsx...')
    xlsx_input = '../data_prng/miejscowosciP_QID.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["miejscowosciP"]

    print('Wczytanie nazw kolumn xlsx...')
    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    print('Przetwarzanie wierszy pliku xlsx...')
    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1
        if index < 8499:
            continue
        # wczytanie danych z xlsx
        row_qid = row[col_names['QID']].value
        row_prng = row[col_names['IDENTYFIKA']].value
        #row_simc = row[col_names['IDENTYFI_1']].value
        if not row_qid:
            continue

        # przygotowanie struktur wikibase
        data = []

        if row_prng:
            statement = create_statement_data(properties['prng id'], row_prng,
                None, None, add_ref_dict=references)
            if statement:
                data.append(statement)

        # if row_simc:
        #     statement = create_statement_data(properties['SIMC place ID'], row_simc,
        #         None, None, add_ref_dict=references)
        #     if statement:
        #         data.append(statement)

        # jeżeli nie ma nic do uzupełnienia
        if not data:
            continue

        if WIKIBASE_WRITE:
            test = 1
            while True:
                try:
                    wb_item = wbi_core.ItemEngine(item_id=row_qid, data=data)
                    wb_item.write(login_instance, bot_account=True, entity_type='item')
                    print(f"{index}/{max_row - 1} Uzupełniono właściwość PRNG")
                    break
                except MWApiError as wbdelreference_error:
                    err_code = wbdelreference_error.error_msg['error']['code']
                    message = wbdelreference_error.error_msg['error']['info']
                    print(f'ERROR: {err_code}, {message}')
                    if err_code in ['assertuserfailed', 'badtoken']:
                        if test == 1:
                            print('Generate edit credentials...')
                            login_instance.generate_edit_credentials()
                            test += 1
                            continue
                    sys.exit(1)
        else:
            print(f"{index}/{max_row - 1} Przygotowano uzupełnienie właściwości PRNG.")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

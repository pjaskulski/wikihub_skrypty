""" import gmin z pliku gminy.xlsx z danymi z PRG"""
import os
import time
import sys
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from wikidariahtools import element_search_adv, search_by_purl, get_properties, get_elements
from wikidariahtools import read_qid_from_text
from property_import import create_statement_data, has_statement


# adresy wikibase
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# login i hasło ze zmiennych środowiskowych
env_path = Path(".") / ".env"
load_dotenv(dotenv_path=env_path)

# OAuth
WIKIDARIAH_CONSUMER_TOKEN = os.environ.get('WIKIDARIAH_CONSUMER_TOKEN')
WIKIDARIAH_CONSUMER_SECRET = os.environ.get('WIKIDARIAH_CONSUMER_SECRET')
WIKIDARIAH_ACCESS_TOKEN = os.environ.get('WIKIDARIAH_ACCESS_TOKEN')
WIKIDARIAH_ACCESS_SECRET = os.environ.get('WIKIDARIAH_ACCESS_SECRET')

# pomiar czasu wykonania
start_time = time.time()

WIKIBASE_WRITE = True

# standardowe właściwości
properties = get_properties(['instance of', 'stated as', 'reference URL', 'retrieved',
                            'id SDI', 'part of', 'has part or parts', 'TERYT', 'commune type',
                            'stated in', 'point in time'])

# elementy definicyjne
elements = get_elements(['administrative unit',
                         'onto.kul.pl/ontohgis/administrative_type_45',
                         'urban commune',
                         'rural commune',
                         'urban-rural commune',
                         'city in urban-rural commune',
                         'rural area in urban-rural commune',
                         'Warsaw district',
                         'delegatury w miastach: Kraków, Łódź, Poznań i Wrocław'
                         ])

# typy gmin
typ_gminy_element = {}
typ_gminy_text = {}
typ_gminy_text_en = {}
# 1 – gmina miejska
typ_gminy_element['1'] = elements['urban commune']
typ_gminy_text['1'] = 'gmina miejska'
typ_gminy_text_en['1'] = 'urban commune'
# 2 – gmina wiejska
typ_gminy_element['2'] = elements['rural commune']
typ_gminy_text['2'] = 'gmina wiejska'
typ_gminy_text_en['2'] = 'rural commune'
# 3 – gmina miejsko-wiejska
typ_gminy_element['3'] = elements['urban-rural commune']
typ_gminy_text['3'] ='gmina miejsko-wiejska'
typ_gminy_text_en['3'] ='urban-rural commune'
# 4 – miasto w gminie miejsko-wiejskiej
typ_gminy_element['4'] = elements['city in urban-rural commune']
typ_gminy_text['4'] = 'miasto w gminie miejsko-wiejskiej'
typ_gminy_text_en['4'] = 'city in urban-rural commune'
# 5 – obszar wiejski w gminie miejsko-wiejskiej
typ_gminy_element['5'] = elements['rural area in urban-rural commune']
typ_gminy_text['5'] = 'obszar wiejski w gminie miejsko-wiejskiej'
typ_gminy_text_en['5'] = 'rural area in urban-rural commune'
# 8 – dzielnice m. st. Warszawy
typ_gminy_element['8'] = elements['Warsaw district']
typ_gminy_text['8'] = 'dzielnice m. st. Warszawy'
typ_gminy_text_en['8'] = 'Warsaw district'
# 9 – delegatury w miastach: Kraków, Łódź, Poznań i Wrocław
typ_gminy_element['9'] = elements['delegatury w miastach: Kraków, Łódź, Poznań i Wrocław']
typ_gminy_text['9'] = 'delegatury w miastach: Kraków, Łódź, Poznań i Wrocław'
typ_gminy_text_en['9'] = 'delegatury w miastach: Kraków, Łódź, Poznań i Wrocław'


# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':

    # wspólna referencja dla wszystkich deklaracji z PRG
    references = {}
    references[properties['reference URL']] = 'https://mapy.geoportal.gov.pl/wss/service/PZGIK/PRG/WFS/AdministrativeBoundaries'
    references[properties['retrieved']] = '2022-09-05'

    # kwalifikator z punktem czasowym
    qualifiers = {}
    qualifiers[properties['point in time']] = '+2022-00-00T00:00:00Z/9' # rok 2022

    # wspólna referencja dla wszystkich deklaracji z ontohgis
    onto_references = {}
    onto_references[properties['stated in']] = 'Q233549'

    # logowanie do instancji wikibase
    login_instance = wbi_login.Login(consumer_key=WIKIDARIAH_CONSUMER_TOKEN,
                                     consumer_secret=WIKIDARIAH_CONSUMER_SECRET,
                                     access_token=WIKIDARIAH_ACCESS_TOKEN,
                                     access_secret=WIKIDARIAH_ACCESS_SECRET,
                                     token_renew_period=14400)

    xlsx_input = '../data_prng/gminy.xlsx'
    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["gminy"]

    # nazwy kolumn w xlsx
    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    parts = {}
    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1
        # wczytanie danych z xlsx
        nazwa = row[col_names['JPT_NAZWA_']].value
        if not nazwa:
            continue

        label_pl = 'gmina ' + nazwa
        label_en = 'commune ' + nazwa

        teryt = row[col_names['JPT_KOD_JE']].value
        idiip = row[col_names['IIP_IDENTY']].value
        part_of = teryt[:4]
        typ_gm = teryt[-1]

        description_pl = 'gmina - współczesna jednostka administracyjna według Państwowego Rejestru Granic (PRG)'
        description_en = 'commune - a modern administrative unit according to the National Register of Boundaries (PRG)'

        # przygotowanie struktur wikibase
        data = []
        aliasy = []

        # instance of
        q_gmina = elements['onto.kul.pl/ontohgis/administrative_type_45']
        statement = create_statement_data(prop=properties['instance of'],
                                          value=q_gmina,
                                          reference_dict=None,
                                          qualifier_dict=None,
                                          add_ref_dict=onto_references)
        if statement:
            data.append(statement)

        # id SDI
        if idiip:
            statement = create_statement_data(prop=properties['id SDI'],
                                              value=idiip,
                                              reference_dict=None,
                                              qualifier_dict=qualifiers,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)

        # TERYT
        if teryt:
            statement = create_statement_data(prop=properties['TERYT'],
                                              value=teryt,
                                              reference_dict=None,
                                              qualifier_dict=qualifiers,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)

        # JPT_NAZWA_
        statement = create_statement_data(prop=properties['stated as'],
                                          value=f'pl:"{nazwa}"',
                                          reference_dict=None,
                                          qualifier_dict=qualifiers,
                                          add_ref_dict=references)
        if statement:
            data.append(statement)

        # typ gminy
        if typ_gm:
            statement = create_statement_data(prop=properties['commune type'],
                                              value=typ_gminy_element[typ_gm],
                                              reference_dict=None,
                                              qualifier_dict=qualifiers,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)

        # part of i has part or parts
        ok, pow_qid = search_by_purl(properties['TERYT'], part_of)
        if ok:
            statement = create_statement_data(prop=properties['part of'],
                                              value=pow_qid,
                                              reference_dict=None,
                                              qualifier_dict=qualifiers,
                                              add_ref_dict=references)
            if statement:
                data.append(statement)
        else:
            pow_qid = ''

        # etykiety, description, aliasy
        wb_item = wbi_core.ItemEngine(new_item=True, data=data)
        wb_item.set_label(label_en, lang='en')
        wb_item.set_label(label_pl,lang='pl')

        # description
        if pow_qid:
            wb_pow = wbi_core.ItemEngine(item_id=pow_qid)
            pow_label_pl = wb_pow.get_label('pl')
            pow_label_en = wb_pow.get_label('en')
        else:
            pow_label_pl = ''
            pow_label_en = ''
            print(f'ERROR: brak powiatu dla gminy: {label_pl}.')

        wb_item.set_description(f"{description_en} ({pow_label_en}, {typ_gminy_text_en[typ_gm]})", 'en')
        wb_item.set_description(f"{description_pl} ({pow_label_pl}, {typ_gminy_text[typ_gm]})", 'pl')

        if aliasy:
            for value_alias in aliasy:
                wb_item.set_aliases(value_alias, 'pl')

        if WIKIBASE_WRITE:
            test = 1
            while True:
                try:
                    new_id = wb_item.write(login_instance, bot_account=True, entity_type='item')
                    if new_id:
                        print(f'{index}/{max_row - 1} # [https://prunus-208.man.poznan.pl/wiki/Item:{new_id} {label_en}/{label_pl}')
                        if pow_qid:
                            if pow_qid in parts:
                                parts[pow_qid].append(new_id)
                            else:
                                parts[pow_qid] = [new_id]
                    break
                except MWApiError as wb_error:
                    err_code = wb_error.error_msg['error']['code']
                    message = wb_error.error_msg['error']['info']
                    if 'already has label' in message and err_code == 'modification-failed':
                        match_qid = read_qid_from_text(message)
                        print(f'{index}/{max_row - 1} Element: {label_en} / {label_pl} już istnieje {match_qid}.')
                        break
                    else:
                        # jeżeli jest to problem z tokenem to próba odświeżenia tokena i powtórzenie
                        # zapisu, ale tylko raz, w razie powtórnego błędu bad token, skrypt kończy pracę
                        if err_code in ['assertuserfailed', 'badtoken']:
                            if test == 1:
                                print('Generate edit credentials...')
                                login_instance.generate_edit_credentials()
                                test += 1
                                continue
                        print(err_code, message)
                        sys.exit(1)
        else:
            # wyszukiwanie po etykiecie, właściwości instance of oraz po opisie
            parameters = [(properties['instance of'], q_gmina)]
            ok, item_id = element_search_adv(label_en, 'en', parameters, f"{description_en} ({pow_label_en}, {typ_gminy_text_en[typ_gm]})")
            if not ok:
                new_id = 'TEST'
                print(f"{index}/{max_row - 1} Przygotowano dodanie elementu - {label_en} / {label_pl}  = {new_id}")
                if pow_qid:
                    if pow_qid in parts:
                        parts[pow_qid].append(new_id)
                    else:
                        parts[pow_qid] = [new_id]
            else:
                print(f'{index}/{max_row - 1} Element: {label_en} / {label_pl} już istnieje: {item_id}')

    # uzupełnienie właściwości 'has part or parts' dla powiatów
    if parts:
        print("\nUzupełnienie 'has part or parts' dla powiatów.\n")

        for item_qid, part_qids in parts.items():
            tmp_data = []
            data = []
            for part_qid in part_qids:
                if not has_statement(item_qid, properties['has part or parts'], part_qid):
                    if WIKIBASE_WRITE:
                        statement = create_statement_data(prop=properties['has part or parts'],
                                                          value=part_qid,
                                                          reference_dict=references,
                                                          qualifier_dict=qualifiers,
                                                          add_ref_dict=None,
                                                          if_exists='APPEND')
                        if statement:
                            data.append(statement)
                    else:
                        tmp_data.append(part_qid)

            if WIKIBASE_WRITE:
                if data:
                    wd_item = wbi_core.ItemEngine(item_id=item_qid, data=data, debug=False)
                    wd_item.write(login_instance, entity_type='item')
                    print(f"Do elementu {item_qid} dodano właściwości {properties['has part or parts']}.")
                else:
                    print(f"ERROR: {item_qid}, wystąpił problem z przygotowaniem danych dla właściwości 'has part or parts'")
            else:
                if tmp_data:
                    print(f"Przygotowano dodanie właściwości {properties['has part or parts']}: {','.join(tmp_data)}")
                else:
                    print(f"ERROR: {item_qid}, wystąpił problem z przygotowaniem danych dla właściwości 'has part or parts'")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')

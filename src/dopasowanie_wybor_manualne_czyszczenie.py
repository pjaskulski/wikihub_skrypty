""" wybór z wyników """

from pathlib import Path
import openpyxl


lista_path = Path('..') / 'data' / '7x.txt'
with open(lista_path, 'r', encoding='utf-8') as f:
    lista = f.readlines()
    lista = [x.strip() for x in lista]

input_path = Path('..') / 'data' / 'dopasowanie_prng_v7y.xlsx'
output_path = Path('..') / 'data' / 'dopasowanie_prng_v7yc.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['wynik_pandas_v2']

col_names = {}
nr_col = 0
for column in ws.iter_cols(1, ws.max_column):
    col_names[column[1].value] = nr_col
    nr_col += 1

rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row))

for row in reversed(rows):
    id_ahp = row[col_names['id_ahp']].value
    if id_ahp in lista:
        ws.delete_rows(row[0].row, 1)

wb.save(output_path)

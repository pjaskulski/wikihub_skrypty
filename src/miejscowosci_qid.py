""" Uzupełnianie danych miejscowosci w pliku miejscowosciU.xlsx (dane z PRG) o QID z wikibase """
import openpyxl


# ----------------------------------- MAIN -------------------------------------

if __name__ == '__main__':
    xlsx_input = '../data_prng/miejscowosciU.xlsx'
    xlsx_output = '../data_prng/miejscowosciU_QID.xlsx'
    qid_input = '../data_prng/lista_miejscowosci_qid.txt'

    wb = openpyxl.load_workbook(xlsx_input)
    ws = wb["miejscowosciU"]

    lista_qid = {}
    # plik z QID
    with open(qid_input, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    for line in lines:
        tmp_line = line.split(';')
        line_index = int(tmp_line[0].strip())
        line_qid= tmp_line[2].strip()
        lista_qid[line_index] = line_qid

    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    index = 0
    max_row = ws.max_row
    for row in ws.iter_rows(2, max_row):
        index += 1

        print(f"Przetwarzanie {index}/{max_row}.")

        # uzupełnienie danych (QID z linkiem)
        row[col_names['QID']].value = lista_qid[index]
        row[col_names['QID']].hyperlink = f"https://prunus-208.man.poznan.pl/wiki/Item:{lista_qid[index]}"
        row[col_names['QID']].style = "Hyperlink"

    wb.save(xlsx_output)
